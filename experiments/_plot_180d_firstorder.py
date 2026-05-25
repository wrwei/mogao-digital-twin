"""First-order saturation variant: ΔE*(t) = ΔE*_max · (1 - exp(-k·t)).

Per spot: fit linear slope on day 0-34 measured data, then back out k = slope / ΔE*_max
using pigment-specific ΔE*_max (literature-informed). Extend in L,a,b space by preserving
the per-spot direction-of-fade. Noise σ matched to the linear-fit residuals.
"""
import openpyxl, math, re
import numpy as np
import matplotlib.pyplot as plt

np.random.seed(42)

DATE_RE = re.compile(r"^\s*\d+\.\d+\s*$")
def to_f(v):
    try: return float(v)
    except Exception: return float("nan")
def is_date(v):
    if v is None: return False
    return bool(DATE_RE.match(str(v).strip()))
def date_to_day(d):
    s = str(d).strip(); m, day = s.split(".")
    dpm = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    return (sum(dpm[: int(m) - 1]) + int(day)) - (sum(dpm[:3]) + 17)

spec_cols = {
    "R":  [2, 5, 8],   "G":  [11, 14, 17], "B":  [20, 23, 26],
    "TR": [29, 32],    "TG": [35, 38],     "TB": [41, 44],
}
PIG = {"R": "Vermilion", "G": "Malachite", "B": "Azurite",
       "TR": "Vermilion", "TG": "Malachite", "TB": "Azurite"}

# Literature-informed asymptotic ΔE* (pigment fully degraded → substrate-like grey).
# Vermilion → meta-cinnabar (β-HgS, near-black): largest range; bright red has |a*|~40 and L*~45.
# Malachite → tenorite / basic copper black: bright green |a*|~30 and L*~60.
# Azurite → paratacamite / blue-green-grey: bright blue |b*|~33 and L*~40.
DELTAE_MAX = {"R": 60, "G": 50, "B": 40, "TR": 60, "TG": 50, "TB": 40}

SRC_OUT = {
    ("室温", "R",  0, "5.12"),
    ("恒温", "TR", 0, "5.14"),
}

src = "experiments/敦煌变化数据(1).xlsx"
wb = openpyxl.load_workbook(src, data_only=True)

q = 0.8; RH_t = 80; RH_ref = 40
rh_ratio = (RH_t / RH_ref) ** q

extra_days = list(range(42, 181, 14))
if extra_days[-1] != 180:
    extra_days.append(180)

# --- Compute first-order trajectory per spot per sheet ---
arms = {}
arm_labels = {
    "室温":     "23 C / 40%RH, ambient light  (measured + projected)",
    "恒温":     "40 C / 10%RH, dark  (measured + projected)",
    "恒温恒湿": "40 C / 80%RH  (RH-scaled prediction)",
}

# First build extended-trajectory in L,a,b per spot for 室温 and 恒温
def build_spot_traj(sheet):
    ws = wb[sheet]
    rows = [r for r in range(4, ws.max_row + 1) if is_date(ws.cell(r, 1).value)]
    meas_days = np.array([date_to_day(ws.cell(r, 1).value) for r in rows])
    date_strs = [str(ws.cell(r, 1).value).strip() for r in rows]
    all_days = np.array(list(meas_days) + extra_days)
    out = {"_t": all_days, "_n_meas": len(rows)}
    for grp, cols in spec_cols.items():
        dE_max = DELTAE_MAX[grp]
        for sidx, c in enumerate(cols):
            Ls, As, Bs = [], [], []
            for i_t, r in enumerate(rows):
                if (sheet, grp, sidx, date_strs[i_t]) in SRC_OUT:
                    Ls.append(np.nan); As.append(np.nan); Bs.append(np.nan); continue
                Ls.append(to_f(ws.cell(r, c).value))
                As.append(to_f(ws.cell(r, c + 1).value))
                Bs.append(to_f(ws.cell(r, c + 2).value))
            Ls = np.array(Ls); As = np.array(As); Bs = np.array(Bs)
            mask = ~np.isnan(Ls) & ~np.isnan(As) & ~np.isnan(Bs)
            t_fit = meas_days[mask]
            # Linear fit per channel — gives slope vector (= direction-of-fade × rate)
            if mask.sum() >= 2:
                sL, iL = np.polyfit(t_fit, Ls[mask], 1)
                sa, ia = np.polyfit(t_fit, As[mask], 1)
                sb, ib = np.polyfit(t_fit, Bs[mask], 1)
            else:
                sL = sa = sb = 0.0
                iL = np.nanmean(Ls); ia = np.nanmean(As); ib = np.nanmean(Bs)
            # Direction of fade and rate magnitude (in ΔE* / day)
            slope_vec = np.array([sL, sa, sb])
            slope_mag = float(np.linalg.norm(slope_vec))
            if slope_mag > 1e-9:
                u_hat = slope_vec / slope_mag
            else:
                u_hat = np.zeros(3)
            # First-order rate constant from initial slope: slope ≈ ΔE_max · k  =>  k = slope / ΔE_max
            k = slope_mag / dE_max
            # Baseline = fit at t=0
            L0 = iL; a0 = ia; b0 = ib
            # Residual sigmas for noise injection
            sigma_L = float(np.std(Ls[mask] - (sL * t_fit + iL), ddof=1)) if mask.sum() >= 2 else 0.0
            sigma_a = float(np.std(As[mask] - (sa * t_fit + ia), ddof=1)) if mask.sum() >= 2 else 0.0
            sigma_b = float(np.std(Bs[mask] - (sb * t_fit + ib), ddof=1)) if mask.sum() >= 2 else 0.0

            L_traj, a_traj, b_traj = [], [], []
            for i_t, t in enumerate(all_days):
                if i_t < len(rows) and not (sheet, grp, sidx, date_strs[i_t]) in SRC_OUT:
                    # measured rows: keep raw values (could be NaN for outliers)
                    if mask[i_t]:
                        L_traj.append(Ls[i_t]); a_traj.append(As[i_t]); b_traj.append(Bs[i_t])
                    else:
                        L_traj.append(np.nan); a_traj.append(np.nan); b_traj.append(np.nan)
                elif i_t < len(rows):
                    # outlier — mask
                    L_traj.append(np.nan); a_traj.append(np.nan); b_traj.append(np.nan)
                else:
                    # extrapolated — first-order model
                    dE_t = dE_max * (1.0 - np.exp(-k * t))
                    L_pred = L0 + dE_t * u_hat[0] + np.random.normal(0, sigma_L)
                    a_pred = a0 + dE_t * u_hat[1] + np.random.normal(0, sigma_a)
                    b_pred = b0 + dE_t * u_hat[2] + np.random.normal(0, sigma_b)
                    L_traj.append(L_pred); a_traj.append(a_pred); b_traj.append(b_pred)
            out[(grp, sidx, c)] = {
                "L": np.array(L_traj), "a": np.array(a_traj), "b": np.array(b_traj),
                "L0": L0, "a0": a0, "b0": b0,
                "u": u_hat, "k": k, "slope_mag": slope_mag, "dE_max": dE_max,
                "sigma": (sigma_L, sigma_a, sigma_b),
            }
    return out

traj_room = build_spot_traj("室温")
traj_const = build_spot_traj("恒温")

# Build predicted (恒温恒湿) by scaling the 恒温 slope_mag by rh_ratio (k scales by same factor)
def build_predicted(traj_src):
    out = {"_t": traj_src["_t"], "_n_meas": 0}  # all "extrapolated" for predicted
    for key, spot in traj_src.items():
        if isinstance(key, str): continue
        grp, sidx, c = key
        dE_max = spot["dE_max"]
        u_hat = spot["u"]
        L0, a0, b0 = spot["L0"], spot["a0"], spot["b0"]
        sigma_L, sigma_a, sigma_b = spot["sigma"]
        k_pred = spot["k"] * rh_ratio   # higher RH → faster k (because slope scales by rh_ratio
                                         # and we keep dE_max fixed: k = slope/dE_max)
        L_traj, a_traj, b_traj = [], [], []
        for t in traj_src["_t"]:
            dE_t = dE_max * (1.0 - np.exp(-k_pred * t))
            L_traj.append(L0 + dE_t * u_hat[0] + np.random.normal(0, sigma_L))
            a_traj.append(a0 + dE_t * u_hat[1] + np.random.normal(0, sigma_a))
            b_traj.append(b0 + dE_t * u_hat[2] + np.random.normal(0, sigma_b))
        out[key] = {
            "L": np.array(L_traj), "a": np.array(a_traj), "b": np.array(b_traj),
            "L0": L0, "a0": a0, "b0": b0,
        }
    return out

traj_pred = build_predicted(traj_const)

# Compute ΔE* per spot for plotting
def trajs_to_dE(traj_obj):
    out = {}
    for key, spot in traj_obj.items():
        if isinstance(key, str): continue
        L = spot["L"]; a = spot["a"]; b = spot["b"]
        L0 = traj_obj[key].get("L0", L[0])
        a0 = traj_obj[key].get("a0", a[0])
        b0 = traj_obj[key].get("b0", b[0])
        # Use the actual first measured baseline (not the fit intercept) for ΔE* reference
        # to match the convention used in earlier plots.
        # But for first-order this is the fit intercept; close enough.
        dE = np.sqrt((L - L0) ** 2 + (a - a0) ** 2 + (b - b0) ** 2)
        out[key] = dE
    out["_t"] = traj_obj["_t"]
    return out

dE_room  = trajs_to_dE(traj_room)
dE_const = trajs_to_dE(traj_const)
dE_pred  = trajs_to_dE(traj_pred)

arm_data = {
    "23 C / 40%RH, ambient light  (measured + projected)":  dE_room,
    "40 C / 10%RH, dark  (measured + projected)":           dE_const,
    "40 C / 80%RH  (RH-scaled prediction)":                  dE_pred,
}

styles = {
    "23 C / 40%RH, ambient light  (measured + projected)":  {"color": "#2E86AB", "ls": "-",  "marker": "o"},
    "40 C / 10%RH, dark  (measured + projected)":           {"color": "#E63946", "ls": "-",  "marker": "s"},
    "40 C / 80%RH  (RH-scaled prediction)":                  {"color": "#6A4C93", "ls": "--", "marker": "^"},
}

fig, axes = plt.subplots(2, 3, figsize=(14, 8), sharex=True)
geom_groups = [("block (10x10x4 cm)", ["R", "G", "B"]),
               ("pedestal (1:5 lotus base)", ["TR", "TG", "TB"])]

for row, (geom, groups) in enumerate(geom_groups):
    row_max = 0.0
    for grp in groups:
        cols = spec_cols[grp]
        for arm_label, dE_data in arm_data.items():
            for c in cols:
                for key in dE_data:
                    if isinstance(key, str): continue
                    if key[0] == grp and key[2] == c:
                        m = np.nanmax(dE_data[key])
                        if not np.isnan(m): row_max = max(row_max, m)

    for col, grp in enumerate(groups):
        ax = axes[row, col]
        cols = spec_cols[grp]
        for arm_label, dE_data in arm_data.items():
            t = dE_data["_t"]
            spot_dEs = []
            for c in cols:
                for key in dE_data:
                    if isinstance(key, str): continue
                    if key[0] == grp and key[2] == c:
                        spot_dEs.append(dE_data[key])
            spot_arr = np.array(spot_dEs)
            mean = np.nanmean(spot_arr, axis=0)
            mn = np.nanmin(spot_arr, axis=0); mx = np.nanmax(spot_arr, axis=0)
            st = styles[arm_label]
            ax.fill_between(t, mn, mx, alpha=0.12, color=st["color"])
            for s in spot_arr:
                ax.plot(t, s, color=st["color"], linestyle=st["ls"],
                        alpha=0.25, linewidth=0.8)
            ax.plot(t, mean, color=st["color"], linestyle=st["ls"],
                    marker=st["marker"], markersize=3.5, linewidth=1.8,
                    label=arm_label if (row == 0 and col == 0) else None)
        ax.set_title(f"{PIG[grp]} — {geom}", fontsize=10)
        ax.grid(True, alpha=0.3)
        ax.set_xlim(-3, 195)
        ax.set_ylim(-0.2, max(row_max * 1.15 + 1, DELTAE_MAX[grp] * 1.08))
        if col == 0: ax.set_ylabel(r"$\Delta E^*_{ab}$")
        if row == 1: ax.set_xlabel("Days since 17 Apr 2026")
        n_spots = len(cols)
        ax.text(0.97, 0.04, f"n={n_spots} spots",
                transform=ax.transAxes, ha="right", va="bottom",
                fontsize=7, alpha=0.6, style="italic")

handles, labels = axes[0, 0].get_legend_handles_labels()
fig.legend(handles, labels, loc="upper center", ncol=3,
           frameon=False, fontsize=9.5, bbox_to_anchor=(0.5, 0.965))

fig.suptitle("Per-pigment dE* trajectories: 34-day measurement and first-order projection to day 180",
             y=0.985, fontsize=12, fontweight="bold")
# Vertical line marking measurement / projection boundary
for row in range(2):
    for col in range(3):
        axes[row, col].axvline(34, color="gray", lw=0.8, ls=":", alpha=0.6)
plt.subplots_adjust(top=0.90, bottom=0.10, left=0.07, right=0.99, hspace=0.32, wspace=0.22)
# Footer annotation
fig.text(0.5, 0.02,
         "Solid markers (days 0-34): measured. Lines past dotted vertical (day 34): "
         "first-order projection with matched per-spot noise (sigma from days 0-34 residuals).",
         ha="center", va="bottom", fontsize=8, style="italic", color="#555")

out = "experiments/deltaE_trajectories_180d_firstorder.png"
plt.savefig(out, dpi=150, bbox_inches="tight")
print(f"Saved: {out}")

# Summary table
print("\n=== Day-180 ΔE* (artefact-mean) — first-order model ===")
print(f"{'Pigment / geom':24s}  {'室温 (23 C)':>12s}  {'恒温 (40 C)':>12s}  {'恒温恒湿':>10s}")
for grp_label, grp_key in [("Vermilion (block)","R"), ("Malachite (block)","G"), ("Azurite (block)","B"),
                            ("Vermilion (pedestal)","TR"), ("Malachite (pedestal)","TG"), ("Azurite (pedestal)","TB")]:
    cols = spec_cols[grp_key]
    means = []
    for dE_data in (dE_room, dE_const, dE_pred):
        spot_vals = []
        for c in cols:
            for key in dE_data:
                if isinstance(key, str): continue
                if key[0] == grp_key and key[2] == c:
                    spot_vals.append(dE_data[key][-1])
        means.append(np.nanmean(spot_vals) if spot_vals else float("nan"))
    print(f"  {grp_label:22s}  {means[0]:>12.2f}  {means[1]:>12.2f}  {means[2]:>10.2f}")
