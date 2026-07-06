"""Convert the RAW measured workbook (experiments/mogao_data.xlsx) into the same
tidy form as mogao_data_consolidated.xlsx -> experiments/mogao_data_measured_consolidated.xlsx

This is a FORMAT conversion only: the L*a*b* values are the actual measurements,
untouched (no model, no denoising, no gap-filling). Only the two experimentally-run
arms are present (the 40C/80%RH sheet is empty in the raw file). Known gross outliers
are left in place but reported in the README.

Sheets: README | data_long | deltaE_mean
"""
import openpyxl, math, re
from openpyxl.styles import Font, PatternFill

DATE_RE = re.compile(r"^\s*\d+\.\d+\s*$")
def to_f(v):
    try: return float(v)
    except Exception: return None
def is_date(v): return v is not None and bool(DATE_RE.match(str(v).strip()))
def date_to_day(d):
    m, day = str(d).strip().split("."); dpm=[31,28,31,30,31,30,31,31,30,31,30,31]
    return (sum(dpm[:int(m)-1]) + int(day)) - (sum(dpm[:3]) + 17)

spec_cols = {"R":[2,5,8],"G":[11,14,17],"B":[20,23,26],"TR":[29,32],"TG":[35,38],"TB":[41,44]}
PIGMENT   = {"R":"Vermilion","TR":"Vermilion","G":"Malachite","TG":"Malachite","B":"Azurite","TB":"Azurite"}
# 'T' in the label = Tile = BLOCK; no-T = PEDESTAL (lotus base). (Was swapped.)
GEOMETRY  = {"R":"pedestal","G":"pedestal","B":"pedestal","TR":"block","TG":"block","TB":"block"}
SHEETS = {   # only the two experimentally-measured arms (40C/80%RH sheet is empty)
    # NB: both arms share the SAME lighting (per experimenter) -- they differ only in
    # temperature and humidity. Earlier "lit"/"dark" labels were incorrect and removed.
    "室温": ("23C_40RH_room",    "23 C / 40%RH (room-temp arm)"),
    "恒温": ("40C_10RH_chamber", "40 C / 10%RH (chamber arm)"),
}
OUTLIER_dE = 8.0   # ΔE above this flagged as a gross outlier in the README

wb = openpyxl.load_workbook("experiments/mogao_data.xlsx", data_only=True)

rows_out, outliers = [], []
for sh, (arm_code, condition) in SHEETS.items():
    ws = wb[sh]
    drows = [r for r in range(1, ws.max_row+1) if is_date(ws.cell(r,1).value)]
    days  = [date_to_day(ws.cell(r,1).value) for r in drows]
    for grp, cols in spec_cols.items():
        for sidx, c in enumerate(cols):
            label = ws.cell(2, c).value or f"{grp}.{sidx+1}"
            r0 = drows[0]
            L0,a0,b0 = (to_f(ws.cell(r0,c).value), to_f(ws.cell(r0,c+1).value), to_f(ws.cell(r0,c+2).value))
            for r, d in zip(drows, days):
                L,a,b = (to_f(ws.cell(r,c).value), to_f(ws.cell(r,c+1).value), to_f(ws.cell(r,c+2).value))
                dE = None
                if None not in (L,a,b,L0,a0,b0):
                    dE = round(math.sqrt((L-L0)**2+(a-a0)**2+(b-b0)**2), 2)
                    if dE > OUTLIER_dE:
                        outliers.append((condition, str(label), str(ws.cell(r,1).value), d, dE))
                rows_out.append([arm_code, condition, "measured", PIGMENT[grp], GEOMETRY[grp],
                                 str(label), str(ws.cell(r,1).value), d, L, a, b, dE, "measured"])

def write_dE_formula(ws, rows):
    """Append rows, writing the deltaE cell (col L) as a live Excel formula from
    L*a*b* (cols I,J,K) vs each spot's own day-0 baseline row."""
    base = None
    for row in rows:
        ws.append(row); er = ws.max_row
        day, L, a, b = row[7], row[8], row[9], row[10]
        if day == 0:
            base = er; ws.cell(er, 12).value = 0
        elif base and None not in (L, a, b):
            ws.cell(er, 12).value = (f"=ROUND(SQRT((I{er}-I{base})^2+"
                                     f"(J{er}-J{base})^2+(K{er}-K{base})^2),2)")

# ---- write workbook ----
out = openpyxl.Workbook()
hdr_font = Font(bold=True, color="FFFFFF"); hdr_fill = PatternFill("solid", fgColor="365F91")

rm = out.active; rm.title = "README"
readme = [
    ("Mogao pigment colour-change — RAW MEASURED dataset (tidy form)", True),
    ("", False),
    ("Source: experiments/mogao_data.xlsx (the actual instrument measurements).", False),
    ("This is a FORMAT conversion only. L*a*b* values are the real measurements, UNMODIFIED:", False),
    ("no model, no outlier removal, no denoising, no gap-filling.", False),
    ("", False),
    ("deltaE = CIE76 ΔE*ab of each spot vs its own day-0 (17 Apr) L*a*b* baseline.", False),
    ("", False),
    ("Arms: two experimentally-run conditions only, distinguished by temperature and humidity.", False),
    ("Lighting was ambient / uncontrolled and IDENTICAL for both arms (NOT a lit/dark pair;", False),
    ("light was not a controlled variable, so it cannot drive any room-vs-chamber difference).", False),
    ("The 40C/80%RH condition was never run (empty in the source) and is absent here.", False),
    ("", False),
    ("Cadence: 14 real timepoints over days 0-61, with two holiday gaps (day 12->25 and 41->53).", False),
    ("Spots: 3 per pigment on pedestals, 2 per pigment on block tiles.", False),
    ("", False),
    (f"KNOWN GROSS OUTLIERS (ΔE > {OUTLIER_dE:.0f}, left in place - review before use):", True),
]
if outliers:
    for cond,label,date,day,dE in outliers:
        readme.append((f"  {cond} | spot {label} | {date} (day {day}) | ΔE = {dE}", False))
else:
    readme.append(("  none", False))
for i,(txt,bold) in enumerate(readme, 1):
    cc = rm.cell(i,1,txt); cc.font = Font(bold=True) if bold else Font()
rm.column_dimensions["A"].width = 100

cols = ["arm_code","condition","status","pigment","geometry","spot","date","day",
        "L*","a*","b*","deltaE","point_type"]

# ---- one sheet per pigment (data_long / deltaE_mean intentionally omitted) ----
def _split_sheet(name, keep):
    ws = out.create_sheet(name[:31]); ws.append(cols)
    for j in range(1,len(cols)+1):
        ws.cell(1,j).font=hdr_font; ws.cell(1,j).fill=hdr_fill
    write_dE_formula(ws, [r for r in rows_out if keep(r)])
    ws.freeze_panes="A2"
    for j,w in enumerate([22,30,12,11,10,9,7,6,8,8,8,9,12],1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width=w
for pig in ["Vermilion","Malachite","Azurite"]:              # one sheet per pigment
    _split_sheet(pig, lambda r, p=pig: r[3]==p)

dst="experiments/mogao_data_measured_consolidated.xlsx"
out.save(dst)
print(f"Saved {dst}")
print(f"  sheets: {out.sheetnames}")
print(f"  gross outliers flagged (ΔE > {OUTLIER_dE:.0f}): {len(outliers)}")
for o in outliers: print("   ", o)
