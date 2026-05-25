"""Extend the per-spot trajectories to day 180 via linear fit + matched-amplitude noise.

Reads:  experiments/敦煌变化数据(1).xlsx
Writes: experiments/敦煌变化数据_180d.xlsx
"""
import openpyxl, math, re
import numpy as np
from openpyxl.styles import Font, PatternFill

np.random.seed(42)  # reproducible noise

DATE_RE = re.compile(r"^\s*\d+\.\d+\s*$")

def to_f(v):
    try: return float(v)
    except Exception: return float("nan")

def is_date(v):
    if v is None: return False
    return bool(DATE_RE.match(str(v).strip()))

def date_to_day(d):
    s = str(d).strip(); m, day = s.split(".")
    dpm = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    return (sum(dpm[: int(m) - 1]) + int(day)) - (sum(dpm[:3]) + 17)

def day_to_date(d):
    dpm = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    base = sum(dpm[:3]) + 17
    target = base + int(d)
    month = 1
    while target > dpm[month - 1]:
        target -= dpm[month - 1]; month += 1
    return f"{month}.{target:02d}"

spec_cols = {
    "R":  [2, 5, 8],
    "G":  [11, 14, 17],
    "B":  [20, 23, 26],
    "TR": [29, 32],
    "TG": [35, 38],
    "TB": [41, 44],
}
SRC_OUT = {
    ("室温", "R",  0, "5.12"),  # all 3 channels off-scale
    ("恒温", "TR", 0, "5.14"),  # a* sign-flip
}

src = "experiments/敦煌变化数据(1).xlsx"
dst = "experiments/敦煌变化数据_180d.xlsx"

wb = openpyxl.load_workbook(src)

q = 0.8
T_t = 40; RH_t = 80; RH_ref = 40
rh_ratio = (RH_t / RH_ref) ** q

extra_days = list(range(42, 181, 14))
if extra_days[-1] != 180:
    extra_days.append(180)
print(f"Extrapolated time points (days): {extra_days}")

measured_fill = PatternFill(start_color="E8F4FF", end_color="E8F4FF", fill_type="solid")
extrap_fill   = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
predicted_fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
outlier_fill   = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")

# --- Step 1: fit per spot, build extended trajectories WITH matched noise ---
extended = {}  # extended[sheet][col] = {days, L, a, b, group, spot, date_strs, sigmas}
for sheet in ("室温", "恒温"):
    ws = wb[sheet]
    rows_existing = [r for r in range(4, ws.max_row + 1) if is_date(ws.cell(r, 1).value)]
    measured_days = np.array([date_to_day(ws.cell(r, 1).value) for r in rows_existing])
    date_strs = [str(ws.cell(r, 1).value).strip() for r in rows_existing]
    extended[sheet] = {"_measured_days": measured_days, "_date_strs": date_strs}
    for grp, cols in spec_cols.items():
        for sidx, c in enumerate(cols):
            L_meas, a_meas, b_meas = [], [], []
            for i_t, r in enumerate(rows_existing):
                if (sheet, grp, sidx, date_strs[i_t]) in SRC_OUT:
                    L_meas.append(np.nan); a_meas.append(np.nan); b_meas.append(np.nan)
                    continue
                L_meas.append(to_f(ws.cell(r, c).value))
                a_meas.append(to_f(ws.cell(r, c + 1).value))
                b_meas.append(to_f(ws.cell(r, c + 2).value))
            L_meas = np.array(L_meas); a_meas = np.array(a_meas); b_meas = np.array(b_meas)
            mask = ~np.isnan(L_meas) & ~np.isnan(a_meas) & ~np.isnan(b_meas)
            t_fit = measured_days[mask]
            if mask.sum() >= 2:
                sL, iL = np.polyfit(t_fit, L_meas[mask], 1)
                sa, ia = np.polyfit(t_fit, a_meas[mask], 1)
                sb, ib = np.polyfit(t_fit, b_meas[mask], 1)
            else:
                sL = sa = sb = 0.0
                iL = np.nanmean(L_meas); ia = np.nanmean(a_meas); ib = np.nanmean(b_meas)
            # Residual std per channel = noise amplitude to inject in extrapolation
            res_L = L_meas[mask] - (sL * t_fit + iL)
            res_a = a_meas[mask] - (sa * t_fit + ia)
            res_b = b_meas[mask] - (sb * t_fit + ib)
            sigma_L = float(np.std(res_L, ddof=1)) if res_L.size >= 2 else 0.0
            sigma_a = float(np.std(res_a, ddof=1)) if res_a.size >= 2 else 0.0
            sigma_b = float(np.std(res_b, ddof=1)) if res_b.size >= 2 else 0.0

            all_days = list(measured_days) + extra_days
            all_L = list(L_meas); all_a = list(a_meas); all_b = list(b_meas)
            for d in extra_days:
                Lp = sL * d + iL + np.random.normal(0, sigma_L)
                ap = sa * d + ia + np.random.normal(0, sigma_a)
                bp = sb * d + ib + np.random.normal(0, sigma_b)
                all_L.append(Lp); all_a.append(ap); all_b.append(bp)

            extended[sheet][c] = {
                "days": all_days, "L": all_L, "a": all_a, "b": all_b,
                "group": grp, "spot": sidx, "date_strs": date_strs,
                "sigmas": (sigma_L, sigma_a, sigma_b),
                "n_measured": len(rows_existing),
            }

# --- Step 2: rewrite 室温 and 恒温 sheets ---
for sheet in ("室温", "恒温"):
    ws = wb[sheet]
    for r in range(4, ws.max_row + 30):
        for c in range(1, ws.max_column + 2):
            ws.cell(r, c).value = None
            ws.cell(r, c).fill = PatternFill(fill_type=None)
    sample_traj = extended[sheet][spec_cols["R"][0]]
    all_days_template = sample_traj["days"]
    for i, d in enumerate(all_days_template):
        ws.cell(4 + i, 1).value = day_to_date(d)
    for c, traj in extended[sheet].items():
        if isinstance(c, str): continue  # skip _measured_days etc.
        grp = traj["group"]; sidx = traj["spot"]; date_strs = traj["date_strs"]
        n_meas = traj["n_measured"]
        for i in range(len(traj["days"])):
            row = 4 + i
            L = traj["L"][i]; a = traj["a"][i]; b = traj["b"][i]
            is_extrap = i >= n_meas
            is_outlier = (i < n_meas) and ((sheet, grp, sidx, date_strs[i]) in SRC_OUT)
            if is_outlier:
                for off in (0, 1, 2):
                    ws.cell(row, c + off).value = None
                    ws.cell(row, c + off).fill = outlier_fill
            else:
                ws.cell(row, c).value     = round(L, 2) if not np.isnan(L) else None
                ws.cell(row, c + 1).value = round(a, 2) if not np.isnan(a) else None
                ws.cell(row, c + 2).value = round(b, 2) if not np.isnan(b) else None
                fill = extrap_fill if is_extrap else measured_fill
                for off in (0, 1, 2): ws.cell(row, c + off).fill = fill

# --- Step 3: rebuild 恒温恒湿 = RH-scaled 恒温 extended trajectory + FRESH noise ---
ws_hh = wb["恒温恒湿"]
for r in range(4, ws_hh.max_row + 40):
    for c in range(1, ws_hh.max_column + 2):
        ws_hh.cell(r, c).value = None
        ws_hh.cell(r, c).fill = PatternFill(fill_type=None)
ws_hh.cell(1, 1).value = (
    f"恒温恒湿组（PREDICTED）— {T_t}°C, {RH_t}%RH, q={q}, extended to 180 d"
)
ws_hh.cell(1, 1).font = Font(bold=True, color="B00020")

sample_traj_h = extended["恒温"][spec_cols["R"][0]]
all_days_template = sample_traj_h["days"]
for i, d in enumerate(all_days_template):
    ws_hh.cell(4 + i, 1).value = day_to_date(d)

# To build the predicted trajectory, scale the *underlying linear trend* (not noisy realisations).
# Recompute slopes from the cleaned source so we don't double-inject noise.
for c, traj in extended["恒温"].items():
    if isinstance(c, str): continue
    grp = traj["group"]; sidx = traj["spot"]
    n_meas = traj["n_measured"]
    # Re-derive baseline from the original measured points (first non-outlier)
    L_meas_raw, a_meas_raw, b_meas_raw, t_meas = [], [], [], []
    for i in range(n_meas):
        if (sheet := "恒温") and (grp, sidx, traj["date_strs"][i]) in {x[1:] for x in SRC_OUT if x[0] == "恒温"}:
            continue
        L = traj["L"][i]; a = traj["a"][i]; b = traj["b"][i]
        if np.isnan(L) or np.isnan(a) or np.isnan(b): continue
        L_meas_raw.append(L); a_meas_raw.append(a); b_meas_raw.append(b)
        t_meas.append(traj["days"][i])
    L_meas_raw = np.array(L_meas_raw); a_meas_raw = np.array(a_meas_raw); b_meas_raw = np.array(b_meas_raw)
    t_meas = np.array(t_meas, dtype=float)
    if len(t_meas) >= 2:
        sL, iL = np.polyfit(t_meas, L_meas_raw, 1)
        sa, ia = np.polyfit(t_meas, a_meas_raw, 1)
        sb, ib = np.polyfit(t_meas, b_meas_raw, 1)
    else:
        sL = sa = sb = 0.0
        iL = np.nanmean(L_meas_raw); ia = np.nanmean(a_meas_raw); ib = np.nanmean(b_meas_raw)
    sigma_L, sigma_a, sigma_b = traj["sigmas"]
    # Baseline at t=0 for the RH-scaling reference
    L0 = sL * 0 + iL; a0 = sa * 0 + ia; b0 = sb * 0 + ib
    # Build predicted trajectory: predicted_centre(t) = baseline + rh_ratio*(linear_slope*t)
    for i, d in enumerate(traj["days"]):
        row = 4 + i
        dL = sL * d  # change relative to t=0
        da = sa * d
        db = sb * d
        Lp = L0 + dL * rh_ratio + np.random.normal(0, sigma_L)
        ap = a0 + da * rh_ratio + np.random.normal(0, sigma_a)
        bp = b0 + db * rh_ratio + np.random.normal(0, sigma_b)
        ws_hh.cell(row, c).value     = round(float(Lp), 2)
        ws_hh.cell(row, c + 1).value = round(float(ap), 2)
        ws_hh.cell(row, c + 2).value = round(float(bp), 2)
        for off in (0, 1, 2): ws_hh.cell(row, c + off).fill = predicted_fill

# Update sheet titles
wb["室温"].cell(1, 1).value = "室温组 (23°C, ~40%RH) — measured 0–34 d, linearly extrapolated 35–180 d (matched-noise)"
wb["室温"].cell(1, 1).font = Font(bold=True)
wb["恒温"].cell(1, 1).value = "恒温组 (40°C, ~40%RH) — measured 0–34 d, linearly extrapolated 35–180 d (matched-noise)"
wb["恒温"].cell(1, 1).font = Font(bold=True)

note_row = 4 + len(all_days_template) + 2
notes = [
    "EXTRAPOLATION + PREDICTION — values past day 34 are model output, not measurement.",
    "Method (室温 / 恒温 extrapolated rows): linear least-squares fit of L*(t), a*(t), b*(t) per spot",
    "  on the 9 measured time points; extended at 14-day intervals from day 42 to day 180; Gaussian",
    "  noise added per channel with σ matched to the residuals of the linear fit on that spot.",
    f"Method (恒温恒湿 sheet): per-spot RH scaling of the linear-trend slope by factor",
    f"  (RH_target/RH_ref)^q = (80/40)^0.8 = {rh_ratio:.3f}; fresh noise drawn with the same σ",
    "  as the 恒温 source (instrument/surface noise is treated as condition-independent).",
    "Cell colours: pale blue = measured; pale yellow = extrapolated (恒温/室温);",
    "              pale purple = predicted (恒温恒湿); pale red = outlier (masked from fit).",
    "Random seed = 42 (rerun reproduces the same noise realisation).",
    "",
    "Outliers masked before linear fit:",
    "  • 室温 R.1 @ 5.12: L,a,b all off-scale (instrument misfire)",
    "  • 恒温 TR.1 @ 5.14: a* sign-flip",
    "",
    "Caveats:",
    "  • Linear extrapolation assumes the per-spot rate observed over days 0–34 holds to day 180;",
    "    pigment fading may saturate (sub-linear) over this longer horizon.",
    "  • The 34-day window did not show detectable thermal acceleration for block specimens;",
    "    extrapolation inherits this, so 室温 and 恒温 trajectories at day 180 are similar.",
    "  • Noise injected is uncorrelated Gaussian per spot per channel; true noise may have",
    "    day-of-measurement correlation across spots which is not captured here.",
]
for i, line in enumerate(notes):
    cell = ws_hh.cell(note_row + i, 1)
    cell.value = line
    cell.font = Font(
        italic=(line not in ("", "EXTRAPOLATION + PREDICTION — values past day 34 are model output, not measurement.",
                             "Caveats:", "Outliers masked before linear fit:")),
        bold=(i == 0),
    )

wb.save(dst)
print(f"Saved extended workbook → {dst}")

# --- Summary: predicted ΔE* at day 180 (using noisy realisation) ---
print("\n=== Day-180 ΔE* (artefact-mean across spots, after extrapolation + matched noise) ===")
fmt = "  {:22s}  {:>14s}  {:>14s}  {:>16s}"
print(fmt.format("Pigment / geom", "室温 (23°C)", "恒温 (40°C)", "恒温恒湿 (pred)"))

ws_hh_read = openpyxl.load_workbook(dst, data_only=True)["恒温恒湿"]
for label, grp_key in [("Vermilion (block)","R"), ("Malachite (block)","G"), ("Azurite (block)","B"),
                       ("Vermilion (pedestal)","TR"), ("Malachite (pedestal)","TG"), ("Azurite (pedestal)","TB")]:
    cols = spec_cols[grp_key]
    cell_means = []
    for sheet in ("室温", "恒温"):
        ws_re = openpyxl.load_workbook(dst, data_only=True)[sheet]
        dEs = []
        for c in cols:
            L0 = to_f(ws_re.cell(4, c).value); a0 = to_f(ws_re.cell(4, c+1).value); b0 = to_f(ws_re.cell(4, c+2).value)
            # find last row with values for this column
            last_row = None
            for rr in range(4, 4 + 30):
                if ws_re.cell(rr, c).value is not None and isinstance(ws_re.cell(rr, c).value, (int, float)):
                    last_row = rr
            if last_row is None or np.isnan(L0): continue
            Lf = to_f(ws_re.cell(last_row, c).value); af = to_f(ws_re.cell(last_row, c+1).value); bf = to_f(ws_re.cell(last_row, c+2).value)
            dEs.append(math.sqrt((Lf - L0) ** 2 + (af - a0) ** 2 + (bf - b0) ** 2))
        cell_means.append(np.mean(dEs) if dEs else float("nan"))
    # predicted from 恒温恒湿 sheet
    dEs_p = []
    for c in cols:
        L0 = to_f(ws_hh_read.cell(4, c).value); a0 = to_f(ws_hh_read.cell(4, c+1).value); b0 = to_f(ws_hh_read.cell(4, c+2).value)
        last_row = None
        for rr in range(4, 4 + 30):
            v = ws_hh_read.cell(rr, c).value
            if v is not None and isinstance(v, (int, float)): last_row = rr
        if last_row is None or np.isnan(L0):
            dEs_p.append(float("nan")); continue
        Lf = to_f(ws_hh_read.cell(last_row, c).value); af = to_f(ws_hh_read.cell(last_row, c+1).value); bf = to_f(ws_hh_read.cell(last_row, c+2).value)
        dEs_p.append(math.sqrt((Lf - L0) ** 2 + (af - a0) ** 2 + (bf - b0) ** 2))
    cell_means.append(np.mean(dEs_p) if dEs_p else float("nan"))
    print(fmt.format(label, f"{cell_means[0]:.2f}", f"{cell_means[1]:.2f}", f"{cell_means[2]:.2f}"))
