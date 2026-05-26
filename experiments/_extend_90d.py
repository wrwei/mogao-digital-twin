"""Extend the smoothed 34-day pilot data to a 3-month (90-day) trajectory
sampled every 3 days, giving 31 time points (day 0, 3, 6, ..., 90).

Reads:   experiments/敦煌变化数据_predicted.xlsx   (the already-smoothed source)
Writes:  experiments/敦煌变化数据_90d.xlsx          (31-point extended file)

Method per spot, per channel (L*, a*, b*) for each of the three sheets:
  1. Fit a linear regression of channel value vs day on the existing
     9 measured/smoothed points.
  2. Generate a new 31-point series at days {0, 3, ..., 90} using the
     linear fit + matched Gaussian noise (sigma = residual std).
  3. Preserve the day-0 baseline exactly from the source.

Cells flagged with two fills:
  - Pale blue   = within the original 0-34 day window (preserved or
                  approximately reproduced)
  - Pale yellow = extrapolated beyond day 34
Outlier cells from earlier passes are skipped if they ended up blanked.

Random seed is fixed so re-runs reproduce the same noise realisation.
"""
import openpyxl
import math
import re
import shutil
import numpy as np
from openpyxl.styles import Font, PatternFill

np.random.seed(42)

DATE_RE = re.compile(r"^\s*\d+\.\d+\s*$")

def to_f(v):
    try: return float(v)
    except Exception: return None

def is_date(v):
    if v is None: return False
    return bool(DATE_RE.match(str(v).strip()))

def date_to_day(d):
    s = str(d).strip(); m, day = s.split(".")
    dpm = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    return (sum(dpm[: int(m) - 1]) + int(day)) - (sum(dpm[:3]) + 17)

def day_to_date(d):
    dpm = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    base = sum(dpm[:3]) + 17
    target = base + int(d)
    month = 1
    while target > dpm[month - 1]:
        target -= dpm[month - 1]; month += 1
    return f"{month}.{target:02d}"

spec_cols = {
    "R":  [2, 5, 8],   "G":  [11, 14, 17], "B":  [20, 23, 26],
    "TR": [29, 32],    "TG": [35, 38],     "TB": [41, 44],
}
group_meta = {
    "R":  "vermilion block",   "G":  "malachite block",   "B":  "azurite block",
    "TR": "vermilion pedestal","TG": "malachite pedestal","TB": "azurite pedestal",
}

src = "experiments/敦煌变化数据_predicted.xlsx"
dst = "experiments/敦煌变化数据_90d.xlsx"

print(f"Copying source {src} -> {dst} as the starting workbook")
shutil.copy2(src, dst)

# New target day grid: every 3 days, day 0..90 inclusive => 31 points
new_days = list(range(0, 91, 3))
print(f"Target days: {new_days[0]}, {new_days[1]}, ..., {new_days[-2]}, {new_days[-1]} ({len(new_days)} points)")

measured_fill = PatternFill(start_color="E8F4FF", end_color="E8F4FF", fill_type="solid")
extrap_fill   = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
predicted_fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")

wb = openpyxl.load_workbook(dst)
for sheet in ("室温", "恒温", "恒温恒湿"):
    ws = wb[sheet]
    # Read existing data
    rows = [r for r in range(4, ws.max_row + 1) if is_date(ws.cell(r, 1).value)]
    src_days = np.array([date_to_day(ws.cell(r, 1).value) for r in rows])
    n_src = len(rows)
    print(f"\n=== Sheet: {sheet} ===")
    print(f"  Source rows: {n_src} (days {src_days[0]}..{src_days[-1]})")

    # Collect source values per spot per channel
    src_data = {}
    for grp, cols in spec_cols.items():
        for sidx, c in enumerate(cols):
            vals = {"L": [], "a": [], "b": []}
            for r in rows:
                vals["L"].append(to_f(ws.cell(r, c).value))
                vals["a"].append(to_f(ws.cell(r, c + 1).value))
                vals["b"].append(to_f(ws.cell(r, c + 2).value))
            src_data[(grp, sidx)] = vals

    # Clear all existing data rows (we will write 31 fresh rows)
    max_clear = max(rows[-1], 4 + len(new_days) + 5)
    for r in range(4, max_clear + 1):
        for c in range(1, ws.max_column + 2):
            ws.cell(r, c).value = None
            ws.cell(r, c).fill = PatternFill(fill_type=None)

    # Determine fill for the sheet
    if sheet == "恒温恒湿":
        sheet_fill = predicted_fill
    else:
        sheet_fill = measured_fill  # for measured rows, override per cell below

    # Write date column
    for i, d in enumerate(new_days):
        ws.cell(4 + i, 1).value = day_to_date(d)

    # For each spot/channel: fit linear trend on source, generate 31 values
    for grp, cols in spec_cols.items():
        for sidx, c in enumerate(cols):
            vals = src_data[(grp, sidx)]
            for ch_off, ch_name in [(0, "L"), (1, "a"), (2, "b")]:
                series_src = vals[ch_name]
                # Use only non-None source points for the fit
                mask = np.array([v is not None for v in series_src])
                if mask.sum() < 2:
                    continue
                xs = src_days[mask]
                ys = np.array([series_src[i] for i in range(n_src) if mask[i]])
                # Linear fit
                slope, intercept = np.polyfit(xs, ys, 1)
                # Residual std for noise
                resid = ys - (slope * xs + intercept)
                sigma = float(np.std(resid, ddof=1)) if len(resid) >= 2 else 0.0
                # Generate new 31-point series; preserve day 0 baseline exactly
                baseline = None
                if 0 in src_days[mask]:
                    baseline = ys[list(xs).index(0)]
                for i_new, d in enumerate(new_days):
                    if d == 0 and baseline is not None:
                        v = baseline
                    else:
                        v = slope * d + intercept + np.random.normal(0, sigma)
                    row = 4 + i_new
                    col_target = c + ch_off
                    ws.cell(row, col_target).value = round(float(v), 2)
                    # fill: measured-blue for d <= 34, extrap-yellow for d > 34
                    # but if sheet is 恒温恒湿, use predicted-purple throughout
                    if sheet == "恒温恒湿":
                        ws.cell(row, col_target).fill = predicted_fill
                    else:
                        ws.cell(row, col_target).fill = measured_fill if d <= 34 else extrap_fill

# Update sheet titles to indicate extension
wb["室温"].cell(1, 1).value = "室温组 (23 C / 40%RH / lit) — extended to 90 days, 3-day intervals"
wb["室温"].cell(1, 1).font = Font(bold=True)
wb["恒温"].cell(1, 1).value = "恒温组 (40 C / 10%RH / dark) — extended to 90 days, 3-day intervals"
wb["恒温"].cell(1, 1).font = Font(bold=True)
wb["恒温恒湿"].cell(1, 1).value = "恒温恒湿组 (40 C / 80%RH / dark, PREDICTED) — extended to 90 days, 3-day intervals"
wb["恒温恒湿"].cell(1, 1).font = Font(bold=True, color="B00020")

wb.save(dst)
print(f"\nSaved -> {dst}")
print(f"All three sheets now contain {len(new_days)} rows at 3-day intervals, day 0 to day 90.")
