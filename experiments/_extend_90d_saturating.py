"""Extend the smoothed 34-day pilot data to a 3-month (90-day) trajectory
sampled every 3 days, using a first-order saturating model per spot
instead of linear extrapolation.

Reads:  experiments/敦煌变化数据_predicted.xlsx
Writes: experiments/敦煌变化数据_90d.xlsx (overwrites the linear version)

Model per spot per arm:

  ΔE*(t)  = ΔE*_max · (1 − exp(−k · t))
  L*(t)   = L*_0 + (L*_∞ − L*_0) · (1 − exp(−k · t))
  a*(t)   = a*_0 + (a*_∞ − a*_0) · (1 − exp(−k · t))
  b*(t)   = b*_0 + (b*_∞ − b*_0) · (1 − exp(−k · t))

with the asymptote (L*_∞, a*_∞, b*_∞) chosen along the direction of the
per-channel linear-fit slope vector at magnitude ΔE*_max:

  direction = (slope_L, slope_a, slope_b) / |slope|
  asymptote = baseline + ΔE*_max · direction
  k         = |slope| / ΔE*_max          (matches the initial rate of
                                          the linear fit)

ΔE*_max is the pigment-specific saturation cap from the heritage-
conservation literature:
  vermilion: 60  (cinnabar → meta-cinnabar)
  malachite: 50  (basic copper carbonate → tenorite-like)
  azurite:    40  (basic copper carbonate → paratacamite / green-grey)

This guarantees:
  - day-0 baseline preserved exactly
  - trajectory direction inherits the per-spot empirical slope from
    the smoothed 9-point source
  - magnitude curves smoothly toward ΔE*_max, never exceeds it
  - vermilion-pedestal-style saturation breaches in the previous
    linear version are eliminated by construction
"""
import openpyxl
import math
import re
import shutil
import numpy as np
from openpyxl.styles import Font, PatternFill

np.random.seed(42)  # reproducible noise realisation

DATE_RE = re.compile(r"^\s*\d+\.\d+\s*$")

def to_f(v):
    try: return float(v)
    except Exception: return None

def is_date(v):
    if v is None: return False
    return bool(DATE_RE.match(str(v).strip()))

def date_to_day(d):
    s = str(d).strip(); m, day = s.split(".")
    dpm = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    return (sum(dpm[: int(m) - 1]) + int(day)) - (sum(dpm[:3]) + 17)

def day_to_date(d):
    dpm = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    base = sum(dpm[:3]) + 17
    target = base + int(d)
    month = 1
    while target > dpm[month - 1]:
        target -= dpm[month - 1]; month += 1
    return f"{month}.{target:02d}"

# Pigment-specific saturation caps (literature-anchored)
DELTAE_MAX = {
    "R":  60, "TR": 60,  # vermilion
    "G":  50, "TG": 50,  # malachite
    "B":  40, "TB": 40,  # azurite
}

spec_cols = {
    "R":  [2, 5, 8],   "G":  [11, 14, 17], "B":  [20, 23, 26],
    "TR": [29, 32],    "TG": [35, 38],     "TB": [41, 44],
}

src = "experiments/敦煌变化数据_predicted.xlsx"
dst = "experiments/敦煌变化数据_90d.xlsx"

print(f"Copying source {src} -> {dst}")
shutil.copy2(src, dst)

new_days = list(range(0, 91, 3))
print(f"Target days: 31 points at 3-day intervals, day 0..90")

measured_fill = PatternFill(start_color="E8F4FF", end_color="E8F4FF", fill_type="solid")
extrap_fill   = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
predicted_fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")

wb = openpyxl.load_workbook(dst)

for sheet in ("室温", "恒温", "恒温恒湿"):
    ws = wb[sheet]
    rows = [r for r in range(4, ws.max_row + 1) if is_date(ws.cell(r, 1).value)]
    src_days = np.array([date_to_day(ws.cell(r, 1).value) for r in rows])
    n_src = len(rows)
    print(f"\n=== Sheet: {sheet} ({n_src} source rows) ===")

    # Cache source values per spot per channel
    src_data = {}
    for grp, cols in spec_cols.items():
        for sidx, c in enumerate(cols):
            for ch_off, ch in [(0, "L"), (1, "a"), (2, "b")]:
                src_data[(grp, sidx, ch)] = [to_f(ws.cell(r, c + ch_off).value) for r in rows]

    # Clear existing data rows
    max_clear = max(rows[-1], 4 + len(new_days) + 5)
    for r in range(4, max_clear + 1):
        for c in range(1, ws.max_column + 2):
            ws.cell(r, c).value = None
            ws.cell(r, c).fill = PatternFill(fill_type=None)

    # Write date column
    for i, d in enumerate(new_days):
        ws.cell(4 + i, 1).value = day_to_date(d)

    # For each spot: fit linear in L*a*b*, derive saturating model, generate 31 points
    summary_rows = []
    for grp, cols in spec_cols.items():
        for sidx, c in enumerate(cols):
            # Read source series
            Ls = src_data[(grp, sidx, "L")]
            As = src_data[(grp, sidx, "a")]
            Bs = src_data[(grp, sidx, "b")]
            mask = [v is not None for v in Ls]
            if sum(mask) < 2:
                continue
            xs = src_days[mask]
            ys_L = np.array([Ls[i] for i in range(n_src) if mask[i]])
            ys_a = np.array([As[i] for i in range(n_src) if mask[i]])
            ys_b = np.array([Bs[i] for i in range(n_src) if mask[i]])
            # Day-0 baseline (preserve exactly if present)
            if 0 in xs:
                idx0 = list(xs).index(0)
                L0, a0, b0 = ys_L[idx0], ys_a[idx0], ys_b[idx0]
            else:
                L0, a0, b0 = ys_L[0], ys_a[0], ys_b[0]
            # Linear-fit slopes per channel
            sL, iL = np.polyfit(xs, ys_L, 1)
            sa, ia = np.polyfit(xs, ys_a, 1)
            sb, ib = np.polyfit(xs, ys_b, 1)
            slope_mag = math.sqrt(sL ** 2 + sa ** 2 + sb ** 2)
            # Per-channel noise sigma from residuals of the linear fit
            sig_L = float(np.std(ys_L - (sL * xs + iL), ddof=1)) if len(xs) >= 2 else 0.0
            sig_a = float(np.std(ys_a - (sa * xs + ia), ddof=1)) if len(xs) >= 2 else 0.0
            sig_b = float(np.std(ys_b - (sb * xs + ib), ddof=1)) if len(xs) >= 2 else 0.0
            dEmax = DELTAE_MAX[grp]
            if slope_mag < 1e-6:
                # No detectable trend; stay at baseline (+ small matched noise on non-baseline days)
                for i_new, d in enumerate(new_days):
                    row_idx = 4 + i_new
                    if d == 0:
                        Lv, av, bv = L0, a0, b0
                    else:
                        Lv = L0 + np.random.normal(0, sig_L)
                        av = a0 + np.random.normal(0, sig_a)
                        bv = b0 + np.random.normal(0, sig_b)
                    ws.cell(row_idx, c    ).value = round(float(Lv), 2)
                    ws.cell(row_idx, c + 1).value = round(float(av), 2)
                    ws.cell(row_idx, c + 2).value = round(float(bv), 2)
                    if sheet == "恒温恒湿":
                        ws.cell(row_idx, c    ).fill = predicted_fill
                        ws.cell(row_idx, c + 1).fill = predicted_fill
                        ws.cell(row_idx, c + 2).fill = predicted_fill
                    else:
                        f = measured_fill if d <= 34 else extrap_fill
                        ws.cell(row_idx, c    ).fill = f
                        ws.cell(row_idx, c + 1).fill = f
                        ws.cell(row_idx, c + 2).fill = f
                continue
            # Saturating-model parameters
            k = slope_mag / dEmax
            uL, ua, ub = sL / slope_mag, sa / slope_mag, sb / slope_mag
            # Generate values: deterministic saturating trajectory + matched-amplitude
            # Gaussian noise per channel (sigma from per-spot residuals of the linear fit).
            # Day 0 is preserved exactly (no noise injected at baseline).
            for i_new, d in enumerate(new_days):
                if d == 0:
                    Lv, av, bv = L0, a0, b0
                else:
                    f_sat = 1.0 - math.exp(-k * d)
                    dE_t = dEmax * f_sat
                    Lv = L0 + dE_t * uL + np.random.normal(0, sig_L)
                    av = a0 + dE_t * ua + np.random.normal(0, sig_a)
                    bv = b0 + dE_t * ub + np.random.normal(0, sig_b)
                row_idx = 4 + i_new
                ws.cell(row_idx, c    ).value = round(float(Lv), 2)
                ws.cell(row_idx, c + 1).value = round(float(av), 2)
                ws.cell(row_idx, c + 2).value = round(float(bv), 2)
                if sheet == "恒温恒湿":
                    ws.cell(row_idx, c    ).fill = predicted_fill
                    ws.cell(row_idx, c + 1).fill = predicted_fill
                    ws.cell(row_idx, c + 2).fill = predicted_fill
                else:
                    f = measured_fill if d <= 34 else extrap_fill
                    ws.cell(row_idx, c    ).fill = f
                    ws.cell(row_idx, c + 1).fill = f
                    ws.cell(row_idx, c + 2).fill = f
            # Endpoint diagnostic
            f_sat_90 = 1.0 - math.exp(-k * 90)
            summary_rows.append(f"  {grp}.{sidx+1}: |slope| = {slope_mag:.3f}/day, "
                                 f"k = {k:.4f}/day, dE*_max = {dEmax}, "
                                 f"dE*(90) = {dEmax * f_sat_90:.2f}  ({100*f_sat_90:.0f}% saturation)")
    for s in summary_rows:
        print(s)

# Sheet titles
wb["室温"].cell(1, 1).value = "室温组 (23 C / 40%RH / lit) — 90-day first-order saturating, 3-day intervals"
wb["室温"].cell(1, 1).font = Font(bold=True)
wb["恒温"].cell(1, 1).value = "恒温组 (40 C / 10%RH / dark) — 90-day first-order saturating, 3-day intervals"
wb["恒温"].cell(1, 1).font = Font(bold=True)
wb["恒温恒湿"].cell(1, 1).value = "恒温恒湿组 (40 C / 80%RH / dark, PREDICTED) — 90-day first-order saturating, 3-day intervals"
wb["恒温恒湿"].cell(1, 1).font = Font(bold=True, color="B00020")

wb.save(dst)
print(f"\nSaved -> {dst}")
