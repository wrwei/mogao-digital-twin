"""Plot the current state of experiments/敦煌变化数据_predicted.xlsx (9 measured
time points over 34 days) per pigment x geometry, with all three arms
overlaid (room-T lit, chamber dark, predicted high-RH).

Reflects the latest denoising state of the workbook: HR.1/2/3 user-edited,
plus the Hampel filter catches in 恒温 and 室温.
"""
import openpyxl, math, re
import numpy as np
import matplotlib.pyplot as plt

DATE_RE = re.compile(r"^\s*\d+\.\d+\s*$")

def to_f(v):
    try: return float(v)
    except Exception: return float("nan")

def is_date(v):
    if v is None: return False
    return bool(DATE_RE.match(str(v).strip()))

def date_to_day(d):
    s = str(d).strip(); m, day = s.split(".")
    dpm = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    return (sum(dpm[: int(m) - 1]) + int(day)) - (sum(dpm[:3]) + 17)

spec_cols = {
    "R":  [2, 5, 8],   "G":  [11, 14, 17], "B":  [20, 23, 26],
    "TR": [29, 32],    "TG": [35, 38],     "TB": [41, 44],
}
PIG = {"R": "Vermilion", "G": "Malachite", "B": "Azurite",
       "TR": "Vermilion", "TG": "Malachite", "TB": "Azurite"}

src = "experiments/敦煌变化数据_predicted.xlsx"
wb = openpyxl.load_workbook(src, data_only=True)

# Load arms — only the measured/predicted rows (not the notes block)
def load_arm(sheet_name):
    ws = wb[sheet_name]
    rows = [r for r in range(4, ws.max_row + 1) if is_date(ws.cell(r, 1).value)]
    times = np.array([date_to_day(ws.cell(r, 1).value) for r in rows])
    per_spot = {}
    for grp, cols in spec_cols.items():
        for sidx, c in enumerate(cols):
            baseline = (to_f(ws.cell(rows[0], c).value),
                        to_f(ws.cell(rows[0], c + 1).value),
                        to_f(ws.cell(rows[0], c + 2).value))
            dE_series = []
            for r in rows:
                L = to_f(ws.cell(r, c).value)
                a = to_f(ws.cell(r, c + 1).value)
                b = to_f(ws.cell(r, c + 2).value)
                L0, a0, b0 = baseline
                if any(np.isnan([L, a, b, L0, a0, b0])):
                    dE_series.append(np.nan)
                else:
                    dE_series.append(math.sqrt((L - L0) ** 2 + (a - a0) ** 2 + (b - b0) ** 2))
            per_spot[(grp, sidx)] = (times, np.array(dE_series))
    return per_spot

arms = {
    "23 C / 40%RH (room, lit)":         load_arm("室温"),
    "40 C / 10%RH (chamber, dark)":     load_arm("恒温"),
    "40 C / 80%RH (predicted, dark)":   load_arm("恒温恒湿"),
}

styles = {
    "23 C / 40%RH (room, lit)":       {"color": "#2E86AB", "marker": "o", "ls": "-"},
    "40 C / 10%RH (chamber, dark)":   {"color": "#E63946", "marker": "s", "ls": "-"},
    "40 C / 80%RH (predicted, dark)": {"color": "#6A4C93", "marker": "^", "ls": "--"},
}

fig, axes = plt.subplots(2, 3, figsize=(13, 7.5), sharex=True)
geom_groups = [
    ("block (10x10x4 cm tile)",   ["TR", "TG", "TB"]),
    ("pedestal (1:5 lotus base)", ["R",  "G",  "B"]),
]

for row, (geom, grps) in enumerate(geom_groups):
    row_max = 0.0
    for grp in grps:
        cols = spec_cols[grp]
        for arm_label, per_spot in arms.items():
            for sidx in range(len(cols)):
                t, dE = per_spot[(grp, sidx)]
                m = np.nanmax(dE)
                if not np.isnan(m): row_max = max(row_max, m)
    for col, grp in enumerate(grps):
        ax = axes[row, col]
        cols = spec_cols[grp]
        for arm_label, per_spot in arms.items():
            st = styles[arm_label]
            # spot traces
            spot_curves = []
            for sidx in range(len(cols)):
                t, dE = per_spot[(grp, sidx)]
                spot_curves.append(dE)
                ax.plot(t, dE, color=st["color"], linestyle=st["ls"],
                        alpha=0.30, linewidth=0.9)
            stacked = np.array(spot_curves, dtype=float)
            mean = np.nanmean(stacked, axis=0)
            mn = np.nanmin(stacked, axis=0); mx = np.nanmax(stacked, axis=0)
            t = arms[arm_label][(grp, 0)][0]
            ax.fill_between(t, mn, mx, color=st["color"], alpha=0.12)
            ax.plot(t, mean, color=st["color"], linestyle=st["ls"],
                    marker=st["marker"], markersize=5, linewidth=2,
                    label=arm_label if (row == 0 and col == 0) else None)
        ax.set_title(f"{PIG[grp]} -- {geom}", fontsize=10)
        ax.grid(True, alpha=0.3)
        ax.set_xlim(-1, 36)
        ax.set_ylim(-0.2, row_max * 1.10 + 0.5)
        if col == 0: ax.set_ylabel(r"$\Delta E^*_{ab}$")
        if row == 1: ax.set_xlabel("Days since 17 Apr 2026")
        n_spots = len(cols)
        ax.text(0.97, 0.04, f"n = {n_spots} spots",
                transform=ax.transAxes, ha="right", va="bottom",
                fontsize=7, alpha=0.6, style="italic")

handles, labels = axes[0, 0].get_legend_handles_labels()
fig.legend(handles, labels, loc="upper center", ncol=3,
           frameon=False, fontsize=9.5, bbox_to_anchor=(0.5, 0.965))

fig.suptitle("Per-pigment dE* trajectories after denoising  (predicted.xlsx, 9 time points, 34 days)",
             y=0.99, fontsize=12, fontweight="bold")
plt.subplots_adjust(top=0.88, bottom=0.10, left=0.07, right=0.99, hspace=0.30, wspace=0.20)

out = "experiments/deltaE_trajectories_predicted_current.png"
plt.savefig(out, dpi=150, bbox_inches="tight")
print(f"Saved: {out}")
