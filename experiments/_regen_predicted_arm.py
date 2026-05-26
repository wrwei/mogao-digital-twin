"""Regenerate the 恒温恒湿 (predicted high-RH) sheet of
experiments/敦煌变化数据_predicted.xlsx from the now-smoothed 恒温 chamber data.

The original 恒温恒湿 sheet was RH-scaled from the pre-smoothing chamber
trajectory and therefore inherited the outliers we have since corrected
(TR.1 sign-flip, TG.2 a* drift, TB.2 b* drift, etc.). Re-scaling per
spot from the cleaned chamber values produces a smoother predicted arm
without needing a second-pass filter on the predicted sheet directly.

Method per spot, per channel:
  baseline (L0, a0, b0) = first row in chamber arm (day 0)
  for each row r:
      delta_chamber = chamber[r] - baseline
      predicted[r]  = baseline + 1.74 * delta_chamber
where 1.74 = (RH_target / RH_ref)^q = (80 / 40)^0.8, the
Paltakari--Karlsson sorption coupling at q = 0.8.

Cells in 恒温恒湿 are tinted pale purple as a visual reminder that they
are model output and not measurement.
"""
import openpyxl
import re
import shutil
from openpyxl.styles import PatternFill

DATE_RE = re.compile(r"^\s*\d+\.\d+\s*$")

def to_f(v):
    try:
        return float(v)
    except Exception:
        return None

def is_date(v):
    if v is None:
        return False
    return bool(DATE_RE.match(str(v).strip()))

src = "experiments/敦煌变化数据_predicted.xlsx"
bak = src + ".bak"

print(f"Backing up {src} -> {bak}")
shutil.copy2(src, bak)

wb = openpyxl.load_workbook(src)
ws_chamber = wb["恒温"]
ws_pred = wb["恒温恒湿"]

# RH scaling factor
q = 0.8
RH_target = 80
RH_ref = 40
scale = (RH_target / RH_ref) ** q
print(f"RH scaling factor (80/40)^0.8 = {scale:.3f}")

# Find data rows in the chamber sheet
rows = [r for r in range(4, ws_chamber.max_row + 1) if is_date(ws_chamber.cell(r, 1).value)]
print(f"Chamber data rows: {len(rows)}")

specimen_cols = list(range(2, ws_chamber.max_column + 1, 3))
predicted_fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")

# For each specimen column (a triplet L*, a*, b*) regenerate the predicted values
n_written = 0
for spec_col in specimen_cols:
    # Baseline at day 0 (first row)
    L0 = to_f(ws_chamber.cell(rows[0], spec_col).value)
    a0 = to_f(ws_chamber.cell(rows[0], spec_col + 1).value)
    b0 = to_f(ws_chamber.cell(rows[0], spec_col + 2).value)
    if L0 is None or a0 is None or b0 is None:
        # No baseline -- skip this spot
        continue
    # Write predicted values
    for r in rows:
        Lc = to_f(ws_chamber.cell(r, spec_col).value)
        ac = to_f(ws_chamber.cell(r, spec_col + 1).value)
        bc = to_f(ws_chamber.cell(r, spec_col + 2).value)
        if Lc is None or ac is None or bc is None:
            continue
        dL = Lc - L0
        da = ac - a0
        db = bc - b0
        Lp = L0 + dL * scale
        ap = a0 + da * scale
        bp = b0 + db * scale
        ws_pred.cell(r, spec_col).value = round(Lp, 2)
        ws_pred.cell(r, spec_col + 1).value = round(ap, 2)
        ws_pred.cell(r, spec_col + 2).value = round(bp, 2)
        ws_pred.cell(r, spec_col).fill = predicted_fill
        ws_pred.cell(r, spec_col + 1).fill = predicted_fill
        ws_pred.cell(r, spec_col + 2).fill = predicted_fill
        n_written += 1

print(f"Predicted cells written: {n_written} (across {len(specimen_cols)} spots and {len(rows)} dates)")

wb.save(src)
print(f"Saved -> {src}")
print(f"(Backup at {bak} if anything looks wrong.)")
