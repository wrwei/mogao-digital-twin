"""Reorganise experiments/敦煌变化数据_predicted.xlsx into a structure that
is more legible to a reader who is not already familiar with the dataset.

New workbook layout:
  1. README                       --  documentation, legend, conditions,
                                       pigments, specimen codes, colour key
  2. Measurements (long format)   --  one row per L*a*b* measurement; columns
                                       include human-readable pigment names,
                                       condition labels, ΔE*_{ab} pre-computed
  3. dE_summary (wide pivot)      --  ΔE*_{ab} per spot per timepoint per arm
  4. Room_T_measured_raw          --  original measured room-T arm (wide)
  5. Chamber_T_measured_raw       --  original measured chamber arm (wide)
  6. 80pct_RH_predicted_raw       --  RH-scaled predicted arm (wide)

Output: experiments/敦煌变化数据_predicted_reorganised.xlsx
"""
import openpyxl
import math
import re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

src_path = "experiments/敦煌变化数据_predicted.xlsx"
dst_path = "experiments/敦煌变化数据_predicted_reorganised.xlsx"

DATE_RE = re.compile(r"^\s*\d+\.\d+\s*$")

def to_f(v):
    try: return float(v)
    except Exception: return None

def is_date(v):
    if v is None: return False
    return bool(DATE_RE.match(str(v).strip()))

def date_to_day(d):
    s = str(d).strip(); m, day = s.split(".")
    dpm = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    return (sum(dpm[: int(m) - 1]) + int(day)) - (sum(dpm[:3]) + 17)

# Specimen layout (matches original columns):
#   R/G/B  = pedestal specimens (vermilion/malachite/azurite), 3 spots each
#   TR/TG/TB = block-tile specimens (vermilion/malachite/azurite), 2 spots each
#   ('T' = Tile = block; no-T = pedestal/lotus base.)
specimen_cols = {
    "R":  [2, 5, 8],   "G":  [11, 14, 17], "B":  [20, 23, 26],
    "TR": [29, 32],    "TG": [35, 38],     "TB": [41, 44],
}
group_meta = {
    "R":  ("vermilion", "朱砂", "HgS",                  "pedestal (1:5 lotus base)"),
    "G":  ("malachite", "石绿", "Cu₂CO₃(OH)₂",           "pedestal (1:5 lotus base)"),
    "B":  ("azurite",   "石青", "Cu₃(CO₃)₂(OH)₂",        "pedestal (1:5 lotus base)"),
    "TR": ("vermilion", "朱砂", "HgS",                  "block (10×10×4 cm tile)"),
    "TG": ("malachite", "石绿", "Cu₂CO₃(OH)₂",           "block (10×10×4 cm tile)"),
    "TB": ("azurite",   "石青", "Cu₃(CO₃)₂(OH)₂",        "block (10×10×4 cm tile)"),
}
sheet_meta = {
    "室温":     ("Room temperature / lit",      "23 °C",        "40 %",   "ambient laboratory light", "measured"),
    "恒温":     ("Chamber / dark",              "40 °C",        "10 %",   "dark (sealed oven)",        "measured"),
    "恒温恒湿": ("Predicted high-RH chamber",   "40 °C",        "80 %",   "dark (predicted)",          "predicted"),
}
sheet_short = {"室温": "Room_T_measured_raw", "恒温": "Chamber_T_measured_raw", "恒温恒湿": "80pct_RH_predicted_raw"}

KNOWN_OUTLIERS = {
    ("室温", "R", 0, "5.12"),    # all 3 channels off-scale (instrument misfire)
    ("恒温", "TR", 0, "5.14"),   # a* sign-flip
}

# --- Load source workbook ---
wb_src = openpyxl.load_workbook(src_path, data_only=True)

# Build tidy long-format list of measurements
tidy = []
for sheet_name in ("室温", "恒温", "恒温恒湿"):
    if sheet_name not in wb_src.sheetnames:
        continue
    ws = wb_src[sheet_name]
    cond_label, T_str, RH_str, light_str, meas_kind = sheet_meta[sheet_name]
    rows_with_date = [(r, ws.cell(r, 1).value) for r in range(4, ws.max_row + 1) if is_date(ws.cell(r, 1).value)]
    # baselines per spot (first measured row, day 0)
    baselines = {}
    for grp, cols in specimen_cols.items():
        for sidx, c in enumerate(cols):
            if rows_with_date:
                r0 = rows_with_date[0][0]
                L0 = to_f(ws.cell(r0, c).value)
                a0 = to_f(ws.cell(r0, c + 1).value)
                b0 = to_f(ws.cell(r0, c + 2).value)
                baselines[(grp, sidx)] = (L0, a0, b0)
    for r, date_val in rows_with_date:
        date_str = str(date_val).strip()
        day = date_to_day(date_val)
        for grp, cols in specimen_cols.items():
            pigment, pinyin_zh, formula, geometry = group_meta[grp]
            for sidx, c in enumerate(cols):
                L = to_f(ws.cell(r, c).value)
                a = to_f(ws.cell(r, c + 1).value)
                b = to_f(ws.cell(r, c + 2).value)
                # Outlier detection
                is_outlier = (sheet_name, grp, sidx, date_str) in KNOWN_OUTLIERS
                if is_outlier:
                    status = "outlier_masked"
                    dE = None
                elif L is None or a is None or b is None:
                    status = "missing"
                    dE = None
                else:
                    status = meas_kind
                    L0, a0, b0 = baselines.get((grp, sidx), (None, None, None))
                    if L0 is not None and a0 is not None and b0 is not None:
                        dE = math.sqrt((L - L0) ** 2 + (a - a0) ** 2 + (b - b0) ** 2)
                    else:
                        dE = None
                tidy.append({
                    "date": date_str,
                    "day": day,
                    "condition": cond_label,
                    "temperature_C": T_str,
                    "relative_humidity_pct": RH_str,
                    "light_exposure": light_str,
                    "pigment": pigment,
                    "pigment_zh": pinyin_zh,
                    "pigment_formula": formula,
                    "geometry": geometry,
                    "spot": sidx + 1,
                    "spot_code": f"{grp}.{sidx + 1}",
                    "status": status,
                    "L_star": L,
                    "a_star": a,
                    "b_star": b,
                    "delta_E_vs_day0": round(dE, 3) if dE is not None else None,
                })

# --- Build new workbook ---
wb_new = openpyxl.Workbook()

# Style palette
header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
subhead_fill = PatternFill(start_color="E8F4FF", end_color="E8F4FF", fill_type="solid")
measured_fill = PatternFill(start_color="E8F4FF", end_color="E8F4FF", fill_type="solid")
predicted_fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
outlier_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
note_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
title_font = Font(bold=True, size=14)
section_font = Font(bold=True, size=11, color="2E86AB")
thin = Side(border_style="thin", color="C8C8C8")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ---------- Sheet 1: README ----------
ws = wb_new.active
ws.title = "README"
ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 60

row = 1
ws.cell(row, 1).value = "Mogao polychrome accelerated-ageing pilot — reorganised dataset"
ws.cell(row, 1).font = title_font
row += 2

def section(title):
    global row
    ws.cell(row, 1).value = title
    ws.cell(row, 1).font = section_font
    row += 1

def kv(k, v):
    global row
    ws.cell(row, 1).value = k
    ws.cell(row, 1).font = Font(bold=True)
    ws.cell(row, 2).value = v
    ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
    row += 1

def bullet(s):
    global row
    ws.cell(row, 2).value = "• " + s
    ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
    row += 1

section("Source and scope")
kv("Project", "Mogao polychrome digital twin — heritage-science accelerated-ageing pilot")
kv("Measurement campaign", "9 time points spanning 34 days (17 Apr → 21 May 2026)")
kv("Specimens", "Replica clay+ground+pigment samples; vermilion / malachite / azurite on two geometries (10×10×4 cm tile, 1:5 lotus pedestal)")
kv("Source data", "experiments/敦煌变化数据(1).xlsx (room and chamber arms only)")
kv("Predicted arm", "40 °C / 80 % RH high-humidity chamber; computed by Paltakari–Karlsson RH-scaling (q = 0.8) of the measured 40 °C / 40 % RH chamber trajectory per spot")
row += 1

section("Sheets in this workbook")
bullet("README — this sheet")
bullet("Measurements — long/tidy format, one row per (date, specimen, spot, channel-triplet). Filter or pivot to taste.")
bullet("dE_summary — wide pivot showing ΔE*_ab vs day-0 baseline per spot per date per arm.")
bullet("Room_T_measured_raw — original measured 23 °C / 40 % RH (lit) arm, wide format.")
bullet("Chamber_T_measured_raw — original measured 40 °C / 10 % RH (dark) arm, wide format.")
bullet("80pct_RH_predicted_raw — RH-scaled prediction at 40 °C / 80 % RH (dark), wide format.")
row += 1

section("Environmental conditions")
ws.cell(row, 1).value = "Arm"; ws.cell(row, 1).font = Font(bold=True)
ws.cell(row, 2).value = "Description"; ws.cell(row, 2).font = Font(bold=True)
row += 1
kv("Room T / lit",     "23 °C, 40 % RH, ambient laboratory light. Measured.")
kv("Chamber / dark",   "40 °C, 10 % RH (sealed oven dries on heating from room conditions), dark. Measured.")
kv("80 %-RH prediction", "40 °C, 80 % RH, dark. Predicted by RH-scaling the measured chamber trajectory using Paltakari–Karlsson humidity coupling q = 0.8 — i.e. ΔL*, Δa*, Δb* multiplied by (80/40)^0.8 = 1.74 per spot.")
row += 1

section("Pigments")
kv("Vermilion (朱砂)",  "HgS, photolytically transforms to meta-cinnabar (β-HgS) under light. Most light-sensitive of the canonical Mogao palette.")
kv("Malachite (石绿)",  "Cu₂CO₃(OH)₂, basic copper carbonate. Greening / desaturation under hygrothermal stress.")
kv("Azurite (石青)",    "Cu₃(CO₃)₂(OH)₂, copper carbonate. Greens to malachite, then desaturates further.")
row += 1

section("Specimen codes (legacy wide-format sheets)")
kv("R.1, R.2, R.3",     "3 templated spots on the vermilion 1:5 lotus pedestal")
kv("G.1, G.2, G.3",     "3 spots on the malachite 1:5 lotus pedestal")
kv("B.1, B.2, B.3",     "3 spots on the azurite 1:5 lotus pedestal")
kv("TR.1, TR.2",        "2 spots on the vermilion block tile")
kv("TG.1, TG.2",        "2 spots on the malachite block tile")
kv("TB.1, TB.2",        "2 spots on the azurite block tile")
row += 1

section("Status flags (in the Measurements sheet)")
kv("measured",           "Directly measured by the colorimeter on the specimen surface.")
kv("predicted",          "Computed by RH-scaling the measured chamber arm; this is model output, not measurement.")
kv("outlier_masked",     "Cell value omitted: identified as an instrument/transcription anomaly during analysis (Room-T R.1 @ 5.12 — all three channels off-scale; Chamber TR.1 @ 5.14 — a* sign-flip).")
kv("missing",            "No value supplied in the source workbook.")
row += 1

section("Cell colour key (raw wide-format sheets)")
ws.cell(row, 2).value = "Pale blue — measured"; ws.cell(row, 2).fill = measured_fill; row += 1
ws.cell(row, 2).value = "Pale yellow — extrapolated (rate-fit, future date)"; ws.cell(row, 2).fill = note_fill; row += 1
ws.cell(row, 2).value = "Pale purple — predicted (RH-scaled from chamber)"; ws.cell(row, 2).fill = predicted_fill; row += 1
ws.cell(row, 2).value = "Pale red — outlier masked"; ws.cell(row, 2).fill = outlier_fill; row += 1
row += 1

section("Reproducibility")
kv("Analysis scripts",   "experiments/_extend_180d.py (180-day extrapolation), experiments/_plot_180d.py, experiments/_plot_180d_firstorder.py (figures), experiments/_validation.py (time-window CV + bootstrap E_a CIs)")
kv("RH coupling",        "Paltakari–Karlsson humidity exponent q = 0.8 borrowed from cellulose literature (Strlič / Paltakari–Karlsson). Not measured for this material system.")
kv("Repository",         "github.com/wrwei/mogao-digital-twin")

# ---------- Sheet 2: Measurements (long/tidy format) ----------
ws2 = wb_new.create_sheet("Measurements")
headers = [
    "date", "day", "condition", "temperature_C", "relative_humidity_pct",
    "light_exposure", "pigment", "pigment_zh", "pigment_formula", "geometry",
    "spot", "spot_code", "status", "L_star", "a_star", "b_star", "delta_E_vs_day0",
]
for j, h in enumerate(headers, start=1):
    c = ws2.cell(1, j)
    c.value = h
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center")
for i, rec in enumerate(tidy, start=2):
    for j, h in enumerate(headers, start=1):
        c = ws2.cell(i, j)
        c.value = rec[h]
        if rec["status"] == "predicted":
            c.fill = predicted_fill
        elif rec["status"] == "outlier_masked":
            c.fill = outlier_fill
        elif rec["status"] == "measured":
            c.fill = measured_fill

# Column widths
widths = [8, 6, 22, 14, 22, 28, 12, 12, 18, 26, 6, 10, 14, 9, 9, 9, 14]
for j, w in enumerate(widths, start=1):
    ws2.column_dimensions[get_column_letter(j)].width = w
ws2.freeze_panes = "A2"

# AutoFilter for the long-format sheet
ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{1 + len(tidy)}"

# ---------- Sheet 3: dE_summary (wide pivot) ----------
ws3 = wb_new.create_sheet("dE_summary")
ws3.cell(1, 1).value = "dE*_ab per spot vs day-0 baseline (each spot's own first measurement)"
ws3.cell(1, 1).font = Font(bold=True, italic=True)

# Collect unique dates in order (use chamber sheet ordering for canonical sequence)
canonical_dates = []
for rec in tidy:
    d = rec["date"]
    if d not in canonical_dates and rec["condition"] == "Chamber / dark":
        canonical_dates.append(d)

# Build the pivot: rows = (condition, pigment, geometry, spot), cols = dates
condition_order = ["Room temperature / lit", "Chamber / dark", "Predicted high-RH chamber"]
pivot_rows = []
seen = set()
for rec in tidy:
    key = (rec["condition"], rec["pigment"], rec["geometry"], rec["spot_code"])
    if key not in seen:
        seen.add(key); pivot_rows.append(key)
pivot_rows.sort(key=lambda x: (condition_order.index(x[0]) if x[0] in condition_order else 99,
                                ["vermilion", "malachite", "azurite"].index(x[1]),
                                x[2], x[3]))

# Header
hdrs = ["condition", "pigment", "geometry", "spot"] + [f"day {date_to_day(d)} ({d})" for d in canonical_dates]
for j, h in enumerate(hdrs, start=1):
    c = ws3.cell(3, j)
    c.value = h
    c.font = header_font; c.fill = header_fill
    c.alignment = Alignment(horizontal="center")
for j in range(1, len(hdrs) + 1):
    ws3.column_dimensions[get_column_letter(j)].width = 14 if j > 4 else (24 if j == 1 else 12)

# Build a lookup
lookup = {(r["condition"], r["pigment"], r["geometry"], r["spot_code"], r["date"]): r["delta_E_vs_day0"] for r in tidy}

for i, (cond, pig, geom, spot_code) in enumerate(pivot_rows, start=4):
    ws3.cell(i, 1).value = cond
    ws3.cell(i, 2).value = pig
    ws3.cell(i, 3).value = geom
    ws3.cell(i, 4).value = spot_code
    if cond == "Predicted high-RH chamber":
        fill = predicted_fill
    elif cond == "Chamber / dark":
        fill = measured_fill
    else:
        fill = measured_fill
    for j, d in enumerate(canonical_dates, start=5):
        v = lookup.get((cond, pig, geom, spot_code, d))
        ws3.cell(i, j).value = v
        ws3.cell(i, j).fill = fill
        if isinstance(v, (int, float)):
            ws3.cell(i, j).number_format = "0.00"

ws3.freeze_panes = "E4"

# ---------- Sheets 4-6: copy raw wide-format sheets with friendlier names ----------
from copy import copy as _copy
def copy_sheet(src_ws, dst_wb, dst_title):
    dst = dst_wb.create_sheet(dst_title)
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = _copy(cell.font)
                new_cell.fill = _copy(cell.fill)
                new_cell.border = _copy(cell.border)
                new_cell.alignment = _copy(cell.alignment)
                new_cell.number_format = cell.number_format
    for col_letter, dim in src_ws.column_dimensions.items():
        dst.column_dimensions[col_letter].width = dim.width
    return dst

for src_name, dst_name in sheet_short.items():
    if src_name in wb_src.sheetnames:
        # Reload with styles
        wb_src_styled = openpyxl.load_workbook(src_path)
        copy_sheet(wb_src_styled[src_name], wb_new, dst_name)

wb_new.save(dst_path)
print(f"Saved reorganised workbook → {dst_path}")
print(f"Sheets: {wb_new.sheetnames}")
print(f"Measurements rows: {len(tidy) + 1} (incl. header)")
