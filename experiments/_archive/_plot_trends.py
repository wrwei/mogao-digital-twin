"""Trend view: room-temperature (23C/40%RH) vs 40C (40C/10%RH) only.

For each pigment x geometry panel, pool the spots of each arm, show the faint
per-spot / mean data, and overlay a bold least-squares trend line with its
60-day ΔE slope annotated -- so the two temperature conditions can be compared
directly. Reads the model-faithful workbook.
"""
import openpyxl, re
import numpy as np
import matplotlib.pyplot as plt
import math
# Per-point ΔE measurement uncertainty: each marker is the mean of n spots at
# σ=0.7/channel, so the 1σ error on the mean is σ/√n.
SIGMA_MEAS = 0.7

DATE_RE = re.compile(r"^\s*\d+\.\d+\s*$")
def to_f(v):
    try: return float(v)
    except Exception: return float("nan")
def is_date(v):
    return v is not None and bool(DATE_RE.match(str(v).strip()))
def date_to_day(d):
    m, day = str(d).strip().split("."); dpm = [31,28,31,30,31,30,31,31,30,31,30,31]
    return (sum(dpm[:int(m)-1]) + int(day)) - (sum(dpm[:3]) + 17)

spec_cols = {"R":[2,5,8], "G":[11,14,17], "B":[20,23,26],
             "TR":[29,32], "TG":[35,38], "TB":[41,44]}
PIG = {"R":"Vermilion","G":"Malachite","B":"Azurite",
       "TR":"Vermilion","TG":"Malachite","TB":"Azurite"}

wb = openpyxl.load_workbook("experiments/mogao_data_model_faithful.xlsx", data_only=True)

def load_arm(sheet):
    ws = wb[sheet]
    rows = [r for r in range(1, ws.max_row+1) if is_date(ws.cell(r,1).value)]
    t = np.array([date_to_day(ws.cell(r,1).value) for r in rows], float)
    per = {}
    for grp, cols in spec_cols.items():
        curves = []
        for c in cols:
            L = np.array([to_f(ws.cell(r,c).value)   for r in rows])
            a = np.array([to_f(ws.cell(r,c+1).value) for r in rows])
            b = np.array([to_f(ws.cell(r,c+2).value) for r in rows])
            curves.append(np.sqrt((L-L[0])**2 + (a-a[0])**2 + (b-b[0])**2))
        per[grp] = (t, np.array(curves))
    return per

arms = {
    "23 C / 40%RH  (room temperature, lit)": ("室温", "#2E86AB", "o"),
    "40 C / 10%RH  (chamber, dark)":         ("恒温", "#E63946", "s"),
}
data = {label: load_arm(sh) for label,(sh,_,_) in arms.items()}

fig, axes = plt.subplots(2, 3, figsize=(14, 7.6), sharex=True)
geom = [("block (10x10x4 cm tile)", ["TR","TG","TB"]),
        ("pedestal (1:5 lotus base)", ["R","G","B"])]

for row, (gname, grps) in enumerate(geom):
    row_max = 0.0
    for grp in grps:
        for label in arms:
            mx = np.nanmax(data[label][grp][1])
            if not np.isnan(mx): row_max = max(row_max, mx)
    for col, grp in enumerate(grps):
        ax = axes[row, col]
        notes = []
        for label,(sh,color,mk) in arms.items():
            t, curves = data[label][grp]
            mean = np.nanmean(curves, 0)
            sem = SIGMA_MEAS / math.sqrt(curves.shape[0])   # 1σ measurement error on mean
            # faint per-spot + mean markers with per-point measurement error bars
            for c in curves:
                ax.plot(t, c, color=color, alpha=0.14, lw=0.6)
            ax.errorbar(t, mean, yerr=sem, color=color, marker=mk, ms=3.5, lw=0,
                        elinewidth=0.8, capsize=2.0, capthick=0.8, alpha=0.7)
            # bold least-squares trend over pooled spot points
            tt = np.tile(t, curves.shape[0]); yy = curves.ravel()
            ok = ~np.isnan(yy)
            slope, icpt = np.polyfit(tt[ok], yy[ok], 1)
            xline = np.array([0, t.max()])
            ax.plot(xline, slope*xline + icpt, color=color, lw=2.4, ls="-",
                    label=label if (row==0 and col==0) else None)
            notes.append((color, f"{slope*60:+.1f} ΔE / 60 d"))
        ax.set_title(f"{PIG[grp]} -- {gname}", fontsize=10)
        ax.grid(True, alpha=0.3); ax.set_xlim(-2, 64)
        ax.set_ylim(-0.3, row_max*1.10 + 0.5)
        if col == 0: ax.set_ylabel(r"$\Delta E^*_{ab}$")
        if row == 1: ax.set_xlabel("Days since 17 Apr 2026")
        # slope annotations (trend magnitude over the 60-day window)
        for i,(color,txt) in enumerate(notes):
            ax.text(0.03, 0.94 - i*0.09, txt, transform=ax.transAxes, ha="left",
                    va="top", fontsize=8, color=color, fontweight="bold")

handles, labels = axes[0,0].get_legend_handles_labels()
fig.legend(handles, labels, loc="upper center", ncol=2, frameon=False,
           fontsize=9, bbox_to_anchor=(0.5, 0.945))
fig.suptitle("Colour-change trends: room-temperature vs 40 C  (bold = least-squares trend, faint = data)",
             y=0.99, fontsize=12, fontweight="bold")
fig.text(0.5, 0.02,
         "Trend = OLS fit of ΔE vs time over all spots of each arm; annotation is the fitted ΔE change across the 60-day window.  "
         "Error bars = 1σ measurement uncertainty (σ/√n, σ=0.7 ΔE/channel).  Rates from per-pigment Arrhenius Ea (_validation.json) × Paltakari + validated vermilion photolytic pathway.",
         ha="center", va="bottom", fontsize=7.5, style="italic", color="#555555", wrap=True)
plt.subplots_adjust(top=0.85, bottom=0.11, left=0.07, right=0.99, hspace=0.28, wspace=0.20)

out = "experiments/deltaE_trends_roomtemp_vs_40C.png"
plt.savefig(out, dpi=150, bbox_inches="tight")
print(f"Saved: {out}")

# console summary
print("\n60-day trend slopes (ΔE over window):")
for grp in spec_cols:
    line = f"  {PIG[grp]:10s} {'block' if grp in ('R','G','B') else 'pedestal':9s}"
    for label,(sh,_,_) in arms.items():
        t, curves = data[label][grp]
        tt = np.tile(t, curves.shape[0]); yy = curves.ravel(); ok = ~np.isnan(yy)
        s = np.polyfit(tt[ok], yy[ok], 1)[0]
        tag = "room" if "room" in label else "40C "
        line += f" | {tag} {s*60:+5.1f}"
    print(line)
