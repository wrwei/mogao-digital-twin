"""Apply the same binomial 1-2-1 smoother to the vermilion spots in the
恒温恒湿 (predicted high-RH) sheet, matching the malachite + azurite pass
done by _smooth_predicted_mg_az.py.

Targets:
  R.1, R.2, R.3   (vermilion block)
  TR.1, TR.2       (vermilion pedestal)

Per spot, per channel (L*, a*, b*):
  smoothed[i] = (series[i-1] + 2*series[i] + series[i+1]) / 4
  endpoints preserved.

Vermilion is the only remaining unsmoothed pigment in 恒温恒湿; after
this pass all three pigments in both geometries should look the same
visual smoothness.
"""
import openpyxl
import re
import shutil
from openpyxl.styles import PatternFill

DATE_RE = re.compile(r"^\s*\d+\.\d+\s*$")

def to_f(v):
    try:
        return float(v)
    except Exception:
        return None

def is_date(v):
    if v is None:
        return False
    return bool(DATE_RE.match(str(v).strip()))

src = "experiments/敦煌变化数据_predicted.xlsx"
bak = src + ".bak"

print(f"Backing up {src} -> {bak}")
shutil.copy2(src, bak)

TARGETS = [
    ("R",  [2, 5, 8]),    # vermilion block
    ("TR", [29, 32]),     # vermilion pedestal
]

wb = openpyxl.load_workbook(src)
ws = wb["恒温恒湿"]
rows = [r for r in range(4, ws.max_row + 1) if is_date(ws.cell(r, 1).value)]
print(f"Data rows: {len(rows)}")

binomial_fill = PatternFill(start_color="E1BEE7", end_color="E1BEE7", fill_type="solid")

def binomial_smooth(series):
    n = len(series)
    out = list(series)
    for i in range(1, n - 1):
        a, b, c = series[i - 1], series[i], series[i + 1]
        if a is None or b is None or c is None:
            continue
        out[i] = round((a + 2 * b + c) / 4.0, 2)
    return out

print()
print(f"{'Group':6s} {'Spot':5s} {'Chan':5s} {'Changes (row -> new val)':60s}")
print("-" * 90)

total_changes = 0
for grp, cols in TARGETS:
    for sidx, c in enumerate(cols):
        for ch_off, ch_name in [(0, "L*"), (1, "a*"), (2, "b*")]:
            series = [to_f(ws.cell(r, c + ch_off).value) for r in rows]
            smoothed = binomial_smooth(series)
            changes = []
            for i in range(len(rows)):
                if series[i] is None or smoothed[i] is None:
                    continue
                if abs(smoothed[i] - series[i]) >= 0.005:
                    r = rows[i]
                    ws.cell(r, c + ch_off).value = smoothed[i]
                    ws.cell(r, c + ch_off).fill = binomial_fill
                    changes.append(f"r{r} {series[i]:.2f}->{smoothed[i]:.2f}")
            if changes:
                total_changes += len(changes)
                print(f"{grp:6s} {sidx + 1:>5d} {ch_name:5s} {', '.join(changes)}")

print()
print(f"Total cells smoothed: {total_changes}")
wb.save(src)
print(f"Saved -> {src}")
print(f"(Backup at {bak} if anything looks wrong.)")
