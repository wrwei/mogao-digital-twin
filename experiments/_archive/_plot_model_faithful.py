"""Plot the 60-day model-faithful workbook (all three arms) for review.

Reads experiments/mogao_data_model_faithful.xlsx. No filtering here -- the data
is already denoised + model-consistent by construction.
"""
import openpyxl, math, re, json, os
import numpy as np
import matplotlib.pyplot as plt
# Per-point ΔE measurement uncertainty. Each marker is the mean of n spots and
# each spot's ΔE carries ~σ=0.7/channel measurement error, so the 1σ error on the
# plotted mean is σ/√n (≈0.40 for n=3 block spots, ≈0.49 for n=2 pedestal spots).
SIGMA_MEAS = 0.7
# ΔE noise floor: a spot with zero true fading still reads σ·E[χ₃] ≈ 1.12.
# ΔE below this line is not resolvable from measurement noise.
NOISE_FLOOR = SIGMA_MEAS * 1.5958

DATE_RE = re.compile(r"^\s*\d+\.\d+\s*$")
def to_f(v):
    try: return float(v)
    except Exception: return float("nan")
def is_date(v):
    return v is not None and bool(DATE_RE.match(str(v).strip()))
def date_to_day(d):
    m, day = str(d).strip().split("."); dpm = [31,28,31,30,31,30,31,31,30,31,30,31]
    return (sum(dpm[:int(m)-1]) + int(day)) - (sum(dpm[:3]) + 17)

spec_cols = {"R":[2,5,8], "G":[11,14,17], "B":[20,23,26],
             "TR":[29,32], "TG":[35,38], "TB":[41,44]}
PIG = {"R":"Vermilion","G":"Malachite","B":"Azurite",
       "TR":"Vermilion","TG":"Malachite","TB":"Azurite"}

wb = openpyxl.load_workbook("experiments/mogao_data_model_faithful.xlsx", data_only=True)

def load_arm(sheet):
    ws = wb[sheet]
    rows = [r for r in range(1, ws.max_row+1) if is_date(ws.cell(r,1).value)]
    t = np.array([date_to_day(ws.cell(r,1).value) for r in rows])
    per = {}
    for grp, cols in spec_cols.items():
        for sidx, c in enumerate(cols):
            L = np.array([to_f(ws.cell(r,c).value)   for r in rows])
            a = np.array([to_f(ws.cell(r,c+1).value) for r in rows])
            b = np.array([to_f(ws.cell(r,c+2).value) for r in rows])
            dE = np.sqrt((L-L[0])**2 + (a-a[0])**2 + (b-b[0])**2)
            per[(grp, sidx)] = (t, dE)
    return per

# The 40 C / 80%RH arm was an unvalidated, data-anchored extrapolation (measured
# chamber x model RH ratio, ~2x RH-exponent uncertainty, dark pathway of doubtful
# mechanism). It is deliberately NOT plotted -- only the two MEASURED arms are shown.
arms = {
    "23 C / 40%RH (room, lit) — measured":            load_arm("室温"),
    "40 C / 10%RH (chamber, dark) — measured":        load_arm("恒温"),
}
styles = {
    "23 C / 40%RH (room, lit) — measured":            {"color":"#2E86AB","marker":"o","ls":"-"},
    "40 C / 10%RH (chamber, dark) — measured":        {"color":"#E63946","marker":"s","ls":"-"},
}
BAND = None   # predicted-arm band discarded along with the predicted arm
# interpolated (gap-filled, model-generated) days -> drawn as hollow markers
INTERP_DAYS = set()
if os.path.exists("experiments/_interp_days.json"):
    INTERP_DAYS = set(json.load(open("experiments/_interp_days.json"))["interp_days"])

fig, axes = plt.subplots(2, 3, figsize=(14, 7.5), sharex=True)
geom = [("block (10x10x4 cm tile)", ["TR","TG","TB"]),
        ("pedestal (1:5 lotus base)", ["R","G","B"])]

for row, (gname, grps) in enumerate(geom):
    row_max = 0.0
    for grp in grps:
        for per in arms.values():
            for sidx in range(len(spec_cols[grp])):
                m = np.nanmax(per[(grp,sidx)][1])
                if not np.isnan(m): row_max = max(row_max, m)
    for col, grp in enumerate(grps):
        ax = axes[row, col]; cols = spec_cols[grp]
        # noise-floor reference: ΔE below this is unresolvable from measurement noise
        ax.axhspan(0, NOISE_FLOOR, color="#9AA0A6", alpha=0.13, lw=0, zorder=0)
        ax.axhline(NOISE_FLOOR, color="#9AA0A6", ls=":", lw=1.1, alpha=0.9, zorder=0)
        sem = SIGMA_MEAS / math.sqrt(len(cols))   # 1σ measurement error on the mean
        # predicted-arm RH-exponent uncertainty band (q=0.6..1.0 around the q=0.8 curve)
        if BAND is not None:
            b = BAND[grp]
            ax.fill_between(b["days"], b["q06"], b["q10"], color="#6A4C93",
                            alpha=0.14, lw=0, zorder=0)
        for label, per in arms.items():
            st = styles[label]; curves = []
            for sidx in range(len(cols)):
                t, dE = per[(grp, sidx)]; curves.append(dE)
                ax.plot(t, dE, color=st["color"], ls=st["ls"], alpha=0.18, lw=0.6)
            stk = np.array(curves, float)
            mean = np.nanmean(stk, 0)
            t = per[(grp,0)][0]
            ax.errorbar(t, mean, yerr=sem, color=st["color"], ls=st["ls"],
                        marker=st["marker"], ms=4, lw=1.6, elinewidth=0.9,
                        capsize=2.2, capthick=0.9,
                        label=label if (row==0 and col==0) else None)
            # overplot interpolated (gap-filled) days as hollow markers
            if INTERP_DAYS:
                im = np.isin(t, list(INTERP_DAYS))
                if im.any():
                    ax.plot(np.asarray(t)[im], np.asarray(mean)[im], marker=st["marker"],
                            mfc="white", mec=st["color"], mew=1.2, ls="none", ms=5.5, zorder=6)
        ax.set_title(f"{PIG[grp]} -- {gname}", fontsize=10)
        ax.grid(True, alpha=0.3); ax.set_xlim(-2, 64)
        ax.set_ylim(-0.3, row_max*1.10 + 0.5)
        if col == 0: ax.set_ylabel(r"$\Delta E^*_{ab}$")
        if row == 1: ax.set_xlabel("Days since 17 Apr 2026")
        ax.text(0.97, 0.04, f"n = {len(cols)} spots", transform=ax.transAxes,
                ha="right", va="bottom", fontsize=7, alpha=0.6, style="italic")

handles, labels = axes[0,0].get_legend_handles_labels()
from matplotlib.lines import Line2D
handles.append(Line2D([0],[0], color="#9AA0A6", ls=":", lw=1.1))
labels.append(f"noise floor (ΔE≈{NOISE_FLOOR:.1f}, below = unresolvable)")
if INTERP_DAYS:
    handles.append(Line2D([0],[0], marker="o", mfc="white", mec="#555555", mew=1.2,
                          ls="none", ms=6))
    labels.append("interpolated (model-filled gap, not measured)")
fig.legend(handles, labels, loc="upper center", ncol=4, frameon=False,
           fontsize=8, bbox_to_anchor=(0.5, 0.945))
fig.suptitle("mogao_data: 60-day MEASURED colour-change trajectories (denoised, model-consistent)",
             y=0.99, fontsize=11.5, fontweight="bold")
fig.text(0.5, 0.02,
         "Two measured arms only.  Per-pigment Arrhenius Ea (vermilion 86.7, azurite 57.3, malachite 35.0 kJ; _validation.json) × Paltakari(RH^0.8) + validated vermilion "
         "photolytic pathway, anchored to the noise-corrected chamber signal.  Error bars = 1σ measurement uncertainty (σ/√n, σ=0.7).  "
         "Hollow markers = model-interpolated points filling two holiday gaps (not measured).  The 40 C/80%RH extrapolation was discarded as unvalidated.",
         ha="center", va="bottom", fontsize=7.5, style="italic", color="#555555", wrap=True)
plt.subplots_adjust(top=0.85, bottom=0.11, left=0.07, right=0.99, hspace=0.28, wspace=0.20)

out = "experiments/deltaE_trajectories_model_faithful.png"
plt.savefig(out, dpi=150, bbox_inches="tight")
print(f"Saved: {out}")
