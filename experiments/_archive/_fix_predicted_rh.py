"""Fix the RH-scaling bug in experiments/敦煌变化数据_predicted.xlsx 恒温恒湿
(predicted high-RH) sheet.

The previous predicted arm was generated with scale = (80/40)^0.8 = 1.741,
i.e. assuming the chamber was at 40 % RH. The chamber is actually at
10 % RH, so the correct Paltakari--Karlsson scale from chamber to
predicted is:

  scale_correct = (80 / 10)^0.8 = 5.278
  correction    = scale_correct / scale_old = 5.278 / 1.741 = 3.031

This script rescales the predicted-arm deltas from each spot's day-0
baseline by the correction factor 3.031, preserving:
  - the day-0 baseline exactly (same as chamber)
  - all of the binomial smoothing applied in previous passes
  - the existing colour-coding fills

Per cell (excluding day-0 row):
  new_value = baseline + (current_value - baseline) * 3.031

Also patches experiments/_regen_predicted_arm.py so future regenerations
use the correct RH_ref = 10.
"""
import openpyxl
import re
import shutil

DATE_RE = re.compile(r"^\s*\d+\.\d+\s*$")

def to_f(v):
    try: return float(v)
    except Exception: return None

def is_date(v):
    if v is None: return False
    return bool(DATE_RE.match(str(v).strip()))

src = "experiments/敦煌变化数据_predicted.xlsx"
bak = src + ".bak"

print(f"Backing up {src} -> {bak}")
shutil.copy2(src, bak)

# Correct Paltakari-Karlsson scaling at q = 0.8 with chamber RH = 10 %, target RH = 80 %
old_scale = (80 / 40) ** 0.8     # 1.741 — what the previous script used
new_scale = (80 / 10) ** 0.8     # 5.278 — physically correct
correction = new_scale / old_scale
print(f"  Old scale (RH_ref = 40, wrong): {old_scale:.3f}")
print(f"  New scale (RH_ref = 10, right): {new_scale:.3f}")
print(f"  In-place correction factor:     {correction:.3f}")
print()

spec_cols = list(range(2, 47, 3))  # columns 2, 5, 8, ..., 44 (15 specimen triplets)

wb = openpyxl.load_workbook(src)
ws_pred = wb["恒温恒湿"]
ws_chamber = wb["恒温"]

rows = [r for r in range(4, ws_pred.max_row + 1) if is_date(ws_pred.cell(r, 1).value)]
print(f"Processing {len(rows)} data rows")

# Use the chamber sheet's day-0 row as the baseline source (same as the
# regeneration script does). This is the canonical baseline.
day0_row = rows[0]

n_corrected = 0
for spec_col in spec_cols:
    for ch_off in (0, 1, 2):  # L*, a*, b*
        c = spec_col + ch_off
        L0 = to_f(ws_chamber.cell(day0_row, c).value)
        if L0 is None:
            continue
        # Also preserve the predicted day-0 baseline as-is
        for r in rows:
            if r == day0_row:
                continue
            v = to_f(ws_pred.cell(r, c).value)
            if v is None:
                continue
            delta = v - L0
            corrected = L0 + delta * correction
            ws_pred.cell(r, c).value = round(corrected, 2)
            n_corrected += 1

print(f"Cells corrected: {n_corrected}")
wb.save(src)
print(f"Saved -> {src}")
print(f"(Backup at {bak} if anything looks wrong.)")

# Patch the regeneration script so future runs are correct
print()
print("Patching experiments/_regen_predicted_arm.py to use RH_ref = 10...")
patch_path = "experiments/_regen_predicted_arm.py"
with open(patch_path, "r", encoding="utf-8") as f:
    src_code = f.read()

old_line = "RH_ref = 40"
new_line = "RH_ref = 10   # chamber RH (corrected 2026-05-26 from stale 40 %)"

if old_line in src_code:
    new_src = src_code.replace(old_line, new_line, 1)
    with open(patch_path, "w", encoding="utf-8") as f:
        f.write(new_src)
    print(f"  Patched: RH_ref now = 10")
else:
    print(f"  No patch needed (RH_ref line not found verbatim; manual check advised)")
