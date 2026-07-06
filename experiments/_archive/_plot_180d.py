"""Plot the 180-day extended trajectories."""
import openpyxl, math, re
import numpy as np
import matplotlib.pyplot as plt

DATE_RE = re.compile(r"^\s*\d+\.\d+\s*$")
def to_f(v):
    try: return float(v)
    except Exception: return float("nan")
def is_date(v):
    if v is None: return False
    return bool(DATE_RE.match(str(v).strip()))
def date_to_day(d):
    s = str(d).strip(); m, day = s.split(".")
    dpm = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    return (sum(dpm[: int(m) - 1]) + int(day)) - (sum(dpm[:3]) + 17)

spec_cols = {
    "R":  [2, 5, 8],   "G":  [11, 14, 17], "B":  [20, 23, 26],
    "TR": [29, 32],    "TG": [35, 38],     "TB": [41, 44],
}
PIG = {"R": "Vermilion", "G": "Malachite", "B": "Azurite",
       "TR": "Vermilion", "TG": "Malachite", "TB": "Azurite"}
N_MEASURED = 9  # rows 4..12 in the extended sheets remain the measured points

def load(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    arms = {}
    for sheet, label in [
        ("室温",     "23 C / 40%RH  (measured + extrapolated)"),
        ("恒温",     "40 C / 40%RH  (measured + extrapolated)"),
        ("恒温恒湿", "40 C / 80%RH  (predicted)"),
    ]:
        ws = wb[sheet]
        rows = [r for r in range(4, ws.max_row + 1) if is_date(ws.cell(r, 1).value)]
        if not rows:
            arms[label] = None; continue
        times = np.array([date_to_day(ws.cell(r, 1).value) for r in rows])
        per = {}
        for grp, cols in spec_cols.items():
            baselines = [
                (to_f(ws.cell(rows[0], c).value),
                 to_f(ws.cell(rows[0], c + 1).value),
                 to_f(ws.cell(rows[0], c + 2).value))
                for c in cols
            ]
            spot_series = [[] for _ in cols]
            for r in rows:
                for i, c in enumerate(cols):
                    L = to_f(ws.cell(r, c).value)
                    a = to_f(ws.cell(r, c + 1).value)
                    b = to_f(ws.cell(r, c + 2).value)
                    L0, a0, b0 = baselines[i]
                    if any(np.isnan([L, a, b, L0, a0, b0])):
                        spot_series[i].append(np.nan)
                    else:
                        spot_series[i].append(
                            math.sqrt((L - L0) ** 2 + (a - a0) ** 2 + (b - b0) ** 2)
                        )
            per[grp] = {"t": times, "spots": spot_series}
        arms[label] = per
    return arms

arms = load("experiments/敦煌变化数据_180d.xlsx")

styles = {
    "23 C / 40%RH  (measured + extrapolated)":  {"color": "#2E86AB", "ls": "-",  "marker": "o"},
    "40 C / 40%RH  (measured + extrapolated)":  {"color": "#E63946", "ls": "-",  "marker": "s"},
    "40 C / 80%RH  (predicted)":                {"color": "#6A4C93", "ls": "--", "marker": "^"},
}

fig, axes = plt.subplots(2, 3, figsize=(14, 8), sharex=True)
geom_groups = [("block (10x10x4 cm)", ["TR", "TG", "TB"]),
               ("pedestal (1:5 lotus base)", ["R", "G", "B"])]

# Determine per-row ymax for shared scale within row
for row, (geom, groups) in enumerate(geom_groups):
    row_max = 0.0
    for grp in groups:
        for label, arm in arms.items():
            if arm is None: continue
            spots = np.array(arm[grp]["spots"], dtype=float)
            if spots.size:
                m = np.nanmax(spots)
                if not np.isnan(m): row_max = max(row_max, m)

    for col, grp in enumerate(groups):
        ax = axes[row, col]
        # vertical separator at day 34 — measurement / extrapolation boundary
        ax.axvline(34, color="gray", lw=0.8, ls=":", alpha=0.7)
        ax.text(34, ax.get_ylim()[1] if False else row_max * 1.05,
                "measured | extrapolated", rotation=90,
                fontsize=7, color="gray", va="top", ha="right", alpha=0.7) if (row == 0 and col == 0) else None

        for label, arm in arms.items():
            if arm is None: continue
            t = np.array(arm[grp]["t"], dtype=float)
            spots = np.array(arm[grp]["spots"], dtype=float)
            mean = np.nanmean(spots, axis=0)
            mn = np.nanmin(spots, axis=0); mx = np.nanmax(spots, axis=0)
            st = styles[label]
            ax.fill_between(t, mn, mx, alpha=0.12, color=st["color"])
            # individual spot traces (very faint)
            for s in spots:
                ax.plot(t, s, color=st["color"], linestyle=st["ls"],
                        alpha=0.25, linewidth=0.8)
            # mean trace, prominent
            ax.plot(t, mean, color=st["color"], linestyle=st["ls"],
                    marker=st["marker"], markersize=3.5, linewidth=1.8,
                    label=label if (row == 0 and col == 0) else None)
        ax.set_title(f"{PIG[grp]} — {geom}", fontsize=10)
        ax.grid(True, alpha=0.3)
        ax.set_xlim(-3, 185)
        ax.set_ylim(-0.2, row_max * 1.15 + 1)
        if col == 0: ax.set_ylabel(r"$\Delta E^*_{ab}$")
        if row == 1: ax.set_xlabel("Days since 17 Apr 2026")
        n_spots = len(spec_cols[grp])
        ax.text(0.97, 0.04, f"n={n_spots} spots",
                transform=ax.transAxes, ha="right", va="bottom",
                fontsize=7, alpha=0.6, style="italic")

handles, labels = axes[0, 0].get_legend_handles_labels()
fig.legend(handles, labels, loc="upper center", ncol=3,
           frameon=False, fontsize=9.5, bbox_to_anchor=(0.5, 0.965))

footer1 = ("Extrapolation: per-spot linear fit on days 0-34 (outliers masked); extended to day 180 "
           "at 14-day intervals; Gaussian noise added with sigma matched to the spot's residuals.")
footer2 = ("Predicted arm: per-spot RH-scaling of the linear trend, factor (80/40)^0.8 = 1.74; "
           "fresh noise drawn with the 40 C/40%RH sigma. Dotted vertical = measurement/extrapolation boundary.")
fig.text(0.5, 0.030, footer1, ha="center", va="bottom", fontsize=8, style="italic", color="#555")
fig.text(0.5, 0.005, footer2, ha="center", va="bottom", fontsize=8, style="italic", color="#555")

fig.suptitle("Per-pigment dE* trajectories — 34-day measured + extrapolated to 180 d (matched noise)",
             y=0.985, fontsize=12, fontweight="bold")
plt.subplots_adjust(top=0.88, bottom=0.13, left=0.07, right=0.99, hspace=0.32, wspace=0.22)

out = "experiments/deltaE_trajectories_180d.png"
plt.savefig(out, dpi=150, bbox_inches="tight")
print(f"Saved: {out}")
