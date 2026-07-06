"""Internal time-window cross-validation of the per-spot linear rate law,
and bootstrap confidence intervals for the inferred Arrhenius E_a.

For each spot:
  TRAIN on April measurements only (days 0, 3, 6, 9, 12)
  PREDICT May measurements (days 25, 27, 31, 34) via linear extrapolation
  Compute spot-level metrics (RMSE, MAE) on the predicted L*a*b* deltas

Aggregate per pigment x condition into:
  - prediction R² between observed and predicted dE*_{ab}(t) on May data
  - per-pigment 95% bootstrap CI on inferred dark-pathway E_a
"""
import openpyxl
import math
import re
import numpy as np
from scipy import stats

np.random.seed(42)

DATE_RE = re.compile(r"^\s*\d+\.\d+\s*$")

def to_f(v):
    try:
        return float(v)
    except Exception:
        return float("nan")

def is_date(v):
    if v is None:
        return False
    return bool(DATE_RE.match(str(v).strip()))

def date_to_day(d):
    s = str(d).strip()
    m, day = s.split(".")
    dpm = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    return (sum(dpm[: int(m) - 1]) + int(day)) - (sum(dpm[:3]) + 17)

spec_cols = {
    "R":  [2, 5, 8],
    "G":  [11, 14, 17],
    "B":  [20, 23, 26],
    "TR": [29, 32],
    "TG": [35, 38],
    "TB": [41, 44],
}

# Outliers identified in earlier analysis (same as _extend_180d.py)
SRC_OUT = {
    ("室温", "R",  0, "5.12"),
    ("恒温", "TR", 0, "5.14"),
}

xlsx = "experiments/敦煌变化数据(1).xlsx"
wb = openpyxl.load_workbook(xlsx, data_only=True)

# Load each sheet into per-spot (t, L, a, b) arrays
data = {}  # data[sheet][group, spot] = (t, L, a, b)
for sheet in ("室温", "恒温"):
    ws = wb[sheet]
    rows = [r for r in range(4, ws.max_row + 1) if is_date(ws.cell(r, 1).value)]
    times = np.array([date_to_day(ws.cell(r, 1).value) for r in rows])
    date_strs = [str(ws.cell(r, 1).value).strip() for r in rows]
    data[sheet] = {}
    for grp, cols in spec_cols.items():
        for sidx, c in enumerate(cols):
            L, a, b = [], [], []
            for i, r in enumerate(rows):
                if (sheet, grp, sidx, date_strs[i]) in SRC_OUT:
                    L.append(np.nan); a.append(np.nan); b.append(np.nan)
                else:
                    L.append(to_f(ws.cell(r, c).value))
                    a.append(to_f(ws.cell(r, c + 1).value))
                    b.append(to_f(ws.cell(r, c + 2).value))
            data[sheet][(grp, sidx)] = (times, np.array(L), np.array(a), np.array(b))

# Time split
APRIL_MAX = 12  # days 0-12 are training (April 17 - April 29)
TRAIN_MASK = lambda t: t <= APRIL_MAX
TEST_MASK = lambda t: t > APRIL_MAX

def fit_linear_lab(t, L, a, b):
    """Fit linear L(t), a(t), b(t) on a subset of points (NaN-tolerant)."""
    mask = ~np.isnan(L) & ~np.isnan(a) & ~np.isnan(b)
    if mask.sum() < 2:
        return None
    tf = t[mask]; Lf = L[mask]; af = a[mask]; bf = b[mask]
    sL, iL = np.polyfit(tf, Lf, 1)
    sa, ia = np.polyfit(tf, af, 1)
    sb, ib = np.polyfit(tf, bf, 1)
    return (sL, iL, sa, ia, sb, ib)

def predict_dE(t_target, L0, a0, b0, fit):
    """Predict dE*_{ab}(t_target) vs (L0, a0, b0) baseline."""
    sL, iL, sa, ia, sb, ib = fit
    Lp = sL * t_target + iL
    ap = sa * t_target + ia
    bp = sb * t_target + ib
    return math.sqrt((Lp - L0) ** 2 + (ap - a0) ** 2 + (bp - b0) ** 2)

def observed_dE(L, a, b, L0, a0, b0):
    return math.sqrt((L - L0) ** 2 + (a - a0) ** 2 + (b - b0) ** 2)

# -----------------------------------------------------------------------------
# Part 1: Time-window cross-validation
# -----------------------------------------------------------------------------

print("=" * 72)
print("PART 1: Time-window cross-validation (train April, predict May)")
print("=" * 72)
print()
print(f"{'Sheet':6s}  {'Group':4s}  {'Spot':4s}  {'n_train':>7s}  {'n_test':>6s}  "
      f"{'RMSE_dE':>8s}  {'MAE_dE':>7s}")

all_obs = []
all_pred = []
per_group = {}

for sheet in ("室温", "恒温"):
    for (grp, sidx), (t, L, a, b) in data[sheet].items():
        mask_train = TRAIN_MASK(t)
        mask_test  = TEST_MASK(t)
        # baseline from first valid TRAIN point
        Lt_train = L[mask_train]; at_train = a[mask_train]; bt_train = b[mask_train]
        t_train  = t[mask_train]
        valid_train = ~np.isnan(Lt_train) & ~np.isnan(at_train) & ~np.isnan(bt_train)
        if valid_train.sum() < 2:
            continue
        L0 = Lt_train[valid_train][0]; a0 = at_train[valid_train][0]; b0 = bt_train[valid_train][0]
        fit = fit_linear_lab(t_train, Lt_train, at_train, bt_train)
        if fit is None:
            continue
        Lt_test = L[mask_test]; at_test = a[mask_test]; bt_test = b[mask_test]
        t_test  = t[mask_test]
        valid_test = ~np.isnan(Lt_test) & ~np.isnan(at_test) & ~np.isnan(bt_test)
        if valid_test.sum() < 1:
            continue
        # Predicted vs observed dE* at each test time
        obs = []; pred = []
        for i in range(len(t_test)):
            if not valid_test[i]:
                continue
            o = observed_dE(Lt_test[i], at_test[i], bt_test[i], L0, a0, b0)
            p = predict_dE(t_test[i], L0, a0, b0, fit)
            obs.append(o); pred.append(p)
        obs = np.array(obs); pred = np.array(pred)
        rmse = float(np.sqrt(np.mean((pred - obs) ** 2)))
        mae  = float(np.mean(np.abs(pred - obs)))
        print(f"{sheet:6s}  {grp:4s}  {sidx:>4d}  "
              f"{int(valid_train.sum()):>7d}  {int(valid_test.sum()):>6d}  "
              f"{rmse:>8.2f}  {mae:>7.2f}")
        all_obs.extend(obs); all_pred.extend(pred)
        per_group.setdefault((sheet, grp), {"obs": [], "pred": []})
        per_group[(sheet, grp)]["obs"].extend(obs)
        per_group[(sheet, grp)]["pred"].extend(pred)

all_obs = np.array(all_obs); all_pred = np.array(all_pred)
overall_rmse = float(np.sqrt(np.mean((all_pred - all_obs) ** 2)))
overall_mae  = float(np.mean(np.abs(all_pred - all_obs)))
overall_r2   = 1 - np.sum((all_obs - all_pred) ** 2) / np.sum((all_obs - np.mean(all_obs)) ** 2)
print()
print(f"OVERALL across all spots: RMSE = {overall_rmse:.2f}, MAE = {overall_mae:.2f}, "
      f"R^2 = {overall_r2:.3f}, n = {len(all_obs)} (spot, test-time) pairs")
print()

print("Per (sheet, pigment) R^2 of predicted dE* on May test points:")
for (sheet, grp), v in per_group.items():
    o = np.array(v["obs"]); p = np.array(v["pred"])
    if len(o) < 2:
        continue
    if np.var(o) == 0:
        r2 = float("nan")
    else:
        r2 = 1 - np.sum((o - p) ** 2) / np.sum((o - np.mean(o)) ** 2)
    rmse = float(np.sqrt(np.mean((p - o) ** 2)))
    print(f"  {sheet:6s}  {grp:4s}  RMSE={rmse:5.2f}  R^2={r2:+.2f}  n={len(o)}")
print()

# -----------------------------------------------------------------------------
# Part 2: Bootstrap CI for inferred dark-pathway E_a per pigment (blocks only)
# -----------------------------------------------------------------------------

print("=" * 72)
print("PART 2: Bootstrap 95% CI for E_a per pigment (q = 0.8 fixed)")
print("=" * 72)
print()
print("Resampling 1000 times with replacement over the spot index per pigment,")
print("refitting the linear rate, taking the per-arm mean slope, computing")
print("apparent E_a via Arrhenius+Paltakari-Karlsson with q = 0.8 and the")
print("chamber RH = 10%, room-T RH = 40%.")
print()

R_gas = 8.314
T_room = 296.15
T_chamber = 313.15
RH_room = 40
RH_chamber = 10
q = 0.8
arrhenius_factor = 1.0 / T_room - 1.0 / T_chamber  # positive (1/296 - 1/313 > 0)

def per_spot_slope_mag(t, L, a, b):
    mask = ~np.isnan(L) & ~np.isnan(a) & ~np.isnan(b)
    if mask.sum() < 2: return None
    sL, _ = np.polyfit(t[mask], L[mask], 1)
    sa, _ = np.polyfit(t[mask], a[mask], 1)
    sb, _ = np.polyfit(t[mask], b[mask], 1)
    return math.sqrt(sL ** 2 + sa ** 2 + sb ** 2)

# For each pigment we bootstrap the headline geometry as cited in the paper:
#   Vermilion: PEDESTAL (TR), n=2 spots per arm
#   Azurite:   BLOCK    (B),  n=3 spots per arm
#   Malachite: BLOCK    (G),  n=3 spots per arm
# (Vermilion block returns the ~18 kJ/mol photolytic residual; we report
#  that separately below as a sanity check on the two-pathway interpretation.)
n_boot = 1000
headline_pairing = [
    ("TR", "Vermilion (pedestal, headline)"),
    ("B",  "Azurite (block, headline)"),
    ("G",  "Malachite (block, headline)"),
    ("R",  "Vermilion (block, photolytic-pathway residual)"),
]

results = {}
for grp, label in headline_pairing:
    slopes_room = []; slopes_chamber = []
    for sidx in range(len(spec_cols[grp])):
        t_r, L_r, a_r, b_r = data["室温"][(grp, sidx)]
        t_c, L_c, a_c, b_c = data["恒温"][(grp, sidx)]
        sr = per_spot_slope_mag(t_r, L_r, a_r, b_r)
        sc = per_spot_slope_mag(t_c, L_c, a_c, b_c)
        if sr is not None: slopes_room.append(sr)
        if sc is not None: slopes_chamber.append(sc)
    slopes_room = np.array(slopes_room); slopes_chamber = np.array(slopes_chamber)
    n_r = len(slopes_room); n_c = len(slopes_chamber)
    if n_r < 1 or n_c < 1:
        continue
    boot_Ea = []
    for _ in range(n_boot):
        sr_b = slopes_room[np.random.randint(0, n_r, n_r)]
        sc_b = slopes_chamber[np.random.randint(0, n_c, n_c)]
        k_room_mean = np.mean(sr_b)
        k_chamber_mean = np.mean(sc_b)
        if k_room_mean <= 0 or k_chamber_mean <= 0:
            continue
        ratio = k_room_mean / k_chamber_mean  # k(room)/k(chamber)
        lnR = math.log(ratio)
        rh_term = q * math.log(RH_room / RH_chamber)
        # ln(ratio) = -Ea/R * (1/T_room - 1/T_chamber) + q * ln(RH_room/RH_chamber)
        # arrhenius_factor = 1/T_room - 1/T_chamber > 0
        # so Ea/R = (rh_term - lnR) / arrhenius_factor
        Ea = (rh_term - lnR) / arrhenius_factor * R_gas  # J/mol
        boot_Ea.append(Ea / 1000.0)  # kJ/mol
    if not boot_Ea:
        continue
    boot_Ea = np.array(boot_Ea)
    median = float(np.median(boot_Ea))
    p2_5 = float(np.percentile(boot_Ea, 2.5))
    p97_5 = float(np.percentile(boot_Ea, 97.5))
    print(f"  {label:50s}: median = {median:5.1f} kJ/mol, "
          f"95% CI = [{p2_5:5.1f}, {p97_5:5.1f}], n_spots_per_arm = {n_r}, n_boot = {len(boot_Ea)}")
    results[grp] = {"label": label, "median": median, "p2_5": p2_5, "p97_5": p97_5, "n_boot": len(boot_Ea), "n_spots": n_r}

print()
print("=" * 72)
print("Saving results to experiments/_validation.json")
print("=" * 72)

import json
key_map = {"TR": "vermilion_pedestal", "B": "azurite_block", "G": "malachite_block", "R": "vermilion_block_photolytic"}
out = {
    "time_window_cv": {
        "overall_rmse": overall_rmse,
        "overall_mae": overall_mae,
        "overall_r2": overall_r2,
        "n_pairs": int(len(all_obs)),
    },
    "bootstrap_Ea_95ci_kJmol": {
        key_map[grp]: results[grp] for grp in results if grp in key_map
    },
}
with open("experiments/_validation.json", "w") as f:
    json.dump(out, f, indent=2)
print()
print("Done.")
