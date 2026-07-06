"""Plot the 60-day MEASURED ΔE trajectories from experiments/mogao_data.xlsx.

Same 2x3 geometry as _plot_90d.py (block over pedestal; vermilion / malachite /
azurite columns) but this workbook holds raw pilot measurements only:
  * 室温    -> 23 C / 40%RH (room, lit)
  * 恒温    -> 40 C / 10%RH (chamber, dark)
  * 恒温恒湿 -> 40 C / 80%RH (predicted) -- EMPTY here, skipped.
No model extrapolation, so no day-34 guide line.
"""
import openpyxl, math, re
import numpy as np
import matplotlib.pyplot as plt

DATE_RE = re.compile(r"^\s*\d+\.\d+\s*$")

def to_f(v):
    try: return float(v)
    except Exception: return float("nan")

def is_date(v):
    if v is None: return False
    return bool(DATE_RE.match(str(v).strip()))

def date_to_day(d):
    s = str(d).strip(); m, day = s.split(".")
    dpm = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    return (sum(dpm[: int(m) - 1]) + int(day)) - (sum(dpm[:3]) + 17)

spec_cols = {
    "R":  [2, 5, 8],   "G":  [11, 14, 17], "B":  [20, 23, 26],
    "TR": [29, 32],    "TG": [35, 38],     "TB": [41, 44],
}
PIG = {"R": "Vermilion", "G": "Malachite", "B": "Azurite",
       "TR": "Vermilion", "TG": "Malachite", "TB": "Azurite"}

src = "experiments/mogao_data.xlsx"
wb = openpyxl.load_workbook(src, data_only=True)

# --- Hampel-style per-channel outlier filter (matches _denoise_spots.py) ---
FILTER = True
WINDOW = 3       # centered 3-point window
K_MAD = 3.0      # MAD multiplier threshold
FLOOR = 3.0      # absolute floor on deviation (L*a*b* units) -- skip nudges < 3

def hampel_replace(series):
    """Replace interior outliers (|x-median| > FLOOR and > K_MAD*MAD over a
    3-pt centered window) with the window median. Endpoints preserved."""
    n = len(series)
    out = list(series)
    n_fixed = 0
    for i in range(1, n - 1):
        win = [series[i - 1], series[i], series[i + 1]]
        valid = [v for v in win if v is not None and not np.isnan(v)]
        x = series[i]
        if x is None or np.isnan(x) or len(valid) < 2:
            continue
        m = sorted(valid)[len(valid) // 2]
        absdev = sorted(abs(v - m) for v in valid)
        mad = absdev[len(absdev) // 2]
        dev = abs(x - m)
        if dev > FLOOR and dev > K_MAD * (mad + 1e-6):
            out[i] = m
            n_fixed += 1
    return out, n_fixed

def load_arm(sheet_name):
    ws = wb[sheet_name]
    rows = [r for r in range(1, ws.max_row + 1) if is_date(ws.cell(r, 1).value)]
    times = np.array([date_to_day(ws.cell(r, 1).value) for r in rows])
    per_spot = {}
    fixed_total = 0
    for grp, cols in spec_cols.items():
        for sidx, c in enumerate(cols):
            # gather each channel as a series over time, then filter per channel
            chans = []
            for off in (0, 1, 2):
                s = [to_f(ws.cell(r, c + off).value) for r in rows]
                if FILTER:
                    s, nf = hampel_replace(s)
                    fixed_total += nf
                chans.append(np.array(s, dtype=float))
            L, a, b = chans
            L0, a0, b0 = L[0], a[0], b[0]
            dE = np.sqrt((L - L0) ** 2 + (a - a0) ** 2 + (b - b0) ** 2)
            per_spot[(grp, sidx)] = (times, dE)
    if FILTER:
        print(f"  [{sheet_name}] Hampel-replaced {fixed_total} channel points")
    return per_spot

arms = {
    "23 C / 40%RH (room, lit)":      load_arm("室温"),
    "40 C / 10%RH (chamber, dark)":  load_arm("恒温"),
}
styles = {
    "23 C / 40%RH (room, lit)":      {"color": "#2E86AB", "marker": "o", "ls": "-"},
    "40 C / 10%RH (chamber, dark)":  {"color": "#E63946", "marker": "s", "ls": "-"},
}

fig, axes = plt.subplots(2, 3, figsize=(14, 7.5), sharex=True)
geom_groups = [
    ("block (10x10x4 cm tile)",   ["TR", "TG", "TB"]),
    ("pedestal (1:5 lotus base)", ["R",  "G",  "B"]),
]

for row, (geom, grps) in enumerate(geom_groups):
    row_max = 0.0
    for grp in grps:
        cols = spec_cols[grp]
        for per_spot in arms.values():
            for sidx in range(len(cols)):
                m = np.nanmax(per_spot[(grp, sidx)][1])
                if not np.isnan(m): row_max = max(row_max, m)
    for col, grp in enumerate(grps):
        ax = axes[row, col]
        cols = spec_cols[grp]
        for arm_label, per_spot in arms.items():
            st = styles[arm_label]
            spot_curves = []
            for sidx in range(len(cols)):
                t, dE = per_spot[(grp, sidx)]
                spot_curves.append(dE)
                ax.plot(t, dE, color=st["color"], linestyle=st["ls"],
                        alpha=0.25, linewidth=0.7)
            stacked = np.array(spot_curves, dtype=float)
            mean = np.nanmean(stacked, axis=0)
            mn = np.nanmin(stacked, axis=0); mx = np.nanmax(stacked, axis=0)
            t = per_spot[(grp, 0)][0]
            ax.fill_between(t, mn, mx, color=st["color"], alpha=0.12)
            ax.plot(t, mean, color=st["color"], linestyle=st["ls"],
                    marker=st["marker"], markersize=4, linewidth=1.6,
                    label=arm_label if (row == 0 and col == 0) else None)
        ax.set_title(f"{PIG[grp]} -- {geom}", fontsize=10)
        ax.grid(True, alpha=0.3)
        ax.set_xlim(-2, 64)
        ax.set_ylim(-0.2, row_max * 1.10 + 0.5)
        if col == 0: ax.set_ylabel(r"$\Delta E^*_{ab}$")
        if row == 1: ax.set_xlabel("Days since 17 Apr 2026")
        ax.text(0.97, 0.04, f"n = {len(cols)} spots",
                transform=ax.transAxes, ha="right", va="bottom",
                fontsize=7, alpha=0.6, style="italic")

handles, labels = axes[0, 0].get_legend_handles_labels()
fig.legend(handles, labels, loc="upper center", ncol=2,
           frameon=False, fontsize=8.5, bbox_to_anchor=(0.5, 0.94))

fig.suptitle("Per-pigment $\\Delta E^*_{ab}$ trajectories: 60-day colorimetry pilot, Hampel-filtered (mogao_data.xlsx)",
             y=0.99, fontsize=11.5, fontweight="bold")
fig.text(0.5, 0.02,
         "14 timepoints (17 Apr - 17 Jun 2026), ~4-5 day cadence.  Per-channel Hampel outlier filter "
         "(3-pt window, k=3.0 MAD, floor 3.0, endpoints kept).  40 C/80%RH predicted arm absent from this workbook.",
         ha="center", va="bottom", fontsize=7.5, style="italic", color="#555555", wrap=True)
plt.subplots_adjust(top=0.86, bottom=0.11, left=0.07, right=0.99, hspace=0.28, wspace=0.20)

out = "experiments/deltaE_trajectories_mogao_data_filtered.png"
plt.savefig(out, dpi=150, bbox_inches="tight")
print(f"Saved: {out}")
