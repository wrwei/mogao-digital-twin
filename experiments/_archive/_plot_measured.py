"""Plot the RAW MEASURED data from experiments/mogao_data.xlsx.

No model generation: ΔE is computed directly from the measured L*a*b* (per spot,
vs each spot's own day-0 baseline). Gross instrument outliers are removed with a
disclosed Hampel filter (3-pt window, 3xMAD, floor 3.0). Error bars are the real
between-spot standard error (std/sqrt(n)) -- i.e. actual measured variability, not
an assumed sigma. Real measurement cadence is preserved (holiday gaps left as gaps).
"""
import openpyxl, math, re
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D

NOISE_FLOOR = 0.7 * 1.5958   # ΔE floor implied by the data's ~0.7/channel scatter

DATE_RE = re.compile(r"^\s*\d+\.\d+\s*$")
def to_f(v):
    try: return float(v)
    except Exception: return None
def is_date(v): return v is not None and bool(DATE_RE.match(str(v).strip()))
def date_to_day(d):
    m, day = str(d).strip().split("."); dpm=[31,28,31,30,31,30,31,31,30,31,30,31]
    return (sum(dpm[:int(m)-1]) + int(day)) - (sum(dpm[:3]) + 17)

def hampel(s):
    n=len(s); o=list(s)
    for i in range(1, n-1):
        w=[s[i-1],s[i],s[i+1]]; v=[x for x in w if x is not None]
        if s[i] is None or len(v)<2: continue
        m=sorted(v)[len(v)//2]; mad=sorted(abs(x-m) for x in v)[len(v)//2]
        if abs(s[i]-m)>3 and abs(s[i]-m)>3*(mad+1e-6): o[i]=m
    return o

spec_cols = {"R":[2,5,8],"G":[11,14,17],"B":[20,23,26],"TR":[29,32],"TG":[35,38],"TB":[41,44]}
PIG = {"R":"Vermilion","G":"Malachite","B":"Azurite","TR":"Vermilion","TG":"Malachite","TB":"Azurite"}

wb = openpyxl.load_workbook("experiments/mogao_data.xlsx", data_only=True)

def arm_series(sheet, grp):
    ws = wb[sheet]
    rows = [r for r in range(1, ws.max_row+1) if is_date(ws.cell(r,1).value)]
    t = np.array([date_to_day(ws.cell(r,1).value) for r in rows], float)
    curves = []
    for c in spec_cols[grp]:
        ch = {off: hampel([to_f(ws.cell(r,c+off).value) for r in rows]) for off in (0,1,2)}
        L,a,b = (np.array(ch[0],float), np.array(ch[1],float), np.array(ch[2],float))
        curves.append(np.sqrt((L-L[0])**2 + (a-a[0])**2 + (b-b[0])**2))
    stk = np.array(curves)                       # n_spots x n_time
    mean = np.nanmean(stk, 0)
    n = stk.shape[0]
    sem = np.nanstd(stk, 0, ddof=1)/math.sqrt(n) if n > 1 else np.full_like(mean, np.nan)
    return t, mean, sem, n

ARMS = {"室温":("23 C / 40%RH (room, lit)", "#2E86AB", "o"),
        "恒温":("40 C / 10%RH (chamber, dark)", "#E63946", "s")}

fig, axes = plt.subplots(2, 3, figsize=(14, 7.5), sharex=True)
geom = [("block (10x10x4 cm tile)", ["TR","TG","TB"]),
        ("pedestal (1:5 lotus base)", ["R","G","B"])]

row_max_all = 0.0
cache = {}
for row,(gname,grps) in enumerate(geom):
    for grp in grps:
        for sh in ARMS:
            cache[(sh,grp)] = arm_series(sh, grp)
            row_max_all = max(row_max_all, np.nanmax(cache[(sh,grp)][1]))

for row,(gname,grps) in enumerate(geom):
    for col,grp in enumerate(grps):
        ax = axes[row,col]
        ax.axhspan(0, NOISE_FLOOR, color="#9AA0A6", alpha=0.13, lw=0, zorder=0)
        ax.axhline(NOISE_FLOOR, color="#9AA0A6", ls=":", lw=1.1, alpha=0.9, zorder=0)
        for sh,(label,color,mk) in ARMS.items():
            t,mean,sem,n = cache[(sh,grp)]
            ax.errorbar(t, mean, yerr=sem, color=color, marker=mk, ms=4, lw=1.6,
                        elinewidth=0.9, capsize=2.2, capthick=0.9,
                        label=(label if (row==0 and col==0) else None))
        ax.set_title(f"{PIG[grp]} -- {gname}", fontsize=10)
        ax.grid(True, alpha=0.3); ax.set_xlim(-2, 64)
        ax.set_ylim(-0.3, row_max_all*1.12 + 0.4)
        if col==0: ax.set_ylabel(r"$\Delta E^*_{ab}$")
        if row==1: ax.set_xlabel("Days since 17 Apr 2026")
        ax.text(0.97, 0.04, f"n = {cache[(list(ARMS)[0],grp)][3]} spots",
                transform=ax.transAxes, ha="right", va="bottom", fontsize=7, alpha=0.6, style="italic")

handles, labels = axes[0,0].get_legend_handles_labels()
handles += [Line2D([0],[0], color="#9AA0A6", ls=":", lw=1.1)]
labels  += [f"noise floor (ΔE≈{NOISE_FLOOR:.1f}, below = unresolvable)"]
fig.legend(handles, labels, loc="upper center", ncol=3, frameon=False,
           fontsize=8.5, bbox_to_anchor=(0.5, 0.945))
fig.suptitle("mogao_data.xlsx: RAW MEASURED 60-day colour change (2 arms, Hampel outlier-removed)",
             y=0.99, fontsize=11.5, fontweight="bold")
fig.text(0.5, 0.02,
         "Directly measured ΔE (no model): per spot vs its own day-0 baseline, mean over spots.  Gross outliers removed by Hampel "
         "filter (3-pt, 3×MAD, floor 3.0).  Error bars = between-spot standard error (std/√n).  Real cadence (holiday gaps left as gaps).",
         ha="center", va="bottom", fontsize=7.5, style="italic", color="#555555", wrap=True)
plt.subplots_adjust(top=0.85, bottom=0.11, left=0.07, right=0.99, hspace=0.28, wspace=0.20)

out = "experiments/deltaE_trajectories_measured_raw.png"
plt.savefig(out, dpi=150, bbox_inches="tight")
print(f"Saved: {out}")
