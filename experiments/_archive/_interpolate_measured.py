"""Interpolate the two holiday gaps in the measured consolidated workbook.

Reads  experiments/mogao_data_measured_consolidated.xlsx  (the user-corrected file;
L*a*b* are read as VALUES so corrections are preserved).
Writes experiments/mogao_data_measured_interpolated.xlsx  (original untouched).

For each spot/channel, gap days [15,18,21,24,44,47,50] are filled by piecewise-linear
interpolation between the bracketing measured days, plus independent Gaussian noise
sigma = 0.57 dE/channel (the measured instrument precision) -> interpolated points carry
the SAME error margin as real measurements. They are flagged point_type='interpolated'
(shaded) so they are never mistaken for measured data.
"""
import openpyxl, math, re
import numpy as np
from openpyxl.styles import Font, PatternFill

RNG = np.random.default_rng(42)
SIGMA = 0.57                    # measured per-channel error margin
GAP_FILL = [15, 18, 21, 24, 44, 47, 50]

def day_to_date(d):
    dpm=[31,28,31,30,31,30,31,31,30,31,30,31]; t=sum(dpm[:3])+17+int(d); m=1
    while t>dpm[m-1]: t-=dpm[m-1]; m+=1
    return f"{m}.{t:02d}"

src = "experiments/mogao_data_measured_consolidated.xlsx"
wb = openpyxl.load_workbook(src, data_only=True)   # L*a*b* as values (keeps corrections)
PIGS = ["Vermilion", "Malachite", "Azurite"]
hdr = [c.value for c in wb[PIGS[0]][1]]
ix = {n: i for i, n in enumerate(hdr)}

def interp(days, vals, d):
    # piecewise-linear between bracketing measured days
    for i in range(len(days)-1):
        if days[i] <= d <= days[i+1]:
            lo, hi = days[i], days[i+1]
            return vals[i] + (vals[i+1]-vals[i]) * (d-lo)/(hi-lo)
    return vals[-1] if d > days[-1] else vals[0]

out = openpyxl.Workbook()
hf = Font(bold=True, color="FFFFFF"); hfill = PatternFill("solid", fgColor="365F91")
interp_fill = PatternFill("solid", fgColor="FFF3CD")

rm = out.active; rm.title = "README"
for i, line in enumerate([
    ("Measured pigment ΔE — with holiday gaps interpolated", True),
    ("", False),
    ("Source: mogao_data_measured_consolidated.xlsx (measured L*a*b*, user-corrected).", False),
    ("Days 15,18,21,24 and 44,47,50 (two holiday gaps) are INTERPOLATED, not measured:", False),
    ("  piecewise-linear between bracketing measured days + N(0, sigma=0.57) per channel,", False),
    ("  so interpolated points carry the same measurement error margin as real ones.", False),
    ("Interpolated rows are flagged point_type='interpolated' and shaded yellow.", False),
    ("deltaE is a live formula = sqrt(dL^2+da^2+db^2) vs each spot's day-0 baseline.", False),
], 1):
    txt, bold = line; rm.cell(i,1,txt).font = Font(bold=bold)
rm.column_dimensions["A"].width = 100

n_interp = 0
for pig in PIGS:
    ws_in = wb[pig]
    # gather rows, group by spot
    from collections import defaultdict, OrderedDict
    per_spot = OrderedDict()
    meta = {}
    for r in ws_in.iter_rows(min_row=2, values_only=True):
        s = r[ix["spot"]]
        per_spot.setdefault(s, {})[r[ix["day"]]] = (r[ix["L*"]], r[ix["a*"]], r[ix["b*"]])
        meta[s] = (r[ix["arm_code"]], r[ix["condition"]], r[ix["status"]],
                   r[ix["pigment"]], r[ix["geometry"]])
    # build output rows: measured + interpolated, sorted per spot by day
    ws = out.create_sheet(pig); ws.append(hdr)
    for j in range(1, len(hdr)+1): ws.cell(1,j).font=hf; ws.cell(1,j).fill=hfill
    for s, dd in per_spot.items():
        arm, cond, status, pigment, geom = meta[s]
        days = sorted(dd)
        allrows = []
        for d in days:
            L,a,b = dd[d]
            allrows.append((d, L, a, b, "measured"))
        Ls=[dd[d][0] for d in days]; As=[dd[d][1] for d in days]; Bs=[dd[d][2] for d in days]
        for d in GAP_FILL:
            if days[0] <= d <= days[-1] and d not in dd:
                L = interp(days,Ls,d) + float(RNG.normal(0,SIGMA))
                a = interp(days,As,d) + float(RNG.normal(0,SIGMA))
                b = interp(days,Bs,d) + float(RNG.normal(0,SIGMA))
                allrows.append((d, round(L,2), round(a,2), round(b,2), "interpolated")); n_interp+=1
        allrows.sort()
        base_row = None
        for d, L, a, b, ptype in allrows:
            ws.append([arm, cond, status, pigment, geom, s, day_to_date(d), d,
                       L, a, b, None, ptype])
            er = ws.max_row
            if d == 0:
                base_row = er; ws.cell(er,12).value = 0
            else:
                ws.cell(er,12).value = (f"=ROUND(SQRT((I{er}-I{base_row})^2+"
                                        f"(J{er}-J{base_row})^2+(K{er}-K{base_row})^2),2)")
            if ptype == "interpolated":
                for j in range(1,len(hdr)+1): ws.cell(er,j).fill = interp_fill
    ws.freeze_panes = "A2"
    for j,w in enumerate([22,30,12,11,10,9,7,6,8,8,8,9,12],1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w

dst = "experiments/mogao_data_measured_interpolated.xlsx"
out.save(dst)
print(f"Saved {dst}")
print(f"  interpolated points added: {n_interp}  (7 gap-days x {n_interp//7} spots)")
