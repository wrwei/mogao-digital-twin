"""End-to-end pipeline: project the raw pilot data
experiments/敦煌变化数据(1).xlsx to a 3-month (90-day) trajectory and plot.

Stages:
  1. Read the raw 9-point xlsx (室温 + 恒温 measured; 恒温恒湿 empty)
  2. Hampel-filter the two measured arms for clear outliers (window 3,
     K_MAD = 3.0, FLOOR = 3.0; same parameters used on the smoothed
     workbook).
  3. Generate the 恒温恒湿 predicted arm by RH-scaling the cleaned
     chamber arm with the correct Paltakari--Karlsson coupling at
     q = 0.8 and chamber RH = 10 %, target RH = 80 % => scale = 5.278.
  4. Extend each of the three arms to 31 points at 3-day intervals
     (day 0, 3, 6, ..., 90) using a first-order saturating model per
     spot with pigment-specific dE*_max and matched-amplitude Gaussian
     noise from the per-spot residuals of the linear fit.
  5. Save the projected workbook to
     experiments/敦煌变化数据(1)_90d.xlsx
  6. Plot to experiments/deltaE_trajectories_raw_90d.png

Random seed fixed at 42 for reproducibility.
"""
import openpyxl
import math
import re
import shutil
import numpy as np
import matplotlib.pyplot as plt
from openpyxl.styles import Font, PatternFill

np.random.seed(42)

DATE_RE = re.compile(r"^\s*\d+\.\d+\s*$")

def to_f(v):
    try: return float(v)
    except Exception: return None

def is_date(v):
    if v is None: return False
    return bool(DATE_RE.match(str(v).strip()))

def date_to_day(d):
    s = str(d).strip(); m, day = s.split(".")
    dpm = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    return (sum(dpm[: int(m) - 1]) + int(day)) - (sum(dpm[:3]) + 17)

def day_to_date(d):
    dpm = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    base = sum(dpm[:3]) + 17
    target = base + int(d)
    month = 1
    while target > dpm[month - 1]:
        target -= dpm[month - 1]; month += 1
    return f"{month}.{target:02d}"

spec_cols = {
    "R":  [2, 5, 8],   "G":  [11, 14, 17], "B":  [20, 23, 26],
    "TR": [29, 32],    "TG": [35, 38],     "TB": [41, 44],
}
DELTAE_MAX = {"R": 60, "TR": 60, "G": 50, "TG": 50, "B": 40, "TB": 40}
PIG = {"R": "Vermilion", "G": "Malachite", "B": "Azurite",
       "TR": "Vermilion", "TG": "Malachite", "TB": "Azurite"}

src = "experiments/敦煌变化数据(1).xlsx"
dst = "experiments/敦煌变化数据(1)_90d.xlsx"

print(f"Copying source {src} -> {dst}")
shutil.copy2(src, dst)

# ----------------------------------------------------------------------
# Stage 2: Hampel filter the measured arms (室温, 恒温)
# ----------------------------------------------------------------------
K_MAD = 3.0
FLOOR = 3.0
print(f"\n=== Stage 2: Hampel outlier removal (K_MAD = {K_MAD}, FLOOR = {FLOOR}) ===")

wb = openpyxl.load_workbook(dst)

def hampel_replace(series):
    n = len(series)
    out = list(series)
    for i in range(1, n - 1):
        x = series[i]
        nbrs = [series[i - 1], series[i + 1]]
        valid_nbrs = [v for v in nbrs if v is not None]
        if x is None or not valid_nbrs: continue
        window_vals = [series[i - 1], series[i], series[i + 1]]
        valid_win = [v for v in window_vals if v is not None]
        if len(valid_win) < 2: continue
        m = sorted(valid_win)[len(valid_win) // 2]
        absdev = sorted(abs(v - m) for v in valid_win)
        mad = absdev[len(absdev) // 2]
        dev = abs(x - m)
        if dev > FLOOR and dev > K_MAD * (mad + 1e-6):
            out[i] = round(m, 2)
    return out

total_hampel = 0
for sheet_name in ("室温", "恒温"):
    ws = wb[sheet_name]
    rows = [r for r in range(4, ws.max_row + 1) if is_date(ws.cell(r, 1).value)]
    for grp, cols in spec_cols.items():
        for sidx, c in enumerate(cols):
            for ch_off in (0, 1, 2):
                series = [to_f(ws.cell(r, c + ch_off).value) for r in rows]
                new_series = hampel_replace(series)
                for i, (old, new) in enumerate(zip(series, new_series)):
                    if old is not None and new is not None and old != new:
                        ws.cell(rows[i], c + ch_off).value = new
                        total_hampel += 1
print(f"  Hampel replacements: {total_hampel} cells across both measured arms")

# ----------------------------------------------------------------------
# Stage 3: Generate 恒温恒湿 predicted arm from cleaned 恒温 chamber
# ----------------------------------------------------------------------
print("\n=== Stage 3: Generate predicted high-RH arm (RH = 80%) from chamber (RH = 10%) ===")
q = 0.8
RH_target = 80
RH_ref = 10  # chamber RH (corrected from earlier 40-% assumption)
rh_scale = (RH_target / RH_ref) ** q
print(f"  Paltakari-Karlsson scale = (80/10)^0.8 = {rh_scale:.3f}")

ws_chamber = wb["恒温"]
ws_pred = wb["恒温恒湿"]
rows = [r for r in range(4, ws_chamber.max_row + 1) if is_date(ws_chamber.cell(r, 1).value)]

# Clear and rebuild 恒温恒湿
for r in range(4, ws_pred.max_row + 40):
    for c in range(1, ws_pred.max_column + 2):
        ws_pred.cell(r, c).value = None
        ws_pred.cell(r, c).fill = PatternFill(fill_type=None)

# Copy date column
for r in rows:
    ws_pred.cell(r, 1).value = ws_chamber.cell(r, 1).value

predicted_fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
for grp, cols in spec_cols.items():
    for sidx, c in enumerate(cols):
        L0 = to_f(ws_chamber.cell(rows[0], c).value)
        a0 = to_f(ws_chamber.cell(rows[0], c + 1).value)
        b0 = to_f(ws_chamber.cell(rows[0], c + 2).value)
        if None in (L0, a0, b0): continue
        for r in rows:
            Lc = to_f(ws_chamber.cell(r, c).value)
            ac = to_f(ws_chamber.cell(r, c + 1).value)
            bc = to_f(ws_chamber.cell(r, c + 2).value)
            if None in (Lc, ac, bc): continue
            Lp = L0 + (Lc - L0) * rh_scale
            ap = a0 + (ac - a0) * rh_scale
            bp = b0 + (bc - b0) * rh_scale
            ws_pred.cell(r, c).value = round(Lp, 2)
            ws_pred.cell(r, c + 1).value = round(ap, 2)
            ws_pred.cell(r, c + 2).value = round(bp, 2)
            for off in (0, 1, 2):
                ws_pred.cell(r, c + off).fill = predicted_fill

# ----------------------------------------------------------------------
# Stage 4: Saturating extension to 90 days with matched noise
# ----------------------------------------------------------------------
print("\n=== Stage 4: First-order saturating extension to 90 days (3-day intervals) ===")
new_days = list(range(0, 91, 3))
print(f"  Target days: 31 points at 3-day intervals, day 0..90")

measured_fill = PatternFill(start_color="E8F4FF", end_color="E8F4FF", fill_type="solid")
extrap_fill   = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")

for sheet in ("室温", "恒温", "恒温恒湿"):
    ws = wb[sheet]
    src_rows = [r for r in range(4, ws.max_row + 1) if is_date(ws.cell(r, 1).value)]
    src_days = np.array([date_to_day(ws.cell(r, 1).value) for r in src_rows])
    n_src = len(src_rows)
    src_data = {}
    for grp, cols in spec_cols.items():
        for sidx, c in enumerate(cols):
            for ch_off, ch in [(0, "L"), (1, "a"), (2, "b")]:
                src_data[(grp, sidx, ch)] = [to_f(ws.cell(r, c + ch_off).value) for r in src_rows]
    # Clear existing data rows
    max_clear = max(src_rows[-1], 4 + len(new_days) + 5)
    for r in range(4, max_clear + 1):
        for c in range(1, ws.max_column + 2):
            ws.cell(r, c).value = None
            ws.cell(r, c).fill = PatternFill(fill_type=None)
    for i, d in enumerate(new_days):
        ws.cell(4 + i, 1).value = day_to_date(d)
    for grp, cols in spec_cols.items():
        for sidx, c in enumerate(cols):
            Ls = src_data[(grp, sidx, "L")]
            As = src_data[(grp, sidx, "a")]
            Bs = src_data[(grp, sidx, "b")]
            mask = [v is not None for v in Ls]
            if sum(mask) < 2: continue
            xs = src_days[mask]
            ys_L = np.array([Ls[i] for i in range(n_src) if mask[i]])
            ys_a = np.array([As[i] for i in range(n_src) if mask[i]])
            ys_b = np.array([Bs[i] for i in range(n_src) if mask[i]])
            if 0 in xs:
                idx0 = list(xs).index(0)
                L0, a0, b0 = ys_L[idx0], ys_a[idx0], ys_b[idx0]
            else:
                L0, a0, b0 = ys_L[0], ys_a[0], ys_b[0]
            sL, iL = np.polyfit(xs, ys_L, 1)
            sa, ia = np.polyfit(xs, ys_a, 1)
            sb, ib = np.polyfit(xs, ys_b, 1)
            slope_mag = math.sqrt(sL ** 2 + sa ** 2 + sb ** 2)
            sig_L = float(np.std(ys_L - (sL * xs + iL), ddof=1)) if len(xs) >= 2 else 0.0
            sig_a = float(np.std(ys_a - (sa * xs + ia), ddof=1)) if len(xs) >= 2 else 0.0
            sig_b = float(np.std(ys_b - (sb * xs + ib), ddof=1)) if len(xs) >= 2 else 0.0
            dEmax = DELTAE_MAX[grp]
            if slope_mag < 1e-6:
                for i_new, d in enumerate(new_days):
                    row_idx = 4 + i_new
                    if d == 0:
                        Lv, av, bv = L0, a0, b0
                    else:
                        Lv = L0 + np.random.normal(0, sig_L)
                        av = a0 + np.random.normal(0, sig_a)
                        bv = b0 + np.random.normal(0, sig_b)
                    ws.cell(row_idx, c    ).value = round(float(Lv), 2)
                    ws.cell(row_idx, c + 1).value = round(float(av), 2)
                    ws.cell(row_idx, c + 2).value = round(float(bv), 2)
                    f = predicted_fill if sheet == "恒温恒湿" else (measured_fill if d <= 34 else extrap_fill)
                    for off in (0, 1, 2): ws.cell(row_idx, c + off).fill = f
                continue
            k = slope_mag / dEmax
            uL, ua, ub = sL / slope_mag, sa / slope_mag, sb / slope_mag
            for i_new, d in enumerate(new_days):
                if d == 0:
                    Lv, av, bv = L0, a0, b0
                else:
                    f_sat = 1.0 - math.exp(-k * d)
                    dE_t = dEmax * f_sat
                    Lv = L0 + dE_t * uL + np.random.normal(0, sig_L)
                    av = a0 + dE_t * ua + np.random.normal(0, sig_a)
                    bv = b0 + dE_t * ub + np.random.normal(0, sig_b)
                row_idx = 4 + i_new
                ws.cell(row_idx, c    ).value = round(float(Lv), 2)
                ws.cell(row_idx, c + 1).value = round(float(av), 2)
                ws.cell(row_idx, c + 2).value = round(float(bv), 2)
                f = predicted_fill if sheet == "恒温恒湿" else (measured_fill if d <= 34 else extrap_fill)
                for off in (0, 1, 2): ws.cell(row_idx, c + off).fill = f

wb["室温"].cell(1, 1).value = "室温组 (23 C / 40%RH / lit) — raw -> Hampel-cleaned -> 90-day saturating projection"
wb["室温"].cell(1, 1).font = Font(bold=True)
wb["恒温"].cell(1, 1).value = "恒温组 (40 C / 10%RH / dark) — raw -> Hampel-cleaned -> 90-day saturating projection"
wb["恒温"].cell(1, 1).font = Font(bold=True)
wb["恒温恒湿"].cell(1, 1).value = "恒温恒湿组 (40 C / 80%RH / dark, PREDICTED) — RH-scaled from cleaned chamber, 90-day saturating projection"
wb["恒温恒湿"].cell(1, 1).font = Font(bold=True, color="B00020")

wb.save(dst)
print(f"\nSaved -> {dst}")

# ----------------------------------------------------------------------
# Stage 5: Plot
# ----------------------------------------------------------------------
print("\n=== Stage 5: Plot ===")
wb2 = openpyxl.load_workbook(dst, data_only=True)

def load_arm(sheet_name):
    ws = wb2[sheet_name]
    rows = [r for r in range(4, ws.max_row + 1) if is_date(ws.cell(r, 1).value)]
    times = np.array([date_to_day(ws.cell(r, 1).value) for r in rows])
    per_spot = {}
    for grp, cols in spec_cols.items():
        for sidx, c in enumerate(cols):
            baseline = (to_f(ws.cell(rows[0], c).value),
                        to_f(ws.cell(rows[0], c + 1).value),
                        to_f(ws.cell(rows[0], c + 2).value))
            dE_series = []
            for r in rows:
                L = to_f(ws.cell(r, c).value)
                a = to_f(ws.cell(r, c + 1).value)
                b = to_f(ws.cell(r, c + 2).value)
                L0, a0, b0 = baseline
                if None in (L, a, b, L0, a0, b0):
                    dE_series.append(np.nan)
                else:
                    dE_series.append(math.sqrt((L - L0) ** 2 + (a - a0) ** 2 + (b - b0) ** 2))
            per_spot[(grp, sidx)] = (times, np.array(dE_series))
    return per_spot

arms = {
    "23 C / 40%RH (room, lit)":         load_arm("室温"),
    "40 C / 10%RH (chamber, dark)":     load_arm("恒温"),
    "40 C / 80%RH (predicted, dark)":   load_arm("恒温恒湿"),
}
styles = {
    "23 C / 40%RH (room, lit)":       {"color": "#2E86AB", "marker": "o", "ls": "-"},
    "40 C / 10%RH (chamber, dark)":   {"color": "#E63946", "marker": "s", "ls": "-"},
    "40 C / 80%RH (predicted, dark)": {"color": "#6A4C93", "marker": "^", "ls": "--"},
}

fig, axes = plt.subplots(2, 3, figsize=(14, 7.5), sharex=True)
geom_groups = [
    ("block (10x10x4 cm tile)",   ["R",  "G",  "B"]),
    ("pedestal (1:5 lotus base)", ["TR", "TG", "TB"]),
]
for row, (geom, grps) in enumerate(geom_groups):
    row_max = 0.0
    for grp in grps:
        cols = spec_cols[grp]
        for arm_label, per_spot in arms.items():
            for sidx in range(len(cols)):
                _, dE = per_spot[(grp, sidx)]
                m = np.nanmax(dE)
                if not np.isnan(m): row_max = max(row_max, m)
    for col, grp in enumerate(grps):
        ax = axes[row, col]
        ax.axvline(34, color="gray", lw=0.8, ls=":", alpha=0.6)
        if row == 0 and col == 0:
            ax.text(34, row_max * 1.05, " measured | extrapolated", fontsize=7,
                    color="gray", ha="left", va="top", alpha=0.7)
        cols = spec_cols[grp]
        for arm_label, per_spot in arms.items():
            st = styles[arm_label]
            spot_curves = []
            for sidx in range(len(cols)):
                t, dE = per_spot[(grp, sidx)]
                spot_curves.append(dE)
                ax.plot(t, dE, color=st["color"], linestyle=st["ls"],
                        alpha=0.25, linewidth=0.7)
            stacked = np.array(spot_curves, dtype=float)
            mean = np.nanmean(stacked, axis=0)
            mn = np.nanmin(stacked, axis=0); mx = np.nanmax(stacked, axis=0)
            t = arms[arm_label][(grp, 0)][0]
            ax.fill_between(t, mn, mx, color=st["color"], alpha=0.12)
            ax.plot(t, mean, color=st["color"], linestyle=st["ls"],
                    marker=st["marker"], markersize=3.5, linewidth=1.6,
                    label=arm_label if (row == 0 and col == 0) else None)
        ax.set_title(f"{PIG[grp]} -- {geom}", fontsize=10)
        ax.grid(True, alpha=0.3)
        ax.set_xlim(-2, 93)
        ax.set_ylim(-0.2, row_max * 1.10 + 0.5)
        if col == 0: ax.set_ylabel(r"$\Delta E^*_{ab}$")
        if row == 1: ax.set_xlabel("Days since 17 Apr 2026")
        n_spots = len(cols)
        ax.text(0.97, 0.04, f"n = {n_spots} spots",
                transform=ax.transAxes, ha="right", va="bottom",
                fontsize=7, alpha=0.6, style="italic")

handles, labels = axes[0, 0].get_legend_handles_labels()
fig.legend(handles, labels, loc="upper center", ncol=3,
           frameon=False, fontsize=9.5, bbox_to_anchor=(0.5, 0.965))
fig.suptitle("Per-pigment dE* trajectories projected from raw pilot to 3 months (90 d, 3-day intervals)",
             y=0.99, fontsize=12, fontweight="bold")
plt.subplots_adjust(top=0.88, bottom=0.10, left=0.07, right=0.99, hspace=0.30, wspace=0.20)

out = "experiments/deltaE_trajectories_raw_90d.png"
plt.savefig(out, dpi=150, bbox_inches="tight")
print(f"Saved: {out}")
