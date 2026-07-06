"""Plot the 90-day extended trajectories from experiments/敦煌变化数据_90d.xlsx.

Same layout as _plot_predicted.py (2 rows x 3 cols: block over pedestal,
vermilion / malachite / azurite columns). Three arms overlaid per panel.
"""
import openpyxl, math, re
import numpy as np
import matplotlib.pyplot as plt

DATE_RE = re.compile(r"^\s*\d+\.\d+\s*$")

def to_f(v):
    try: return float(v)
    except Exception: return float("nan")

def is_date(v):
    if v is None: return False
    return bool(DATE_RE.match(str(v).strip()))

def date_to_day(d):
    s = str(d).strip(); m, day = s.split(".")
    dpm = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    return (sum(dpm[: int(m) - 1]) + int(day)) - (sum(dpm[:3]) + 17)

spec_cols = {
    "R":  [2, 5, 8],   "G":  [11, 14, 17], "B":  [20, 23, 26],
    "TR": [29, 32],    "TG": [35, 38],     "TB": [41, 44],
}
PIG = {"R": "Vermilion", "G": "Malachite", "B": "Azurite",
       "TR": "Vermilion", "TG": "Malachite", "TB": "Azurite"}

src = "experiments/敦煌变化数据_90d.xlsx"
wb = openpyxl.load_workbook(src, data_only=True)

def load_arm(sheet_name):
    ws = wb[sheet_name]
    rows = [r for r in range(4, ws.max_row + 1) if is_date(ws.cell(r, 1).value)]
    times = np.array([date_to_day(ws.cell(r, 1).value) for r in rows])
    per_spot = {}
    for grp, cols in spec_cols.items():
        for sidx, c in enumerate(cols):
            baseline = (to_f(ws.cell(rows[0], c).value),
                        to_f(ws.cell(rows[0], c + 1).value),
                        to_f(ws.cell(rows[0], c + 2).value))
            dE_series = []
            for r in rows:
                L = to_f(ws.cell(r, c).value)
                a = to_f(ws.cell(r, c + 1).value)
                b = to_f(ws.cell(r, c + 2).value)
                L0, a0, b0 = baseline
                if any(np.isnan([L, a, b, L0, a0, b0])):
                    dE_series.append(np.nan)
                else:
                    dE_series.append(math.sqrt((L - L0) ** 2 + (a - a0) ** 2 + (b - b0) ** 2))
            per_spot[(grp, sidx)] = (times, np.array(dE_series))
    return per_spot

arms = {
    "23 C / 40%RH (room, lit)  --  measured d 0-34 / saturating model d 36-90":   load_arm("室温"),
    "40 C / 10%RH (chamber, dark)  --  measured d 0-34 / saturating model d 36-90": load_arm("恒温"),
    "40 C / 80%RH (predicted, dark)  --  model output throughout (RH-scaled, never measured)": load_arm("恒温恒湿"),
}

styles = {
    "23 C / 40%RH (room, lit)  --  measured d 0-34 / saturating model d 36-90":   {"color": "#2E86AB", "marker": "o", "ls": "-"},
    "40 C / 10%RH (chamber, dark)  --  measured d 0-34 / saturating model d 36-90": {"color": "#E63946", "marker": "s", "ls": "-"},
    "40 C / 80%RH (predicted, dark)  --  model output throughout (RH-scaled, never measured)": {"color": "#6A4C93", "marker": "^", "ls": "--"},
}

fig, axes = plt.subplots(2, 3, figsize=(14, 7.5), sharex=True)
geom_groups = [
    ("block (10x10x4 cm tile)",   ["TR", "TG", "TB"]),
    ("pedestal (1:5 lotus base)", ["R",  "G",  "B"]),
]

for row, (geom, grps) in enumerate(geom_groups):
    row_max = 0.0
    for grp in grps:
        cols = spec_cols[grp]
        for arm_label, per_spot in arms.items():
            for sidx in range(len(cols)):
                t, dE = per_spot[(grp, sidx)]
                m = np.nanmax(dE)
                if not np.isnan(m): row_max = max(row_max, m)
    for col, grp in enumerate(grps):
        ax = axes[row, col]
        # vertical guide at day 34 (end of measured window) -- now on EVERY subplot
        ax.axvline(34, color="#999999", lw=1.0, ls=":", alpha=0.85)
        ax.text(34 - 1, row_max * 1.05, "measured", fontsize=7.5,
                color="#555555", ha="right", va="top", alpha=0.85, style="italic")
        ax.text(34 + 1, row_max * 1.05, "model extrapolation", fontsize=7.5,
                color="#555555", ha="left", va="top", alpha=0.85, style="italic")

        cols = spec_cols[grp]
        for arm_label, per_spot in arms.items():
            st = styles[arm_label]
            spot_curves = []
            for sidx in range(len(cols)):
                t, dE = per_spot[(grp, sidx)]
                spot_curves.append(dE)
                ax.plot(t, dE, color=st["color"], linestyle=st["ls"],
                        alpha=0.25, linewidth=0.7)
            stacked = np.array(spot_curves, dtype=float)
            mean = np.nanmean(stacked, axis=0)
            mn = np.nanmin(stacked, axis=0); mx = np.nanmax(stacked, axis=0)
            t = arms[arm_label][(grp, 0)][0]
            ax.fill_between(t, mn, mx, color=st["color"], alpha=0.12)
            ax.plot(t, mean, color=st["color"], linestyle=st["ls"],
                    marker=st["marker"], markersize=3.5, linewidth=1.6,
                    label=arm_label if (row == 0 and col == 0) else None)
        ax.set_title(f"{PIG[grp]} -- {geom}", fontsize=10)
        ax.grid(True, alpha=0.3)
        ax.set_xlim(-2, 93)
        ax.set_ylim(-0.2, row_max * 1.10 + 0.5)
        if col == 0: ax.set_ylabel(r"$\Delta E^*_{ab}$")
        if row == 1: ax.set_xlabel("Days since 17 Apr 2026")
        n_spots = len(cols)
        ax.text(0.97, 0.04, f"n = {n_spots} spots",
                transform=ax.transAxes, ha="right", va="bottom",
                fontsize=7, alpha=0.6, style="italic")

handles, labels = axes[0, 0].get_legend_handles_labels()
fig.legend(handles, labels, loc="upper center", ncol=1,
           frameon=False, fontsize=7.5, bbox_to_anchor=(0.5, 0.93))

fig.suptitle("Per-pigment $\\Delta E^*_{ab}$ trajectories: 34-day measured pilot + 56-day first-order saturating model projection",
             y=0.99, fontsize=11.5, fontweight="bold")
fig.text(0.5, 0.025,
         "Days 0-34: outlier-filtered colorimetry pilot data.  Days 36-90: first-order saturating model per spot with matched-amplitude noise.  "
         "40 C/80%RH arm is RH-scaled prediction from 10%RH chamber, never measured experimentally.",
         ha="center", va="bottom", fontsize=7.5, style="italic", color="#555555", wrap=True)
plt.subplots_adjust(top=0.83, bottom=0.13, left=0.07, right=0.99, hspace=0.30, wspace=0.20)

out = "experiments/deltaE_trajectories_90d.png"
plt.savefig(out, dpi=150, bbox_inches="tight")
print(f"Saved: {out}")
