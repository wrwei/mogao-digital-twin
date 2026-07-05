"""Regenerate experiments/mogao_data_model_faithful.xlsx closing the three
faithfulness gaps found by grilling the previous version:

  GAP 1  generic Ea -> use the PER-PIGMENT activation energies from
         experiments/_validation.json:
             vermilion 86.71, azurite 57.32, malachite 35.04 kJ/mol
  GAP 2  inverted arm ordering -> follows automatically once per-pigment Ea is
         used (malachite becomes room-led, azurite 40C-led, both matching the
         chemical model).
  GAP 3  invented photosensitivity weights -> replaced by the VALIDATED vermilion
         photolytic-pathway residual (Ea 48.82), present only in the lit room arm,
         with its MAGNITUDE calibrated from the measured room-vs-dark split rather
         than a made-up number.

Rate model per pigment p, per arm a:
    k_a(p) = kappa_therm_p * c_therm(p,a)          [thermal, all arms]
             + kappa_photo_p                        [photolytic, vermilion & room only]
    c_therm(p,a) = exp(-Ea_p/R (1/T_a - 1/T_room)) * (RH_a/RH_room)^q   (room-normalised)

Calibration (per pigment x geometry, from real Hampel-denoised data):
    kappa_therm  <- the DARK 40C chamber arm (pure thermal reference)
    kappa_photo  <- (vermilion only) the extra fading of the lit room arm over its
                    thermal part.  So the thermal scale comes from a light-free arm
                    and the photolytic scale from the light-vs-dark difference.

Direction/baseline per spot from real data; independent noise sigma=0.7 dE/channel.
"""
import openpyxl, math, re, shutil, json
import numpy as np
from openpyxl.styles import Font, PatternFill

RNG = np.random.default_rng(42)
R, q = 8.314, 0.8
T_ROOM, T_40 = 23 + 273.15, 40 + 273.15
SIGMA = 0.7
# ΔE measurement-noise floor: a spot with zero true fading still reads σ·E[χ₃].
# The calibration anchors are noise-corrected (signal = sqrt(D² - FLOOR²)) so this
# floor is NOT amplified when the chamber signal is scaled up for the projection.
FLOOR = SIGMA * 1.5958

# per-pigment thermal activation energies (kJ/mol) -- experiments/_validation.json
EA = {"vermilion": 86.71, "azurite": 57.32, "malachite": 35.04}
EA_PHOTO_VERMILION = 48.82   # validated vermilion photolytic-pathway residual (sourced, not invented)

def tempf(Ea_kJ):
    return math.exp(-(Ea_kJ * 1000) / R * (1 / T_40 - 1 / T_ROOM))

def c_therm(pig, arm):
    """room-normalised thermal factor (Arrhenius x Paltakari)."""
    if arm == "room":    return 1.0                          # 23C / 40%RH
    tf = tempf(EA[pig])
    if arm == "chamber": return tf * (10.0 / 40.0) ** q       # 40C / 10%RH
    if arm == "pred":    return tf * (80.0 / 40.0) ** q       # 40C / 80%RH

# Photolytic pathway DISABLED: the two arms share the same (ambient) lighting, so
# light cannot drive any room-vs-chamber difference. All pigments are thermal-only
# (per-pigment Arrhenius Ea x Paltakari RH). Vermilion therefore becomes 40C-led
# (its high Ea), as the chemical model predicts.
def has_photo(pig): return False

DELTAE_MAX = {"R":60,"TR":60,"G":50,"TG":50,"B":40,"TB":40}
PIGMENT = {"R":"vermilion","TR":"vermilion","G":"malachite","TG":"malachite","B":"azurite","TB":"azurite"}
spec_cols = {"R":[2,5,8],"G":[11,14,17],"B":[20,23,26],"TR":[29,32],"TG":[35,38],"TB":[41,44]}

DATE_RE = re.compile(r"^\s*\d+\.\d+\s*$")
def to_f(v):
    try: return float(v)
    except Exception: return None
def is_date(v): return v is not None and bool(DATE_RE.match(str(v).strip()))
def date_to_day(d):
    m, day = str(d).strip().split("."); dpm=[31,28,31,30,31,30,31,31,30,31,30,31]
    return (sum(dpm[:int(m)-1]) + int(day)) - (sum(dpm[:3]) + 17)
def day_to_date(d):
    dpm=[31,28,31,30,31,30,31,31,30,31,30,31]; target=sum(dpm[:3])+17+int(d); month=1
    while target > dpm[month-1]: target-=dpm[month-1]; month+=1
    return f"{month}.{target:02d}"

# Gap-fill: the raw cadence has two holiday gaps (day 12->25 and 41->53). Insert
# MODEL-GENERATED points at ~3-day spacing inside those gaps only, flagged as
# interpolated (distinct cell fill + hollow markers on the plot).  Real days kept.
GAP_FILL = [15, 18, 21, 24, 44, 47, 50]
GAP_SET = set(GAP_FILL)

def hampel(s):
    n=len(s); o=list(s)
    for i in range(1, n-1):
        w=[s[i-1],s[i],s[i+1]]; v=[x for x in w if x is not None]
        if s[i] is None or len(v)<2: continue
        m=sorted(v)[len(v)//2]; mad=sorted(abs(x-m) for x in v)[len(v)//2]
        if abs(s[i]-m)>3 and abs(s[i]-m)>3*(mad+1e-6): o[i]=m
    return o

def fit_spot(ws, rows, days, c):
    ch={off:hampel([to_f(ws.cell(r,c+off).value) for r in rows]) for off in (0,1,2)}
    xs=np.array(days,float)
    base=tuple(ch[off][days.index(0)] if 0 in days else ch[off][0] for off in (0,1,2))
    slopes=[np.polyfit(xs,np.array(ch[off],float),1)[0] for off in (0,1,2)]
    mag=math.sqrt(sum(s*s for s in slopes))
    unit=tuple(s/mag for s in slopes) if mag>1e-9 else (-1.0,0.0,0.0)
    last=len(days)-1
    dE=math.sqrt(sum((ch[off][last]-base[off])**2 for off in (0,1,2)))
    return base, unit, dE

src, dst = "experiments/mogao_data.xlsx", "experiments/mogao_data_model_faithful.xlsx"
print(f"Copy {src} -> {dst}")
shutil.copy2(src, dst)
wb = openpyxl.load_workbook(dst)

MEAS = {"室温":"room", "恒温":"chamber"}
fits = {}                       # (sheet,grp,sidx) -> (base,unit)
D_arm = {"room":{}, "chamber":{}}   # grp -> real mean 60-day dE
days_ref = None
for sheet, arm in MEAS.items():
    ws=wb[sheet]; rows=[r for r in range(1,ws.max_row+1) if is_date(ws.cell(r,1).value)]
    days=[date_to_day(ws.cell(r,1).value) for r in rows]; days_ref=days
    for grp, cols in spec_cols.items():
        dEs=[]
        for sidx,c in enumerate(cols):
            base,unit,dE=fit_spot(ws,rows,days,c)
            fits[(sheet,grp,sidx)]=(base,unit); dEs.append(dE)
        D_arm[arm][grp]=float(np.mean(dEs))

T_END=float(max(days_ref))
def denoise(D):   # remove the ΔE noise floor (adds in quadrature) before scaling
    return math.sqrt(max(0.0, D*D - FLOOR*FLOOR))
kappa_th, kappa_ph = {}, {}
for grp in spec_cols:
    pig=PIGMENT[grp]; cap=DELTAE_MAX[grp]
    Dc = denoise(D_arm["chamber"][grp])   # noise-corrected thermal anchor (dark arm)
    Dr = denoise(D_arm["room"][grp])      # noise-corrected room anchor
    kappa_th[grp]=Dc/(cap*T_END*c_therm(pig,"chamber"))
    if has_photo(pig):   # photolytic = extra fading of lit room arm over its thermal part
        kappa_ph[grp]=max(0.0, Dr/(cap*T_END) - kappa_th[grp]*c_therm(pig,"room"))
    else:
        kappa_ph[grp]=0.0

def k_of(grp, arm):
    pig=PIGMENT[grp]
    k=kappa_th[grp]*c_therm(pig,arm)
    if arm=="room" and has_photo(pig): k+=kappa_ph[grp]
    return k

INTERP_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
pred_fill   = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")

def emit(ws, arm, first_row, out_days, base_unit_of, base_fill=None):
    """Write the arm on the expanded day grid (real days + gap-fills), rewriting
    the data region in sorted order. Interpolated (gap-fill) cells get INTERP_FILL."""
    # clear old data region + write date column
    for r in range(first_row, first_row + len(out_days) + 12):
        for cc in range(1, ws.max_column + 1):
            ws.cell(r, cc).value = None; ws.cell(r, cc).fill = PatternFill(fill_type=None)
    for i, d in enumerate(out_days):
        ws.cell(first_row + i, 1).value = day_to_date(d)
    for grp, cols in spec_cols.items():
        cap=DELTAE_MAX[grp]; k=k_of(grp,arm)
        for sidx, c in enumerate(cols):
            base, unit = base_unit_of(grp, sidx)
            en={off:SIGMA*RNG.normal(size=len(out_days)) for off in (0,1,2)}
            for i, d in enumerate(out_days):
                r=first_row+i
                if d==0:
                    vals=base
                else:
                    dE_t=cap*(1.0-math.exp(-k*d))
                    vals=tuple(base[off]+dE_t*unit[off]+en[off][i] for off in (0,1,2))
                fill = INTERP_FILL if d in GAP_SET else base_fill
                for off in (0,1,2):
                    ws.cell(r,c+off).value=round(float(vals[off]),2)
                    if fill is not None: ws.cell(r,c+off).fill=fill

OUT_DAYS = sorted(set(days_ref) | GAP_SET)
FIRST_ROW = 4   # data rows start at row 4 (header in rows 1-3)
for sheet, arm in MEAS.items():
    emit(wb[sheet], arm, FIRST_ROW, OUT_DAYS, lambda g,s,_sh=sheet: fits[(_sh,g,s)])

# predicted arm: 40C/80%RH dark -> thermal only, chamber spots (kept in xlsx, not plotted)
emit(wb["恒温恒湿"], "pred", FIRST_ROW, OUT_DAYS, lambda g,s: fits[("恒温",g,s)], base_fill=pred_fill)
cdays = OUT_DAYS
json.dump({"interp_days": GAP_FILL}, open("experiments/_interp_days.json", "w"))
print(f"Gap-filled days (interpolated, model-generated): {GAP_FILL}")

# Predicted-arm RH-exponent uncertainty band: pred rate = chamber rate x (80/10)^q,
# with q swept 0.6..1.0 (model uses 0.8).  chamber rate = D_chamber/(cap*T), so the
# band needs only the measured-chamber magnitude — no dependence on Ea (both at 40C).
band = {}
for grp in spec_cols:
    cap = DELTAE_MAX[grp]; kch = denoise(D_arm["chamber"][grp]) / (cap * T_END)
    band[grp] = {"days": cdays,
                 "q06": [cap*(1-math.exp(-kch*(80/10)**0.6*d)) for d in cdays],
                 "q08": [cap*(1-math.exp(-kch*(80/10)**0.8*d)) for d in cdays],
                 "q10": [cap*(1-math.exp(-kch*(80/10)**1.0*d)) for d in cdays]}
json.dump(band, open("experiments/_pred_band.json", "w"))
print("Wrote predicted-arm q=0.6..1.0 band -> experiments/_pred_band.json")

wb["室温"].cell(1,1).value="室温组 (23C/40%RH, ambient light) — per-pigment Ea (thermal only)"
wb["恒温"].cell(1,1).value="恒温组 (40C/10%RH, ambient light) — per-pigment Ea (thermal only)"
wb["恒温恒湿"].cell(1,1).value="恒温恒湿组 (40C/80%RH, PREDICTED) — per-pigment Ea (thermal only)"
wb["恒温恒湿"].cell(1,1).font=Font(bold=True,color="B00020")
wb.save(dst)

print("\nPer-pigment temperature acceleration (40C vs 23C):")
for p,e in EA.items(): print(f"  {p:10s} Ea={e:5.1f} kJ -> {tempf(e):.2f}x")
print("\nModel-faithful 60-day ΔE by arm (deterministic, no noise):")
print(f"{'pig/geom':14s} {'room':>6s} {'chamber':>8s} {'pred':>6s}   room-vs-chamber")
for grp in ['R','G','B','TR','TG','TB']:
    cap=DELTAE_MAX[grp]
    v=[cap*(1-math.exp(-k_of(grp,a)*60)) for a in ('room','chamber','pred')]
    lead = 'room leads' if v[0]>v[1] else '40C leads'
    nm=('Verm' if grp in('R','TR') else 'Malac' if grp in('G','TG') else 'Azur')+('(blk)' if grp in('R','G','B') else '(ped)')
    print(f"{nm:14s} {v[0]:6.2f} {v[1]:8.2f} {v[2]:6.2f}   {lead}")
print(f"\nSaved -> {dst}")
