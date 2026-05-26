"""Validation figure: predicted (April-fit, extrapolated) vs observed (May)
dE*_{ab} for each spot, plus bootstrap CI bars per pigment.

Outputs experiments/validation_predicted_vs_observed.png.
"""
import openpyxl, math, re, json
import numpy as np
import matplotlib.pyplot as plt

DATE_RE = re.compile(r"^\s*\d+\.\d+\s*$")

def to_f(v):
    try: return float(v)
    except Exception: return float("nan")
def is_date(v):
    if v is None: return False
    return bool(DATE_RE.match(str(v).strip()))
def date_to_day(d):
    s = str(d).strip(); m, day = s.split(".")
    dpm = [31,28,31,30,31,30,31,31,30,31,30,31]
    return (sum(dpm[:int(m)-1]) + int(day)) - (sum(dpm[:3]) + 17)

spec_cols = {
    "R":  [2, 5, 8],   "G":  [11, 14, 17], "B":  [20, 23, 26],
    "TR": [29, 32],    "TG": [35, 38],     "TB": [41, 44],
}
SRC_OUT = {("室温","R",0,"5.12"), ("恒温","TR",0,"5.14")}

wb = openpyxl.load_workbook("experiments/敦煌变化数据(1).xlsx", data_only=True)

# Gather predicted vs observed dE* for each spot, train=April, test=May
def fit_lab(t, L, a, b):
    mask = ~np.isnan(L) & ~np.isnan(a) & ~np.isnan(b)
    if mask.sum() < 2: return None
    sL, iL = np.polyfit(t[mask], L[mask], 1)
    sa, ia = np.polyfit(t[mask], a[mask], 1)
    sb, ib = np.polyfit(t[mask], b[mask], 1)
    return (sL, iL, sa, ia, sb, ib)

points_obs = []
points_pred = []
sheets_used = []
groups_used = []

for sheet in ("室温", "恒温"):
    ws = wb[sheet]
    rows = [r for r in range(4, ws.max_row+1) if is_date(ws.cell(r,1).value)]
    times = np.array([date_to_day(ws.cell(r,1).value) for r in rows])
    date_strs = [str(ws.cell(r,1).value).strip() for r in rows]
    for grp, cols in spec_cols.items():
        for sidx, c in enumerate(cols):
            Ls, As, Bs = [], [], []
            for i, r in enumerate(rows):
                if (sheet, grp, sidx, date_strs[i]) in SRC_OUT:
                    Ls.append(np.nan); As.append(np.nan); Bs.append(np.nan)
                else:
                    Ls.append(to_f(ws.cell(r,c).value))
                    As.append(to_f(ws.cell(r,c+1).value))
                    Bs.append(to_f(ws.cell(r,c+2).value))
            L = np.array(Ls); A = np.array(As); B = np.array(Bs)
            train_mask = times <= 12
            test_mask  = times > 12
            valid_train = ~np.isnan(L[train_mask]) & ~np.isnan(A[train_mask]) & ~np.isnan(B[train_mask])
            if valid_train.sum() < 2: continue
            L0 = L[train_mask][valid_train][0]; a0 = A[train_mask][valid_train][0]; b0 = B[train_mask][valid_train][0]
            fit = fit_lab(times[train_mask], L[train_mask], A[train_mask], B[train_mask])
            if fit is None: continue
            sL, iL, sa, ia, sb, ib = fit
            valid_test = ~np.isnan(L[test_mask]) & ~np.isnan(A[test_mask]) & ~np.isnan(B[test_mask])
            t_test_v = times[test_mask][valid_test]
            for i, t in enumerate(t_test_v):
                Lo = L[test_mask][valid_test][i]
                ao = A[test_mask][valid_test][i]
                bo = B[test_mask][valid_test][i]
                obs = math.sqrt((Lo-L0)**2 + (ao-a0)**2 + (bo-b0)**2)
                Lp = sL*t + iL; ap = sa*t + ia; bp = sb*t + ib
                pred = math.sqrt((Lp-L0)**2 + (ap-a0)**2 + (bp-b0)**2)
                points_obs.append(obs); points_pred.append(pred)
                sheets_used.append(sheet); groups_used.append(grp)

obs = np.array(points_obs); pred = np.array(points_pred)
rmse = float(np.sqrt(np.mean((pred - obs)**2)))
mae = float(np.mean(np.abs(pred - obs)))
r2 = 1 - np.sum((obs - pred)**2) / np.sum((obs - np.mean(obs))**2)
print(f"RMSE = {rmse:.2f}, MAE = {mae:.2f}, R^2 = {r2:.2f}, n = {len(obs)}")

# Load bootstrap CIs
with open("experiments/_validation.json") as f:
    valdata = json.load(f)

# 2-panel plot: (a) predicted vs observed; (b) bootstrap CI bars per pigment
fig, axes = plt.subplots(1, 2, figsize=(13, 5.5))

# Panel (a): scatter
ax = axes[0]
# color by sheet
colors = {"室温": "#2E86AB", "恒温": "#E63946"}
labels = {"室温": "23 C / 40%RH (lit)", "恒温": "40 C / 10%RH (dark)"}
for sh in ("室温", "恒温"):
    m = [i for i, s in enumerate(sheets_used) if s == sh]
    ax.scatter(obs[m], pred[m], c=colors[sh], alpha=0.6, s=35, label=labels[sh], edgecolor="white", linewidth=0.5)
lim = max(obs.max(), pred.max(), 1) * 1.1
ax.plot([0, lim], [0, lim], "k--", alpha=0.4, lw=1, label="1:1 line")
ax.set_xlim(-0.5, lim); ax.set_ylim(-0.5, lim)
ax.set_xlabel(r"Observed $\Delta E^*_{ab}$ (May, day 25-34)")
ax.set_ylabel(r"Predicted $\Delta E^*_{ab}$ (linear extrapolation from April fit)")
ax.set_title(f"Time-window cross-validation\n(train April days 0-12, predict May days 25-34)\nRMSE = {rmse:.2f}, MAE = {mae:.2f}, $R^2$ = {r2:.2f}, n = {len(obs)}", fontsize=10)
ax.grid(True, alpha=0.3); ax.legend(loc="upper left", fontsize=9)

# Panel (b): bootstrap CI bars
ax = axes[1]
keys = ["vermilion_pedestal", "azurite_block", "malachite_block", "vermilion_block_photolytic"]
ylabels = ["Vermilion\npedestal\n(n=2)", "Azurite\nblock\n(n=3)", "Malachite\nblock\n(n=3)", "Vermilion block\n(photolytic\nresidual, n=3)"]
y = list(range(len(keys)))
medians = [valdata["bootstrap_Ea_95ci_kJmol"][k]["median"] for k in keys]
los = [valdata["bootstrap_Ea_95ci_kJmol"][k]["p2_5"] for k in keys]
his = [valdata["bootstrap_Ea_95ci_kJmol"][k]["p97_5"] for k in keys]
ax.errorbar(medians, y, xerr=[
    [m - lo for m, lo in zip(medians, los)],
    [hi - m for m, hi in zip(medians, his)]],
    fmt="o", color="#6A4C93", capsize=4, capthick=1.2, markersize=8, lw=1.5)
ax.set_yticks(y); ax.set_yticklabels(ylabels, fontsize=9)
ax.invert_yaxis()
ax.axvline(70, color="gray", lw=0.8, ls=":", alpha=0.6)
ax.text(70, -0.3, "Strlic ~70", fontsize=8, color="gray", ha="center", alpha=0.7)
ax.axvline(100, color="gray", lw=0.8, ls=":", alpha=0.6)
ax.text(100, -0.3, "Michalski ~100", fontsize=8, color="gray", ha="center", alpha=0.7)
ax.set_xlabel(r"Inferred $E_a$ (kJ/mol)")
ax.set_xlim(0, max(his) * 1.1 + 5)
ax.set_title("Bootstrap 95% CI on inferred $E_a$ at $q$ = 0.8\n(1000 resamples over spot index per arm, chamber RH = 10%)", fontsize=10)
ax.grid(True, alpha=0.3, axis="x")

fig.suptitle("Internal validation of the pilot Arrhenius extraction", y=1.0, fontsize=11.5, fontweight="bold")
plt.tight_layout()
out = "experiments/validation_predicted_vs_observed.png"
plt.savefig(out, dpi=150, bbox_inches="tight")
print(f"Saved: {out}")
