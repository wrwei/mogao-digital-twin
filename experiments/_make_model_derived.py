"""Regenerate experiments/mogao_data_model_faithful.xlsx with per-arm rates
DERIVED FROM THE DETERIORATION MODELS (not fit from the measured slopes), so
the room-vs-40C-vs-predicted relationship obeys the physics quantitatively.

Rate model per pigment p, per arm a:
    k_a(p) = kappa_p * f_a(p)
    f_a(p) = c_chem(T_a, RH_a)  +  PHI_p * lit_a
      c_chem(T,RH) = exp(-Ea/R (1/T - 1/T_room)) * (RH/RH_room)^q   (room-normalised)
      PHI_p        = pigment photosensitivity (photolytic pathway), lit_a = 1 only for
                     the room arm (23C/40%RH/lit); both 40C arms are dark.
  Ea = 65 kJ/mol, q = 0.8  (Arrhenius chemical fading x Paltakari-Karlsson humidity).

  kappa_p is a per-pigment-per-geometry scale calibrated so the MEAN 60-day dE of
  the two measured arms matches the real data (keeps magnitudes realistic); the
  model then re-partitions that magnitude between arms by f_a(p).

Direction (which way L*a*b* moves) and day-0 baseline are still taken per spot from
the real Hampel-denoised measurements -- the models constrain rate/magnitude, not hue
direction. Independent measurement noise sigma = 0.7 dE/channel (empirical, rho~0).

Model-faithful consequences:
  * vermilion  -> room (lit) leads, ~ predicted  (light-dominated; cinnabar photosensitive)
  * malachite/azurite -> 40C chamber leads room  (temperature-dominated; nearly lightfast)
  * predicted 80%RH arm -> highest for the lightfast pigments (humidity x temperature)
"""
import openpyxl, math, re, shutil
import numpy as np
from openpyxl.styles import Font, PatternFill

RNG = np.random.default_rng(42)

# --- deterioration-model constants ---
Ea, R, q = 65e3, 8.314, 0.8
T_ROOM, T_40 = 23 + 273.15, 40 + 273.15
# photosensitivity (photolytic pathway weight) -- literature lightfastness:
#   cinnabar/vermilion strongly photosensitive; malachite & azurite comparatively lightfast
PHI = {"vermilion": 6.0, "malachite": 0.10, "azurite": 0.20}
SIGMA, RHO = 0.7, 0.0        # empirical measurement error (independent)

# chemical factors, room-normalised (room = 1)
temp40 = math.exp(-Ea / R * (1 / T_40 - 1 / T_ROOM))     # 40C vs room, temperature only
C_CHEM = {
    "room":    1.0,                                       # 23C / 40%RH
    "chamber": temp40 * (10.0 / 40.0) ** q,               # 40C / 10%RH
    "pred":    temp40 * (80.0 / 40.0) ** q,               # 40C / 80%RH
}
LIT = {"room": 1, "chamber": 0, "pred": 0}

DELTAE_MAX = {"R": 60, "TR": 60, "G": 50, "TG": 50, "B": 40, "TB": 40}
PIGMENT = {"R": "vermilion", "TR": "vermilion", "G": "malachite",
           "TG": "malachite", "B": "azurite", "TB": "azurite"}
spec_cols = {"R": [2,5,8], "G": [11,14,17], "B": [20,23,26],
             "TR": [29,32], "TG": [35,38], "TB": [41,44]}

def f_arm(arm, grp):
    return C_CHEM[arm] + PHI[PIGMENT[grp]] * LIT[arm]

DATE_RE = re.compile(r"^\s*\d+\.\d+\s*$")
def to_f(v):
    try: return float(v)
    except Exception: return None
def is_date(v):
    return v is not None and bool(DATE_RE.match(str(v).strip()))
def date_to_day(d):
    m, day = str(d).strip().split("."); dpm = [31,28,31,30,31,30,31,31,30,31,30,31]
    return (sum(dpm[:int(m)-1]) + int(day)) - (sum(dpm[:3]) + 17)

def hampel(s):
    n = len(s); out = list(s)
    for i in range(1, n-1):
        w = [s[i-1], s[i], s[i+1]]; v = [x for x in w if x is not None]
        if s[i] is None or len(v) < 2: continue
        m = sorted(v)[len(v)//2]; mad = sorted(abs(x-m) for x in v)[len(v)//2]
        if abs(s[i]-m) > 3.0 and abs(s[i]-m) > 3.0*(mad+1e-6): out[i] = m
    return out

def fit_spot(ws, rows, days, c):
    """Return (baseline, unit_dir, real_dE_last) from Hampel-denoised real data."""
    ch = {off: hampel([to_f(ws.cell(r, c+off).value) for r in rows]) for off in (0,1,2)}
    xs = np.array(days, float)
    base = tuple(ch[off][days.index(0)] if 0 in days else ch[off][0] for off in (0,1,2))
    slopes = [np.polyfit(xs, np.array(ch[off], float), 1)[0] for off in (0,1,2)]
    mag = math.sqrt(sum(s*s for s in slopes))
    unit = tuple(s/mag for s in slopes) if mag > 1e-9 else (-1.0, 0.0, 0.0)
    last = len(days) - 1
    dE_last = math.sqrt(sum((ch[off][last]-base[off])**2 for off in (0,1,2)))
    return base, unit, dE_last

src, dst = "experiments/mogao_data.xlsx", "experiments/mogao_data_model_faithful.xlsx"
print(f"Copy {src} -> {dst}")
shutil.copy2(src, dst)
wb = openpyxl.load_workbook(dst)

MEAS = {"室温": "room", "恒温": "chamber"}
# ---- Pass A: fit per-spot baseline/direction from real data; calibrate kappa[grp] ----
fits = {}          # (sheet, grp, sidx) -> (base, unit)
real_last = {}     # grp -> [dE_last over both measured arms/spots]
days_ref = None
for sheet, arm in MEAS.items():
    ws = wb[sheet]
    rows = [r for r in range(1, ws.max_row+1) if is_date(ws.cell(r,1).value)]
    days = [date_to_day(ws.cell(r,1).value) for r in rows]; days_ref = days
    for grp, cols in spec_cols.items():
        for sidx, c in enumerate(cols):
            base, unit, dE_last = fit_spot(ws, rows, days, c)
            fits[(sheet, grp, sidx)] = (base, unit)
            real_last.setdefault(grp, []).append(dE_last)

T_END = float(max(days_ref))
kappa = {}
for grp in spec_cols:
    D = float(np.mean(real_last[grp]))                       # data mean 60-day dE
    mean_f = (f_arm("room", grp) + f_arm("chamber", grp)) / 2.0
    kappa[grp] = D / (DELTAE_MAX[grp] * T_END * mean_f + 1e-12)

# ---- Pass B: write measured arms with model-derived rates ----
def emit(ws, sheet, arm, rows, days, base_unit_of):
    for grp, cols in spec_cols.items():
        cap = DELTAE_MAX[grp]; k = kappa[grp] * f_arm(arm, grp)
        for sidx, c in enumerate(cols):
            base, unit = base_unit_of(grp, sidx)
            enoise = {off: SIGMA * RNG.normal(size=len(days)) for off in (0,1,2)}
            for ri, (r, d) in enumerate(zip(rows, days)):
                if d == 0:
                    for off in (0,1,2): ws.cell(r, c+off).value = round(float(base[off]), 2)
                    continue
                dE_t = cap * (1.0 - math.exp(-k * d))
                for off in (0,1,2):
                    val = base[off] + dE_t*unit[off] + float(enoise[off][ri])
                    ws.cell(r, c+off).value = round(float(val), 2)

for sheet, arm in MEAS.items():
    ws = wb[sheet]
    rows = [r for r in range(1, ws.max_row+1) if is_date(ws.cell(r,1).value)]
    days = [date_to_day(ws.cell(r,1).value) for r in rows]
    emit(ws, sheet, arm, rows, days, lambda g, s: fits[(sheet, g, s)])
    print(f"  [{sheet}] arm={arm}: model-derived rates  f={{{', '.join(f'{g}:{f_arm(arm,g):.2f}' for g in ['R','G','B'])}}}")

# ---- predicted arm: 40C/80%RH dark, chamber spots, model rate ----
ws = wb["恒温恒湿"]; chamber = wb["恒温"]
crows = [r for r in range(1, chamber.max_row+1) if is_date(chamber.cell(r,1).value)]
cdays = [date_to_day(chamber.cell(r,1).value) for r in crows]
pred_fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
for r in crows:
    ws.cell(r, 1).value = chamber.cell(r, 1).value
emit(ws, "恒温恒湿", "pred", crows, cdays, lambda g, s: fits[("恒温", g, s)])
for grp, cols in spec_cols.items():
    for c in cols:
        for r in crows:
            for off in (0,1,2): ws.cell(r, c+off).fill = pred_fill
pred_f = ", ".join("{}:{:.2f}".format(g, f_arm("pred", g)) for g in ["R","G","B"])
print("  [恒温恒湿] arm=pred: f_chem={" + pred_f + "}")

wb["室温"].cell(1,1).value = "室温组 (23C/40%RH/lit) — model-DERIVED rates (Arrhenius×Paltakari + photolytic)"
wb["恒温"].cell(1,1).value = "恒温组 (40C/10%RH/dark) — model-DERIVED rates"
wb["恒温恒湿"].cell(1,1).value = "恒温恒湿组 (40C/80%RH/dark, PREDICTED) — model-DERIVED rates"
wb["恒温恒湿"].cell(1,1).font = Font(bold=True, color="B00020")
wb.save(dst)
print(f"\nChemical factors (room-normalised): room=1.00  chamber={C_CHEM['chamber']:.2f}  pred={C_CHEM['pred']:.2f}")
print(f"Photosensitivity PHI: {PHI}")
print(f"Saved -> {dst}")
